"""
ORGANIZAR OFÍCIOS POR ÓRGÃO DEVEDOR
Cria estrutura de pastas e organiza os 15.977 PDFs
"""

import os
import shutil
import PyPDF2
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

class OrganizadorOficios:
    
    def __init__(self, pasta_origem):
        self.pasta_origem = pasta_origem
        self.pasta_organizada = "oficios_organizados"
        self.estatisticas = defaultdict(lambda: {'quantidade': 0, 'arquivos': []})
        
        if not os.path.exists(self.pasta_organizada):
            os.makedirs(self.pasta_organizada)
    
    def extrair_orgao_devedor(self, pdf_path):
        """Extrai o órgão devedor do PDF"""
        
        try:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                
                # Pegar primeira página
                texto = reader.pages[0].extract_text()
                
                # Procurar padrões comuns de órgão devedor
                patterns = [
                    r'Devedor[a]?:\s*([^\n]+)',
                    r'Órgão Devedor:\s*([^\n]+)',
                    r'Entidade Devedora:\s*([^\n]+)',
                    r'Estado de São Paulo',
                    r'Município de [A-Z][^\n]+',
                    r'Prefeitura Municipal de [A-Z][^\n]+',
                    r'FAZENDA.*ESTADO',
                    r'FAZENDA.*MUNICÍPIO',
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, texto, re.IGNORECASE)
                    if match:
                        orgao = match.group(1) if match.lastindex else match.group(0)
                        orgao = orgao.strip()
                        
                        # Limpar e padronizar
                        orgao = self.padronizar_nome_orgao(orgao)
                        
                        if orgao and len(orgao) > 3:
                            return orgao
                
                # Se não encontrou, procurar "Estado de São Paulo" ou similares
                if 'estado de são paulo' in texto.lower():
                    return 'Estado de São Paulo'
                elif 'fazenda' in texto.lower() and 'estado' in texto.lower():
                    return 'Fazenda do Estado de SP'
                
                return 'Não Identificado'
                
        except Exception as e:
            return 'Erro na Leitura'
    
    def padronizar_nome_orgao(self, orgao):
        """Padroniza nome do órgão"""
        
        # Remover caracteres especiais para nome de pasta
        orgao = orgao.strip()
        
        # Padronizações comuns
        if 'estado' in orgao.lower() and 'são paulo' in orgao.lower():
            return 'Estado de São Paulo'
        elif 'fazenda' in orgao.lower() and 'estado' in orgao.lower():
            return 'Fazenda do Estado de SP'
        elif 'município' in orgao.lower() or 'prefeitura' in orgao.lower():
            # Extrair nome do município
            match = re.search(r'(município|prefeitura)\s+(?:municipal\s+)?(?:de\s+)?([a-zà-ú\s]+)', orgao, re.IGNORECASE)
            if match:
                municipio = match.group(2).strip().title()
                return f'Município de {municipio}'
        
        # Capitalizar primeira letra de cada palavra
        orgao = ' '.join(word.capitalize() for word in orgao.split())
        
        return orgao
    
    def criar_nome_pasta_seguro(self, nome):
        """Cria nome de pasta válido"""
        
        # Remover caracteres inválidos
        nome = re.sub(r'[<>:"/\|?*]', '', nome)
        nome = nome.strip()
        
        # Limitar tamanho
        if len(nome) > 100:
            nome = nome[:100]
        
        return nome
    
    def organizar_oficios(self):
        """Organiza todos os ofícios"""
        
        print(f"\n📂 Organizando ofícios...")
        print(f"   Origem: {self.pasta_origem}")
        print(f"   Destino: {self.pasta_organizada}")
        
        # Listar todos os PDFs
        pdfs = [f for f in os.listdir(self.pasta_origem) if f.endswith('.pdf')]
        total = len(pdfs)
        
        print(f"\n   📄 Total de PDFs: {total}")
        
        print(f"\n🔍 Processando...")
        
        for idx, pdf in enumerate(pdfs, 1):
            if idx % 100 == 0 or idx == 1:
                print(f"   📋 {idx}/{total} ({idx/total*100:.1f}%)", end="\r")
            
            try:
                pdf_path = os.path.join(self.pasta_origem, pdf)
                
                # Extrair órgão
                orgao = self.extrair_orgao_devedor(pdf_path)
                
                # Criar pasta do órgão
                nome_pasta = self.criar_nome_pasta_seguro(orgao)
                pasta_orgao = os.path.join(self.pasta_organizada, nome_pasta)
                
                if not os.path.exists(pasta_orgao):
                    os.makedirs(pasta_orgao)
                
                # Copiar arquivo
                destino = os.path.join(pasta_orgao, pdf)
                shutil.copy2(pdf_path, destino)
                
                # Estatísticas
                self.estatisticas[orgao]['quantidade'] += 1
                self.estatisticas[orgao]['arquivos'].append(pdf)
                
            except Exception as e:
                print(f"\n   ❌ Erro em {pdf}: {str(e)[:50]}")
                continue
        
        print(f"\n   ✅ {total} PDFs processados!")
        
        return True
    
    def gerar_relatorio(self):
        """Gera relatório de organização"""
        
        print(f"\n📊 Gerando relatório...")
        
        # Ordenar por quantidade
        orgaos_ordenados = sorted(
            self.estatisticas.items(),
            key=lambda x: x[1]['quantidade'],
            reverse=True
        )
        
        print(f"\n" + "="*70)
        print(f"📊 ESTATÍSTICAS POR ÓRGÃO")
        print("="*70)
        
        print(f"\n   {'Órgão Devedor':<40} {'Quantidade':>10}")
        print(f"   {'-'*40} {'-'*10}")
        
        for orgao, dados in orgaos_ordenados:
            print(f"   {orgao:<40} {dados['quantidade']:>10}")
        
        print(f"\n   {'TOTAL':<40} {sum(d['quantidade'] for d in self.estatisticas.values()):>10}")
        
        # Criar planilha Excel
        self.criar_planilha_relatorio(orgaos_ordenados)
    
    def criar_planilha_relatorio(self, orgaos_ordenados):
        """Cria planilha Excel com relatório"""
        
        print(f"\n📊 Criando planilha de relatório...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório por Órgão"
        
        # Cabeçalhos
        ws['A1'] = 'Órgão Devedor'
        ws['B1'] = 'Quantidade de Ofícios'
        ws['C1'] = 'Percentual'
        
        # Estilo
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in ['A1', 'B1', 'C1']:
            ws[col].fill = header_fill
            ws[col].font = header_font
            ws[col].alignment = header_alignment
        
        # Dados
        total_geral = sum(d['quantidade'] for _, d in orgaos_ordenados)
        
        for row_idx, (orgao, dados) in enumerate(orgaos_ordenados, 2):
            ws.cell(row=row_idx, column=1, value=orgao)
            ws.cell(row=row_idx, column=2, value=dados['quantidade'])
            
            percentual = (dados['quantidade'] / total_geral) * 100
            ws.cell(row=row_idx, column=3, value=f"{percentual:.1f}%")
        
        # Total
        row_total = len(orgaos_ordenados) + 2
        ws.cell(row=row_total, column=1, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row_total, column=2, value=total_geral).font = Font(bold=True)
        ws.cell(row=row_total, column=3, value='100%').font = Font(bold=True)
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        ws.row_dimensions[1].height = 25
        
        # Filtros
        ws.auto_filter.ref = f'A1:C{len(orgaos_ordenados)+1}'
        
        # Salvar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo = f"relatorio_oficios_por_orgao_{timestamp}.xlsx"
        
        wb.save(arquivo)
        
        print(f"   ✅ Planilha salva: {arquivo}")

if __name__ == "__main__":
    print("="*70)
    print("📂 ORGANIZAR OFÍCIOS POR ÓRGÃO DEVEDOR")
    print("="*70)
    
    pasta_origem = "oficios_requisitorios_tjsp"
    
    if not os.path.exists(pasta_origem):
        print(f"\n❌ Pasta não encontrada: {pasta_origem}")
        input("\nENTER para sair...")
        exit()
    
    # Confirmar
    pdfs = [f for f in os.listdir(pasta_origem) if f.endswith('.pdf')]
    print(f"\n📊 Encontrados: {len(pdfs)} PDFs")
    print(f"📁 Origem: {pasta_origem}")
    print(f"📁 Destino: oficios_organizados/")
    
    confirma = input(f"\nOrganizar {len(pdfs)} PDFs por órgão? (s/n): ").lower()
    
    if confirma != 's':
        print("\n❌ Cancelado")
        exit()
    
    try:
        inicio = datetime.now()
        
        organizador = OrganizadorOficios(pasta_origem)
        
        if organizador.organizar_oficios():
            organizador.gerar_relatorio()
            
            fim = datetime.now()
            duracao = fim - inicio
            
            print("\n" + "="*70)
            print("🎉 ORGANIZAÇÃO CONCLUÍDA!")
            print("="*70)
            print(f"\n⏱️  Tempo: {int(duracao.total_seconds()/60)} minutos")
            print(f"📂 Pasta: oficios_organizados/")
            print(f"📊 Órgãos: {len(organizador.estatisticas)}")
        
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nENTER para fechar...\n")
    
    print("\n✅ FIM!")
