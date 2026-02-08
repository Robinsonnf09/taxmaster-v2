"""
Gerador de Planilha Modelo para Busca de Ofícios
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def criar_planilha_modelo():
    """Cria planilha Excel modelo para busca de ofícios"""
    
    print("📊 Criando planilha modelo...")
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos para Busca"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos (colunas obrigatórias)
    headers = ["numero_processo", "tribunal"]
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Adicionar linha de instruções
    ws.merge_cells('A1:B1')
    instruction = ws['A1']
    instruction.value = "⚠️ INSTRUÇÕES: Preencha com números de processos REAIS e PÚBLICOS do tribunal correspondente"
    instruction.font = Font(bold=True, color="FF0000", size=10)
    instruction.alignment = Alignment(horizontal='center', vertical='center')
    instruction.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Dados de exemplo (PROCESSOS FICTÍCIOS - SUBSTITUIR POR REAIS)
    dados_exemplo = [
        ["0000001-00.2020.4.01.3800", "TRF1"],
        ["0000002-00.2021.4.02.5101", "TRF2"],
        ["0000003-00.2022.4.03.6100", "TRF3"],
        ["0000004-00.2019.4.04.7100", "TRF4"],
        ["0000005-00.2023.4.05.8300", "TRF5"],
    ]
    
    # Adicionar dados
    for row_idx, row_data in enumerate(dados_exemplo, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Marcar como exemplo
            if row_idx <= 7:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Ajustar largura
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    
    # Adicionar nota no final
    nota_row = len(dados_exemplo) + 4
    ws.merge_cells(f'A{nota_row}:B{nota_row}')
    nota = ws[f'A{nota_row}']
    nota.value = "💡 DICA: Apague os exemplos e adicione seus processos reais abaixo"
    nota.font = Font(italic=True, color="0066CC")
    nota.alignment = Alignment(horizontal='center')
    
    # Salvar
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'modelo_busca_oficios_{timestamp}.xlsx'
    wb.save(filename)
    
    print(f"✅ Planilha criada: {filename}")
    print(f"\n📋 Localização: {os.path.abspath(filename)}")
    print(f"\n🎯 PRÓXIMO PASSO:")
    print(f"   1. Abra o arquivo: {filename}")
    print(f"   2. Substitua os processos fictícios por REAIS")
    print(f"   3. Use na interface: http://localhost:8080/automacao/busca-oficios")
    
    return filename

if __name__ == "__main__":
    print("="*70)
    print("📊 GERADOR DE PLANILHA MODELO PARA BUSCA DE OFÍCIOS")
    print("="*70)
    
    arquivo = criar_planilha_modelo()
    
    print("\n" + "="*70)
    print("✅ CONCLUÍDO!")
    print("="*70)
