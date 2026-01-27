"""
BUSCAR 500 PROCESSOS TJSP - PERÍODO 11/2024 a 12/2025
Busca processos de requisitórios/precatórios no período
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re

class BuscadorProcessosPeriodo:
    
    def __init__(self):
        self.driver = None
        self.processos = []
    
    def iniciar(self):
        print("\n🌐 Iniciando Chrome...")
        
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-popup-blocking')
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
        print("✅ Chrome iniciado!")
    
    def fazer_login(self):
        print(f"\n🔐 Login...")
        self.driver.get("https://esaj.tjsp.jus.br")
        time.sleep(3)
        
        print("\n" + "="*70)
        print("⚠️  FAÇA LOGIN:")
        print("="*70)
        print("   ❌ Se aparecer 'WEB SIGNER' → CANCELAR")
        print("   ✅ Use LOGIN e SENHA")
        print("="*70)
        
        input("\n>>> ENTER após login <<<\n")
        
        print("✅ Login efetuado!")
        return True
    
    def buscar_processos_periodo(self, data_inicio, data_fim, limite=500):
        """Busca processos no período especificado"""
        
        print(f"\n🔍 Buscando processos...")
        print(f"   📅 Período: {data_inicio} a {data_fim}")
        print(f"   📊 Limite: {limite} processos")
        
        # Acessar consulta de requisitórios
        url = "https://esaj.tjsp.jus.br/cpopg/open.do"
        self.driver.get(url)
        time.sleep(2)
        
        print("\n" + "="*70)
        print("⚠️  INSTRUÇÕES MANUAIS:")
        print("="*70)
        print("   1. Na página que abriu, vá em 'Consulta de Requisitórios'")
        print("   2. Selecione filtros:")
        print(f"      - Data inicial: {data_inicio}")
        print(f"      - Data final: {data_fim}")
        print("      - Natureza: ALIMENTARES (se disponível)")
        print("   3. Clique em CONSULTAR")
        print("   4. Aguarde carregar os resultados")
        print("="*70)
        
        input("\n>>> ENTER após realizar a busca <<<\n")
        
        print(f"\n📋 Extraindo processos da página...")
        
        processos_encontrados = []
        pagina = 1
        
        while len(processos_encontrados) < limite:
            print(f"\n   📄 Página {pagina}...", end=" ", flush=True)
            
            try:
                # Script para extrair processos da tabela
                script = """
                let processos = [];
                
                // Procurar tabelas
                document.querySelectorAll('table').forEach(table => {
                    table.querySelectorAll('tr').forEach(tr => {
                        let texto = tr.textContent;
                        
                        // Procurar números de processo no padrão TJSP
                        let matches = texto.match(/\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}/g);
                        
                        if (matches) {
                            matches.forEach(proc => {
                                if (!processos.includes(proc)) {
                                    processos.push(proc);
                                }
                            });
                        }
                    });
                });
                
                // Procurar também em links
                document.querySelectorAll('a').forEach(a => {
                    let texto = a.textContent;
                    let matches = texto.match(/\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}/g);
                    
                    if (matches) {
                        matches.forEach(proc => {
                            if (!processos.includes(proc)) {
                                processos.push(proc);
                            }
                        });
                    }
                });
                
                return processos;
                """
                
                processos_pagina = self.driver.execute_script(script)
                
                if processos_pagina:
                    for proc in processos_pagina:
                        if proc not in processos_encontrados:
                            processos_encontrados.append(proc)
                            if len(processos_encontrados) >= limite:
                                break
                    
                    print(f"✅ {len(processos_pagina)} novos (total: {len(processos_encontrados)})")
                else:
                    print(f"⚠️  Nenhum processo encontrado")
                
                # Verificar se tem próxima página
                if len(processos_encontrados) >= limite:
                    break
                
                # Tentar ir para próxima página
                try:
                    proxima = self.driver.find_element(By.LINK_TEXT, "Próxima")
                    proxima.click()
                    time.sleep(2)
                    pagina += 1
                except:
                    print(f"\n   ℹ️  Não há mais páginas")
                    break
                
            except Exception as e:
                print(f"❌ Erro: {str(e)[:50]}")
                break
        
        self.processos = processos_encontrados[:limite]
        
        print(f"\n" + "="*70)
        print(f"✅ TOTAL ENCONTRADO: {len(self.processos)} processos")
        print("="*70)
        
        return self.processos
    
    def criar_planilha(self, arquivo_saida):
        """Cria planilha com os processos"""
        
        print(f"\n📊 Criando planilha Excel...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Processos TJSP"
        
        # Cabeçalho
        ws['A1'] = 'Nº Processo'
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = header_alignment
        
        # Dados
        for row_idx, processo in enumerate(self.processos, 2):
            ws.cell(row=row_idx, column=1, value=processo)
        
        ws.column_dimensions['A'].width = 35
        ws.row_dimensions[1].height = 25
        ws.auto_filter.ref = f'A1:A{len(self.processos)+1}'
        
        wb.save(arquivo_saida)
        
        print(f"   ✅ Planilha salva: {arquivo_saida}")
        print(f"   📊 Total: {len(self.processos)} processos")
    
    def fechar(self):
        if self.driver:
            self.driver.quit()

if __name__ == "__main__":
    print("="*70)
    print("🔍 BUSCAR PROCESSOS TJSP - PERÍODO 11/2024 a 12/2025")
    print("="*70)
    
    buscador = BuscadorProcessosPeriodo()
    
    try:
        input("\nENTER para começar...\n")
        
        buscador.iniciar()
        
        if buscador.fazer_login():
            # Buscar processos
            processos = buscador.buscar_processos_periodo(
                data_inicio="01/11/2024",
                data_fim="31/12/2025",
                limite=500
            )
            
            if len(processos) > 0:
                # Preview
                print(f"\n📋 PREVIEW (primeiros 10):")
                for i, proc in enumerate(processos[:10], 1):
                    print(f"   {i:3}. {proc}")
                
                if len(processos) > 10:
                    print(f"\n   ... e mais {len(processos) - 10}")
                
                # Criar planilha
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                arquivo = f"processos_tjsp_periodo_{timestamp}.xlsx"
                
                buscador.criar_planilha(arquivo)
                
                print("\n" + "="*70)
                print("🎉 CONCLUÍDO!")
                print("="*70)
                print(f"\n📁 Arquivo: {arquivo}")
                print(f"📊 Processos: {len(processos)}")
                print(f"📅 Período: 11/2024 a 12/2025")
            else:
                print("\n⚠️  Nenhum processo encontrado no período!")
        
        input("\n\nENTER para fechar...\n")
        
    except KeyboardInterrupt:
        print("\n⚠️  Interrompido")
    
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        buscador.fechar()
    
    print("\n✅ FIM!")
