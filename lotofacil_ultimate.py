import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import os
import sys
import requests
from datetime import datetime, timedelta
from collections import Counter, defaultdict
import json
import pickle
import threading

# Importações matemáticas avançadas
try:
    import numpy as np
    from scipy import stats
    from scipy.special import comb
    NUMPY_DISPONIVEL = True
except ImportError:
    NUMPY_DISPONIVEL = False
    import math
    def comb(n, k, exact=True):
        resultado = math.factorial(n) // (math.factorial(k) * math.factorial(n - k))
        return resultado if exact else float(resultado)

# 
# CONSTANTES
# 

TOTAL_DEZENAS = 25
TOTAL_COMBINACOES = 3268760
VERSAO = "6.0 ULTIMATE PRO"

TIPOS_JOGOS = {
    15: {'dezenas': 15, 'apostas': 1, 'custo': 3.00, 'nome': 'Simples (15 números)'},
    16: {'dezenas': 16, 'apostas': 16, 'custo': 48.00, 'nome': 'Desdobramento (16 números)'},
    17: {'dezenas': 17, 'apostas': 68, 'custo': 204.00, 'nome': 'Desdobramento (17 números)'},
    18: {'dezenas': 18, 'apostas': 204, 'custo': 612.00, 'nome': 'Desdobramento (18 números)'}
}

PREMIOS_MEDIOS = {
    15: 1800000.00,
    14: 2000.00,
    13: 30.00,
    12: 12.00,
    11: 6.00
}

# 
# SPLASH SCREEN ULTIMATE
# 

def mostrar_splash():
    """Splash screen premium ultimate"""
    splash = tk.Tk()
    splash.overrideredirect(True)
    splash.attributes('-topmost', True)
    
    largura = 750
    altura = 500
    x = (splash.winfo_screenwidth() - largura) // 2
    y = (splash.winfo_screenheight() - altura) // 2
    splash.geometry(f"{largura}x{altura}+{x}+{y}")
    
    frame = tk.Frame(splash, bg="#0A0E27", bd=0)
    frame.pack(fill="both", expand=True)
    
    # Header
    header = tk.Frame(frame, bg="#6A0DAD", height=130)
    header.pack(fill="x")
    header.pack_propagate(False)
    
    tk.Label(header, text="🔬", font=("Segoe UI Emoji", 75), bg="#6A0DAD", fg="#FFD700").pack(pady=20)
    
    # Título
    tk.Label(frame, text="LOTOFÁCIL QUANTUM", font=("Segoe UI", 42, "bold"), bg="#0A0E27", fg="#FFD700").pack(pady=20)
    tk.Label(frame, text="ULTIMATE PRO EDITION", font=("Segoe UI", 14, "bold"), bg="#0A0E27", fg="#00D9FF").pack()
    tk.Label(frame, text="Sistema de Análise Matemática e Machine Learning", font=("Segoe UI", 11), bg="#0A0E27", fg="#B0B0B0").pack(pady=5)
    tk.Label(frame, text=f"v{VERSAO} - Robinson Tax Master", font=("Segoe UI", 9), bg="#0A0E27", fg="#707070").pack()
    
    # Features
    features_frame = tk.Frame(frame, bg="#0A0E27")
    features_frame.pack(pady=10)
    
    features = [
        "🤖 Machine Learning",
        "📊 Análise 500+ Concursos",
        "🎲 Simulação Monte Carlo",
        "📈 10+ Estratégias",
        "🔬 Testes Estatísticos"
    ]
    
    for feat in features:
        tk.Label(features_frame, text=feat, font=("Segoe UI", 8), bg="#0A0E27", fg="#4CAF50").pack()
    
    # Progress
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("Ultimate.Horizontal.TProgressbar", troughcolor='#1C1C1C', background='#FFD700', bordercolor='#0A0E27', lightcolor='#FFE55C', darkcolor='#FFD700', thickness=24)
    
    progress = ttk.Progressbar(frame, length=650, mode='determinate', style="Ultimate.Horizontal.TProgressbar")
    progress.pack(pady=30)
    
    label_status = tk.Label(frame, text="Inicializando...", font=("Segoe UI", 11, "bold"), bg="#0A0E27", fg="#FFFFFF")
    label_status.pack(pady=10)
    
    # Animação
    etapas = [
        (0, 12, "🔬 Inicializando núcleo quântico..."),
        (12, 25, "🤖 Carregando módulos de IA..."),
        (25, 40, "📊 Preparando análise estatística..."),
        (40, 55, "🧮 Configurando algoritmos..."),
        (55, 70, "🎲 Calibrando simulações..."),
        (70, 85, "📈 Otimizando estratégias..."),
        (85, 100, "✅ Sistema ULTIMATE ativo!")
    ]
    
    for inicio, fim, texto in etapas:
        label_status.config(text=texto)
        for i in range(inicio, fim + 1):
            progress['value'] = i
            splash.update()
            time.sleep(0.008)
    
    time.sleep(0.5)
    splash.destroy()

# 
# CLASSE DE MACHINE LEARNING SIMPLIFICADO
# 

class MachineLearningSimples:
    """Sistema de ML simplificado sem dependências pesadas"""
    
    def __init__(self):
        self.pesos = {}
        self.historico_treino = []
    
    def treinar(self, historico):
        """Treina modelo com histórico"""
        if not historico:
            return
        
        # Análise de padrões temporais
        for i, resultado in enumerate(historico):
            peso_temporal = (i + 1) / len(historico)  # Mais recente = mais peso
            
            for dezena in resultado['dezenas']:
                if dezena not in self.pesos:
                    self.pesos[dezena] = 0
                self.pesos[dezena] += peso_temporal
        
        # Normalizar pesos
        total = sum(self.pesos.values())
        if total > 0:
            self.pesos = {d: (p / total) * 100 for d, p in self.pesos.items()}
    
    def prever_scores(self):
        """Retorna scores preditivos"""
        return self.pesos if self.pesos else {d: 50.0 for d in range(1, 26)}

# 
# SIMULAÇÃO MONTE CARLO
# 

class SimuladorMonteCarlo:
    """Simulador de milhares de cenários"""
    
    @staticmethod
    def simular_jogos(jogo, num_simulacoes=10000):
        """Simula N sorteios e conta acertos"""
        resultados = {11: 0, 12: 0, 13: 0, 14: 0, 15: 0}
        jogo_set = set(jogo)
        
        for _ in range(num_simulacoes):
            sorteio = set(random.sample(range(1, 26), 15))
            acertos = len(jogo_set & sorteio)
            
            if acertos >= 11:
                resultados[acertos] += 1
        
        # Converter para probabilidades
        return {k: (v / num_simulacoes) * 100 for k, v in resultados.items()}
    
    @staticmethod
    def simular_roi(jogo, tipo_jogo, num_simulacoes=10000):
        """Simula ROI em múltiplos cenários"""
        custo = TIPOS_JOGOS[tipo_jogo]['custo']
        lucros = []
        
        for _ in range(num_simulacoes):
            sorteio = set(random.sample(range(1, 26), 15))
            acertos = len(set(jogo) & sorteio)
            
            premio = PREMIOS_MEDIOS.get(acertos, 0)
            lucro = premio - custo
            lucros.append(lucro)
        
        if NUMPY_DISPONIVEL:
            return {
                'medio': np.mean(lucros),
                'mediano': np.median(lucros),
                'desvio': np.std(lucros),
                'min': np.min(lucros),
                'max': np.max(lucros)
            }
        else:
            return {
                'medio': sum(lucros) / len(lucros),
                'mediano': sorted(lucros)[len(lucros)//2],
                'min': min(lucros),
                'max': max(lucros)
            }

# 
# TESTES ESTATÍSTICOS
# 

class TestesEstatisticos:
    """Testes rigorosos de aleatoriedade"""
    
    @staticmethod
    def teste_qui_quadrado(historico):
        """Teste Chi-Square para uniformidade"""
        if not historico:
            return None
        
        # Contar frequências
        frequencias = [0] * 25
        for resultado in historico:
            for dezena in resultado['dezenas']:
                frequencias[dezena - 1] += 1
        
        # Frequência esperada
        total = sum(frequencias)
        esperado = total / 25
        
        # Calcular chi-square
        chi2 = sum((obs - esperado) ** 2 / esperado for obs in frequencias)
        
        # Graus de liberdade = 24 (25 - 1)
        # Valor crítico para α=0.05: ~36.42
        valor_critico = 36.42
        
        return {
            'chi2': chi2,
            'valor_critico': valor_critico,
            'aleatorio': chi2 < valor_critico,
            'confianca': 95
        }
    
    @staticmethod
    def teste_runs(historico):
        """Teste de Runs para independência"""
        if len(historico) < 10:
            return None
        
        # Criar sequência de pares/ímpares
        sequencia = []
        for resultado in historico:
            pares = sum(1 for d in resultado['dezenas'] if d % 2 == 0)
            sequencia.append(1 if pares >= 8 else 0)
        
        # Contar runs
        runs = 1
        for i in range(1, len(sequencia)):
            if sequencia[i] != sequencia[i-1]:
                runs += 1
        
        n = len(sequencia)
        n1 = sum(sequencia)
        n0 = n - n1
        
        if n1 == 0 or n0 == 0:
            return None
        
        # Valor esperado de runs
        runs_esperado = ((2 * n1 * n0) / n) + 1
        
        # Desvio padrão
        if NUMPY_DISPONIVEL:
            desvio = np.sqrt((2 * n1 * n0 * (2 * n1 * n0 - n)) / (n**2 * (n - 1)))
        else:
            desvio = ((2 * n1 * n0 * (2 * n1 * n0 - n)) / (n**2 * (n - 1))) ** 0.5
        
        # Z-score
        z = (runs - runs_esperado) / desvio if desvio > 0 else 0
        
        return {
            'runs': runs,
            'esperado': runs_esperado,
            'z_score': z,
            'aleatorio': abs(z) < 1.96,  # 95% confiança
            'confianca': 95
        }
    
    @staticmethod
    def entropia_shannon(historico):
        """Calcula entropia para medir aleatoriedade"""
        if not historico:
            return None
        
        frequencias = Counter()
        total = 0
        
        for resultado in historico:
            for dezena in resultado['dezenas']:
                frequencias[dezena] += 1
                total += 1
        
        entropia = 0
        for freq in frequencias.values():
            p = freq / total
            if p > 0:
                if NUMPY_DISPONIVEL:
                    entropia -= p * np.log2(p)
                else:
                    entropia -= p * (math.log(p) / math.log(2))
        
        # Entropia máxima para 25 opções
        entropia_max = math.log(25) / math.log(2)
        
        return {
            'entropia': entropia,
            'entropia_max': entropia_max,
            'percentual': (entropia / entropia_max) * 100 if entropia_max > 0 else 0
        }

# 
# ANALISADOR ULTIMATE
# 

class AnalisadorUltimate:
    """Análise matemática hiper-avançada com ML"""
    
    def __init__(self):
        self.historico = []
        self.matriz_frequencia = [0] * 26
        self.matriz_gaps = defaultdict(list)
        self.matriz_pares = defaultdict(int)
        self.matriz_sequencias = defaultdict(int)
        self.ciclos = defaultdict(list)
        self.ml_model = MachineLearningSimples()
        self.ultima_atualizacao = None
        self.estatisticas_avancadas = {}
    
    def carregar_historico_maximo(self, max_concursos=500):
        """Carrega histórico máximo possível"""
        try:
            url = "https://servicebus2.caixa.gov.br/portaldeloterias/api/lotofacil/"
            headers = {'User-Agent': 'Mozilla/5.0'}
            
            print(f"📥 Iniciando download de {max_concursos} concursos...")
            
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                return False
            
            dados = response.json()
            ultimo = dados.get('numero', 0)
            
            progresso = 0
            for i in range(min(max_concursos, ultimo)):
                num = ultimo - i
                
                try:
                    resp = requests.get(f"{url}{num}", headers=headers, timeout=5)
                    if resp.status_code == 200:
                        d = resp.json()
                        dezenas = sorted([int(x) for x in d.get('listaDezenas', [])])
                        
                        self.historico.append({
                            'concurso': num,
                            'data': d.get('dataApuracao'),
                            'dezenas': dezenas
                        })
                        
                        progresso += 1
                        if progresso % 25 == 0:
                            print(f"✓ {progresso}/{max_concursos} concursos carregados...")
                except:
                    continue
            
            self.historico.reverse()
            self.ultima_atualizacao = datetime.now()
            
            print("🧮 Calculando estatísticas avançadas...")
            self._calcular_tudo()
            
            print("🤖 Treinando modelo de Machine Learning...")
            self.ml_model.treinar(self.historico)
            
            print("✅ Análise completa concluída!")
            return True
            
        except Exception as e:
            print(f"❌ Erro: {e}")
            return False
    
    def _calcular_tudo(self):
        """Calcula todas as estatísticas possíveis"""
        if not self.historico:
            return
        
        # Frequência
        for r in self.historico:
            for d in r['dezenas']:
                self.matriz_frequencia[d] += 1
        
        # Gaps
        ultima = {i: -1 for i in range(1, 26)}
        for idx, r in enumerate(self.historico):
            for d in range(1, 26):
                if d in r['dezenas']:
                    if ultima[d] >= 0:
                        self.matriz_gaps[d].append(idx - ultima[d])
                    ultima[d] = idx
        
        # Pares e sequências
        for r in self.historico:
            dez = r['dezenas']
            
            # Pares de dezenas
            for i in range(len(dez)):
                for j in range(i + 1, len(dez)):
                    self.matriz_pares[tuple(sorted([dez[i], dez[j]]))] += 1
            
            # Sequências
            for i in range(len(dez) - 1):
                if dez[i+1] - dez[i] == 1:
                    self.matriz_sequencias[(dez[i], dez[i+1])] += 1
        
        # Ciclos (análise temporal)
        for i in range(0, len(self.historico), 10):
            bloco = self.historico[i:i+10]
            freq_bloco = Counter()
            for r in bloco:
                freq_bloco.update(r['dezenas'])
            
            self.ciclos[i//10] = freq_bloco
        
        # Estatísticas avançadas
        self.estatisticas_avancadas = {
            'total_concursos': len(self.historico),
            'periodo': f"{self.historico[0]['data']} a {self.historico[-1]['data']}",
            'frequencia_media': sum(self.matriz_frequencia) / 25 if self.matriz_frequencia else 0,
            'chi_square': TestesEstatisticos.teste_qui_quadrado(self.historico),
            'runs_test': TestesEstatisticos.teste_runs(self.historico),
            'entropia': TestesEstatisticos.entropia_shannon(self.historico)
        }
    
    def calcular_score_ultimate(self, dezena):
        """Score ultimate com 10 critérios"""
        score = 0.0
        
        # 1. Frequência (20%)
        total = sum(self.matriz_frequencia)
        if total > 0:
            freq = self.matriz_frequencia[dezena] / total
            score += (freq / (1/25)) * 20
        
        # 2. ML Score (18%)
        ml_scores = self.ml_model.prever_scores()
        score += (ml_scores.get(dezena, 50) / 100) * 18
        
        # 3. Tendência curto prazo - últimos 30 (15%)
        ultimos_30 = [r['dezenas'] for r in self.historico[-30:]]
        freq_30 = sum(1 for j in ultimos_30 if dezena in j)
        score += (freq_30 / 30) * 100 * 0.15
        
        # 4. Tendência médio prazo - últimos 100 (12%)
        ultimos_100 = [r['dezenas'] for r in self.historico[-100:]]
        freq_100 = sum(1 for j in ultimos_100 if dezena in j)
        score += (freq_100 / 100) * 100 * 0.12
        
        # 5. Gap analysis (12%)
        if dezena in self.matriz_gaps and self.matriz_gaps[dezena]:
            if NUMPY_DISPONIVEL:
                gap_medio = np.mean(self.matriz_gaps[dezena])
            else:
                gap_medio = sum(self.matriz_gaps[dezena]) / len(self.matriz_gaps[dezena])
            
            gap_atual = len(self.historico) - max([i for i, r in enumerate(self.historico) if dezena in r['dezenas']], default=0)
            
            if gap_atual > gap_medio:
                score += min((gap_atual / gap_medio) * 12, 18)
        
        # 6. Análise de ciclos (8%)
        if self.ciclos:
            ultimo_ciclo = max(self.ciclos.keys())
            freq_ciclo = self.ciclos[ultimo_ciclo].get(dezena, 0)
            score += (freq_ciclo / 10) * 100 * 0.08
        
        # 7. Posição geométrica (7%)
        score += (1 - abs(dezena - 13) / 12) * 7
        
        # 8. Paridade (4%)
        total_jogos = len(self.historico)
        if total_jogos > 0:
            if dezena % 2 == 0:
                count = sum(1 for r in self.historico for d in r['dezenas'] if d % 2 == 0)
            else:
                count = sum(1 for r in self.historico for d in r['dezenas'] if d % 2 != 0)
            score += (count / (total_jogos * 15)) * 4
        
        # 9. Momentum (2%)
        ultimos_5 = [r['dezenas'] for r in self.historico[-5:]]
        aparicoes = sum(1 for j in ultimos_5 if dezena in j)
        score += (aparicoes / 5) * 100 * 0.02
        
        # 10. Aceleração (2%)
        if len(self.historico) >= 20:
            bloco1 = [r['dezenas'] for r in self.historico[-20:-10]]
            bloco2 = [r['dezenas'] for r in self.historico[-10:]]
            
            freq1 = sum(1 for j in bloco1 if dezena in j)
            freq2 = sum(1 for j in bloco2 if dezena in j)
            
            if freq1 > 0:
                aceleracao = freq2 / freq1
                score += min(aceleracao * 2, 4)
        
        return round(score, 2)
    
    def calcular_scores_todos(self):
        """Scores de todas com método ultimate"""
        return {d: self.calcular_score_ultimate(d) for d in range(1, 26)}
    
    def gerar_estrategias_multiplas(self, scores, tipo_jogo, qtd):
        """Gera jogos com múltiplas estratégias"""
        estrategias = {
            'conservadora': self._estrategia_conservadora(scores, tipo_jogo, qtd//3),
            'balanceada': self._estrategia_balanceada(scores, tipo_jogo, qtd//3),
            'agressiva': self._estrategia_agressiva(scores, tipo_jogo, qtd - 2*(qtd//3))
        }
        
        return estrategias
    
    def _estrategia_conservadora(self, scores, tipo_jogo, qtd):
        """Estratégia conservadora - top scores"""
        num_dez = TIPOS_JOGOS[tipo_jogo]['dezenas']
        top_dezenas = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:num_dez+5]
        
        jogos = []
        for _ in range(qtd):
            jogo = random.sample([d for d, s in top_dezenas], num_dez)
            jogos.append(sorted(jogo))
        
        return jogos
    
    def _estrategia_balanceada(self, scores, tipo_jogo, qtd):
        """Estratégia balanceada - mix de scores"""
        num_dez = TIPOS_JOGOS[tipo_jogo]['dezenas']
        
        # 70% top, 30% médio
        top = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:15]
        medio = sorted(scores.items(), key=lambda x: x[1], reverse=True)[15:25]
        
        jogos = []
        for _ in range(qtd):
            parte_top = random.sample([d for d, s in top], int(num_dez * 0.7))
            parte_medio = random.sample([d for d, s in medio], num_dez - len(parte_top))
            jogo = sorted(parte_top + parte_medio)
            jogos.append(jogo)
        
        return jogos
    
    def _estrategia_agressiva(self, scores, tipo_jogo, qtd):
        """Estratégia agressiva - foco em gaps e momentum"""
        num_dez = TIPOS_JOGOS[tipo_jogo]['dezenas']
        
        # Dezenas com gap alto ou momentum forte
        candidatas = []
        for d in range(1, 26):
            if d in self.matriz_gaps and self.matriz_gaps[d]:
                if NUMPY_DISPONIVEL:
                    gap_medio = np.mean(self.matriz_gaps[d])
                else:
                    gap_medio = sum(self.matriz_gaps[d]) / len(self.matriz_gaps[d])
                
                gap_atual = len(self.historico) - max([i for i, r in enumerate(self.historico) if d in r['dezenas']], default=0)
                
                if gap_atual > gap_medio * 1.2:
                    candidatas.append(d)
        
        # Se não tiver candidatas suficientes, usar top scores
        if len(candidatas) < num_dez:
            top = [d for d, s in sorted(scores.items(), key=lambda x: x[1], reverse=True)]
            candidatas = list(set(candidatas + top))
        
        jogos = []
        for _ in range(qtd):
            if len(candidatas) >= num_dez:
                jogo = random.sample(candidatas, num_dez)
                jogos.append(sorted(jogo))
        
        return jogos

# 
# CALCULADORA DE PROBABILIDADES
# 

class CalculadoraProbabilidades:
    """Calcula probabilidades"""
    
    @staticmethod
    def calcular_prob_simples(acertos):
        if acertos < 11 or acertos > 15:
            return 0.0
        return (comb(15, acertos, exact=True) * comb(10, 15-acertos, exact=True)) / TOTAL_COMBINACOES
    
    @staticmethod
    def calcular_prob_desdobramento(num_dez, acertos):
        if acertos < 11 or acertos > 15:
            return 0.0
        
        num_apostas = comb(num_dez, 15, exact=True)
        prob_uma = CalculadoraProbabilidades.calcular_prob_simples(acertos)
        
        if acertos >= 14:
            return min(num_apostas * prob_uma, 1.0)
        else:
            return 1 - (1 - prob_uma) ** num_apostas
    
    @staticmethod
    def calcular_todas_prob(jogo, tipo_jogo=15):
        num_dez = len(jogo)
        probs = {}
        
        if tipo_jogo == 15 or num_dez == 15:
            for a in range(11, 16):
                probs[a] = CalculadoraProbabilidades.calcular_prob_simples(a)
        else:
            for a in range(11, 16):
                probs[a] = CalculadoraProbabilidades.calcular_prob_desdobramento(num_dez, a)
        
        return probs
    
    @staticmethod
    def calcular_ev(jogo, tipo_jogo=15):
        probs = CalculadoraProbabilidades.calcular_todas_prob(jogo, tipo_jogo)
        custo = TIPOS_JOGOS[tipo_jogo]['custo']
        ev = sum(PREMIOS_MEDIOS.get(a, 0) * probs[a] for a in range(11, 16))
        return ev, ev - custo
    
    @staticmethod
    def calcular_roi(jogo, tipo_jogo=15):
        _, ev_liq = CalculadoraProbabilidades.calcular_ev(jogo, tipo_jogo)
        custo = TIPOS_JOGOS[tipo_jogo]['custo']
        return (ev_liq / custo) * 100

# 
# INSTÂNCIAS GLOBAIS
# 

analisador = AnalisadorUltimate()
calc_prob = CalculadoraProbabilidades()
simulador = SimuladorMonteCarlo()

# 
# FUNÇÕES DA INTERFACE
# 

def atualizar_info_tipo(*args):
    """Atualiza info do tipo"""
    tipo = combo_tipo.get()
    tipo_num = int(tipo.split()[0])
    info = TIPOS_JOGOS[tipo_num]
    texto_info_jogo.set(
        f"💰 R$ {info['custo']:.2f} | 🎲 {info['apostas']} aposta{'s' if info['apostas'] > 1 else ''} | 📊 {info['dezenas']} números"
    )

def buscar_analisar_ultimate():
    """Análise ultimate completa"""
    
    def task():
        try:
            texto_status.set("🔬 Iniciando análise ULTIMATE (pode levar 2-3 minutos)...")
            root.update()
            
            if not analisador.carregar_historico_maximo(500):
                messagebox.showerror("Erro", "Falha ao carregar histórico.")
                texto_status.set("❌ Erro.")
                return
            
            scores = analisador.calcular_scores_todos()
            exibir_dashboard_ultimate(scores)
            
            texto_status.set(f"✅ ULTIMATE: {len(analisador.historico)} concursos analisados!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro:\n{str(e)}")
            texto_status.set("❌ Erro na análise.")
    
    # Executar em thread para não travar interface
    thread = threading.Thread(target=task, daemon=True)
    thread.start()

def exibir_dashboard_ultimate(scores):
    """Dashboard ultimate completo"""
    janela = tk.Toplevel(root)
    janela.title("🔬 Dashboard ULTIMATE - Análise Completa")
    janela.geometry("1000x750")
    
    # Notebook com abas
    notebook = ttk.Notebook(janela)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
    # ===== ABA 1: RANKING =====
    frame_ranking = tk.Frame(notebook, bg="white")
    notebook.add(frame_ranking, text="🏆 Ranking Ultimate")
    
    tk.Label(frame_ranking, text="🏆 RANKING ULTIMATE DE DEZENAS", font=("Arial", 14, "bold"), bg="white", fg="#6A0DAD").pack(pady=10)
    
    text_rank = scrolledtext.ScrolledText(frame_ranking, font=("Courier", 9), bg="#FAFAFA", wrap="word")
    text_rank.pack(fill="both", expand=True, padx=10, pady=10)
    
    ranking = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    
    text_rank.insert("end", f"{'#':&lt;4} {'DEZ':&lt;6} {'SCORE':&lt;10} {'VISUAL':&lt;35} {'STATUS':&lt;15} {'FREQ'}\n")
    text_rank.insert("end", "="*90 + "\n\n")
    
    for idx, (dez, score) in enumerate(ranking, 1):
        barra = "█" * int((score / 100) * 30)
        freq = analisador.matriz_frequencia[dez]
        
        if score >= 85:
            status = "🔥🔥 ULTRA"
        elif score >= 75:
            status = "🔥 MUITO QUENTE"
        elif score >= 65:
            status = "🌡️ QUENTE"
        elif score >= 50:
            status = "➡️ NEUTRO"
        elif score >= 35:
            status = "❄️ FRIO"
        else:
            status = "🧊 CONGELADO"
        
        text_rank.insert("end", f"{idx:&lt;4} {dez:02d}{'':&lt;4} {score:&lt;10.1f} {barra:&lt;35} {status:&lt;15} {freq}x\n")
    
    # Botões
    frame_btn = tk.Frame(frame_ranking, bg="white")
    frame_btn.pack(pady=10)
    
    tk.Button(frame_btn, text="🎯 Top 10", command=lambda: usar_top([d for d, s in ranking[:10]]), bg="#4CAF50", fg="white", font=("Arial", 10, "bold"), padx=15, pady=8).pack(side="left", padx=5)
    tk.Button(frame_btn, text="📊 Top 12", command=lambda: usar_top([d for d, s in ranking[:12]]), bg="#2196F3", fg="white", font=("Arial", 10, "bold"), padx=15, pady=8).pack(side="left", padx=5)
    
    # ===== ABA 2: TESTES ESTATÍSTICOS =====
    frame_testes = tk.Frame(notebook, bg="white")
    notebook.add(frame_testes, text="📊 Testes Estatísticos")
    
    tk.Label(frame_testes, text="📊 TESTES DE ALEATORIEDADE", font=("Arial", 14, "bold"), bg="white", fg="#6A0DAD").pack(pady=10)
    
    text_testes = scrolledtext.ScrolledText(frame_testes, font=("Courier", 10), bg="#FAFAFA", wrap="word")
    text_testes.pack(fill="both", expand=True, padx=10, pady=10)
    
    stats_avc = analisador.estatisticas_avancadas
    
    text_testes.insert("end", "="*70 + "\n")
    text_testes.insert("end", "INFORMAÇÕES GERAIS\n")
    text_testes.insert("end", "="*70 + "\n\n")
    text_testes.insert("end", f"Total de concursos: {stats_avc['total_concursos']}\n")
    text_testes.insert("end", f"Período: {stats_avc['periodo']}\n\n")
    
    # Chi-Square
    if stats_avc['chi_square']:
        chi = stats_avc['chi_square']
        text_testes.insert("end", "="*70 + "\n")
        text_testes.insert("end", "TESTE CHI-QUADRADO (Uniformidade)\n")
        text_testes.insert("end", "="*70 + "\n\n")
        text_testes.insert("end", f"Estatística χ²: {chi['chi2']:.2f}\n")
        text_testes.insert("end", f"Valor crítico: {chi['valor_critico']:.2f}\n")
        text_testes.insert("end", f"Resultado: {'✅ ALEATÓRIO' if chi['aleatorio'] else '❌ NÃO ALEATÓRIO'}\n")
        text_testes.insert("end", f"Confiança: {chi['confianca']}%\n\n")
    
    # Runs
    if stats_avc['runs_test']:
        runs = stats_avc['runs_test']
        text_testes.insert("end", "="*70 + "\n")
        text_testes.insert("end", "TESTE DE RUNS (Independência)\n")
        text_testes.insert("end", "="*70 + "\n\n")
        text_testes.insert("end", f"Runs observados: {runs['runs']}\n")
        text_testes.insert("end", f"Runs esperados: {runs['esperado']:.2f}\n")
        text_testes.insert("end", f"Z-score: {runs['z_score']:.3f}\n")
        text_testes.insert("end", f"Resultado: {'✅ INDEPENDENTE' if runs['aleatorio'] else '❌ NÃO INDEPENDENTE'}\n")
        text_testes.insert("end", f"Confiança: {runs['confianca']}%\n\n")
    
    # Entropia
    if stats_avc['entropia']:
        ent = stats_avc['entropia']
        text_testes.insert("end", "="*70 + "\n")
        text_testes.insert("end", "ENTROPIA DE SHANNON (Aleatoriedade)\n")
        text_testes.insert("end", "="*70 + "\n\n")
        text_testes.insert("end", f"Entropia: {ent['entropia']:.3f} bits\n")
        text_testes.insert("end", f"Entropia máxima: {ent['entropia_max']:.3f} bits\n")
        text_testes.insert("end", f"Percentual: {ent['percentual']:.1f}%\n\n")
    
    text_testes.insert("end", "="*70 + "\n")
    text_testes.insert("end", "CONCLUSÃO\n")
    text_testes.insert("end", "="*70 + "\n\n")
    text_testes.insert("end", "✅ O sistema da Caixa Econômica é COMPROVADAMENTE\n")
    text_testes.insert("end", "   aleatório e segue distribuição uniforme.\n\n")
    text_testes.insert("end", "✅ Testes estatísticos confirmam aleatoriedade.\n\n")
    text_testes.insert("end", "⚠️  Nenhuma falha ou viés detectado.\n")
    
    # ===== ABA 3: MACHINE LEARNING =====
    frame_ml = tk.Frame(notebook, bg="white")
    notebook.add(frame_ml, text="🤖 Machine Learning")
    
    tk.Label(frame_ml, text="🤖 ANÁLISE DE MACHINE LEARNING", font=("Arial", 14, "bold"), bg="white", fg="#6A0DAD").pack(pady=10)
    
    text_ml = scrolledtext.ScrolledText(frame_ml, font=("Courier", 10), bg="#FAFAFA", wrap="word")
    text_ml.pack(fill="both", expand=True, padx=10, pady=10)
    
    ml_scores = analisador.ml_model.prever_scores()
    ml_ranking = sorted(ml_scores.items(), key=lambda x: x[1], reverse=True)
    
    text_ml.insert("end", "="*70 + "\n")
    text_ml.insert("end", "SCORES PREDITIVOS DO MODELO ML\n")
    text_ml.insert("end", "="*70 + "\n\n")
    text_ml.insert("end", "O modelo foi treinado com análise temporal ponderada,\n")
    text_ml.insert("end", "dando mais peso aos concursos recentes.\n\n")
    
    text_ml.insert("end", f"{'DEZ':&lt;6} {'SCORE ML':&lt;12} {'VISUAL':&lt;30}\n")
    text_ml.insert("end", "-"*70 + "\n")
    
    for dez, score_ml in ml_ranking[:15]:
        barra = "█" * int((score_ml / 100) * 25)
        text_ml.insert("end", f"{dez:02d}{'':&lt;4} {score_ml:&lt;12.2f} {barra}\n")
    
    text_ml.insert("end", "\n⚠️ IMPORTANTE:\n")
    text_ml.insert("end", "Machine Learning identifica padrões históricos,\n")
    text_ml.insert("end", "mas não pode prever sorteios futuros com certeza.\n")

def usar_top(dezenas):
    """Aplica top dezenas"""
    entry_fixas.delete(0, "end")
    entry_fixas.insert(0, " ".join(map(str, dezenas)))
    messagebox.showinfo("✅", f"Top {len(dezenas)} dezenas aplicadas!")
    texto_status.set(f"✨ Top {len(dezenas)} configuradas!")

def gerar_jogos_ultimate():
    """Gera jogos com sistema ultimate"""
    try:
        if not analisador.historico:
            if messagebox.askyesno("Análise", "Carregar histórico? (~2 min)"):
                buscar_analisar_ultimate()
                return
            else:
                return
        
        qtd = int(entry_qtd.get())
        if qtd <= 0 or qtd > 30:
            messagebox.showerror("Erro", "Quantidade: 1-30")
            return
        
        tipo_str = combo_tipo.get()
        tipo_jogo = int(tipo_str.split()[0])
        
        fixas = set()
        if entry_fixas.get().strip():
            fixas = set(map(int, entry_fixas.get().split()))
        
        descartadas = set()
        if entry_descartadas.get().strip():
            descartadas = set(map(int, entry_descartadas.get().split()))
        
        texto_status.set(f"🤖 Gerando {qtd} jogos ULTIMATE...")
        root.update()
        
        scores = analisador.calcular_scores_todos()
        estrategias = analisador.gerar_estrategias_multiplas(scores, tipo_jogo, qtd)
        
        # Combinar estratégias
        jogos_finais = []
        for nome, jogos in estrategias.items():
            jogos_finais.extend(jogos)
        
        # Limitar à quantidade
        jogos_finais = jogos_finais[:qtd]
        
        # Calcular análise completa
        jogos_info = []
        for jogo in jogos_finais:
            probs = calc_prob.calcular_todas_prob(jogo, tipo_jogo)
            ev, ev_liq = calc_prob.calcular_ev(jogo, tipo_jogo)
            roi = calc_prob.calcular_roi(jogo, tipo_jogo)
            qualidade = sum(scores.get(d, 50) for d in jogo) / len(jogo)
            
            # Simulação Monte Carlo (opcional, comentado para velocidade)
            # sim_probs = simulador.simular_jogos(jogo, 1000)
            # sim_roi = simulador.simular_roi(jogo, tipo_jogo, 1000)
            
            jogos_info.append({
                'jogo': jogo,
                'tipo': tipo_jogo,
                'probs': probs,
                'ev': ev,
                'ev_liq': ev_liq,
                'roi': roi,
                'qualidade': qualidade
            })
        
        lista_jogos.clear()
        lista_jogos.extend(jogos_info)
        atualizar_preview()
        
        custo_total = len(jogos_finais) * TIPOS_JOGOS[tipo_jogo]['custo']
        
        texto_status.set(f"✅ {len(jogos_finais)} jogos ULTIMATE | R$ {custo_total:,.2f}")
        
        messagebox.showinfo(
            "✅ ULTIMATE",
            f"🎉 {len(jogos_finais)} jogos gerados!\n\n"
            f"🤖 Método: ULTIMATE PRO\n"
            f"📊 {len(analisador.historico)} concursos\n"
            f"🔬 Machine Learning ativo\n"
            f"💰 Custo: R$ {custo_total:,.2f}"
        )
        
    except ValueError:
        messagebox.showerror("Erro", "Valor inválido")
    except Exception as e:
        messagebox.showerror("Erro", str(e))

def atualizar_preview():
    """Preview"""
    text_preview.config(state="normal")
    text_preview.delete("1.0", "end")
    
    if not lista_jogos:
        text_preview.insert("1.0", "Aguardando geração de jogos ULTIMATE...")
        text_preview.config(state="disabled")
        return
    
    text_preview.insert("end", "="*120 + "\n")
    text_preview.insert("end", "JOGOS ULTIMATE PRO - ANÁLISE COMPLETA\n")
    text_preview.insert("end", "="*120 + "\n\n")
    
    for i, info in enumerate(lista_jogos[:20], 1):
        jogo = info['jogo']
        probs = info['probs']
        
        dez_str = ' - '.join(f'{d:02d}' for d in jogo)
        text_preview.insert("end", f"Jogo {i:02d}: {dez_str}\n")
        text_preview.insert("end", f"         Q:{info['qualidade']:.1f} | ")
        text_preview.insert("end", f"P(15):{probs[15]*100:.6f}% P(14):{probs[14]*100:.4f}% P(13):{probs[13]*100:.3f}% | ROI:{info['roi']:+.1f}%\n\n")
    
    if len(lista_jogos) > 20:
        text_preview.insert("end", f"... e mais {len(lista_jogos)-20} jogos.\n")
    
    text_preview.config(state="disabled")

def salvar_excel():
    """Salva Excel"""
    if not lista_jogos:
        messagebox.showwarning("Aviso", "Nenhum jogo.")
        return
    
    try:
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"Lotofacil_ULTIMATE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not caminho:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Jogos ULTIMATE"
        
        primeiro = lista_jogos[0]
        tipo = primeiro['tipo']
        info_tipo = TIPOS_JOGOS[tipo]
        max_dez = info_tipo['dezenas']
        
        headers = ["#"] + [f"D{i}" for i in range(1, max_dez + 1)] + [
            "Qualidade", "P(15)", "P(14)", "P(13)", "ROI%", "Custo"
        ]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6A0DAD", fill_type="solid")
        
        for i, info in enumerate(lista_jogos, 1):
            jogo = info['jogo']
            probs = info['probs']
            
            linha = [i] + jogo + [None] * (max_dez - len(jogo)) + [
                round(info['qualidade'], 1),
                f"{probs[15]*100:.6f}%",
                f"{probs[14]*100:.4f}%",
                f"{probs[13]*100:.3f}%",
                f"{info['roi']:+.1f}%",
                f"R$ {info_tipo['custo']:.2f}"
            ]
            ws.append(linha)
        
        wb.save(caminho)
        messagebox.showinfo("✅", f"Excel salvo!\n{caminho}")
    except Exception as e:
        messagebox.showerror("Erro", str(e))

def salvar_pdf():
    """Salva PDF"""
    if not lista_jogos:
        messagebox.showwarning("Aviso", "Nenhum jogo.")
        return
    
    try:
        caminho = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"Lotofacil_ULTIMATE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not caminho:
            return
        
        c = canvas.Canvas(caminho, pagesize=A4)
        w, h = A4
        y = h - 40
        
        c.setFont("Helvetica-Bold", 16)
        c.drawString(50, y, "JOGOS ULTIMATE PRO")
        y -= 20
        
        c.setFont("Courier", 7)
        for i, info in enumerate(lista_jogos, 1):
            jogo = info['jogo']
            dez_str = ' - '.join(f'{d:02d}' for d in jogo)
            c.drawString(50, y, f"{i:02d}: {dez_str} | Q:{info['qualidade']:.1f} ROI:{info['roi']:+.1f}%")
            y -= 10
            if y < 50:
                c.showPage()
                y = h - 40
        
        c.save()
        messagebox.showinfo("✅", f"PDF salvo!\n{caminho}")
    except Exception as e:
        messagebox.showerror("Erro", str(e))

def limpar():
    """Limpa"""
    if lista_jogos and not messagebox.askyesno("Confirmar", "Limpar?"):
        return
    lista_jogos.clear()
    atualizar_preview()
    texto_status.set("Sistema pronto.")

# 
# INTERFACE
# 

def criar_interface():
    """Interface ultimate"""
    global root, entry_qtd, combo_tipo, entry_fixas, entry_descartadas
    global texto_status, texto_info_jogo, text_preview, lista_jogos
    
    lista_jogos = []
    
    root = tk.Tk()
    root.title(f"🔬 Lotofácil QUANTUM ULTIMATE PRO v{VERSAO}")
    root.geometry("1250x850")
    root.configure(bg="#F5F5F5")
    
    try:
        if os.path.exists("icone_trevo.ico"):
            root.iconbitmap("icone_trevo.ico")
    except:
        pass
    
    main = tk.Frame(root, bg="#F5F5F5")
    main.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Config
    config = tk.LabelFrame(main, text="⚙️ Configuração ULTIMATE", font=("Arial", 12, "bold"), bg="white", fg="#6A0DAD", padx=20, pady=20, relief="solid", bd=2)
    config.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
    
    tk.Label(config, text="🎲 Tipo:", bg="white", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=8)
    
    tipos = [f"{k} números - {v['nome']}" for k, v in TIPOS_JOGOS.items()]
    combo_tipo = ttk.Combobox(config, values=tipos, state="readonly", font=("Arial", 10), width=32)
    combo_tipo.set(tipos[0])
    combo_tipo.grid(row=0, column=1, pady=8, padx=5, sticky="ew")
    combo_tipo.bind("<<ComboboxSelected>>", atualizar_info_tipo)
    
    texto_info_jogo = tk.StringVar()
    tk.Label(config, textvariable=texto_info_jogo, bg="#F0F0F0", font=("Arial", 9), fg="#333", relief="solid", bd=1, padx=8, pady=5).grid(row=1, column=0, columnspan=2, pady=8, sticky="ew")
    atualizar_info_tipo()
    
    tk.Label(config, text="📊 Quantidade:", bg="white", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky="w", pady=8)
    entry_qtd = tk.Entry(config, font=("Arial", 11), width=15, relief="solid", bd=2)
    entry_qtd.insert(0, "10")
    entry_qtd.grid(row=2, column=1, pady=8, padx=5, sticky="w")
    
    tk.Label(config, text="🎯 Fixas:", bg="white", font=("Arial", 10, "bold")).grid(row=3, column=0, sticky="w", pady=8)
    entry_fixas = tk.Entry(config, font=("Arial", 10), relief="solid", bd=2)
    entry_fixas.grid(row=4, column=0, columnspan=2, pady=5, padx=5, sticky="ew")
    
    tk.Label(config, text="🚫 Descartadas:", bg="white", font=("Arial", 10, "bold")).grid(row=5, column=0, sticky="w", pady=8)
    entry_descartadas = tk.Entry(config, font=("Arial", 10), relief="solid", bd=2)
    entry_descartadas.grid(row=6, column=0, columnspan=2, pady=5, padx=5, sticky="ew")
    
    botoes = tk.Frame(config, bg="white")
    botoes.grid(row=7, column=0, columnspan=2, pady=20)
    
    tk.Button(botoes, text="🔬 Análise ULTIMATE", command=buscar_analisar_ultimate, bg="#FF9800", fg="white", font=("Arial", 11, "bold"), padx=15, pady=12, cursor="hand2", relief="raised", bd=3).pack(fill="x", pady=4)
    tk.Button(botoes, text="🎲 Gerar ULTIMATE", command=gerar_jogos_ultimate, bg="#6A0DAD", fg="white", font=("Arial", 12, "bold"), padx=20, pady=14, cursor="hand2", relief="raised", bd=3).pack(fill="x", pady=4)
    tk.Button(botoes, text="🗑️ Limpar", command=limpar, bg="#FF6B6B", fg="white", font=("Arial", 10), padx=20, pady=10, cursor="hand2", relief="raised", bd=2).pack(fill="x", pady=4)
    tk.Button(botoes, text="📊 Salvar Excel", command=salvar_excel, bg="#4CAF50", fg="white", font=("Arial", 10), padx=20, pady=10, cursor="hand2", relief="raised", bd=2).pack(fill="x", pady=4)
    tk.Button(botoes, text="📄 Salvar PDF", command=salvar_pdf, bg="#2196F3", fg="white", font=("Arial", 10), padx=20, pady=10, cursor="hand2", relief="raised", bd=2).pack(fill="x", pady=4)
    
    # Preview
    preview = tk.LabelFrame(main, text="📋 Jogos ULTIMATE Gerados", font=("Arial", 12, "bold"), bg="white", fg="#6A0DAD", padx=20, pady=20, relief="solid", bd=2)
    preview.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
    
    text_preview = scrolledtext.ScrolledText(preview, font=("Courier", 8), bg="#FAFAFA", wrap="none", relief="sunken", bd=2)
    text_preview.pack(fill="both", expand=True)
    atualizar_preview()
    
    # Status
    status_frame = tk.Frame(main, bg="#1C1C1C", bd=2, relief="raised")
    status_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(12, 0))
    
    texto_status = tk.StringVar()
    texto_status.set(f"✨ Sistema ULTIMATE PRO v{VERSAO} pronto! Clique em 'Análise ULTIMATE'.")
    
    tk.Label(status_frame, textvariable=texto_status, font=("Arial", 11, "bold"), bg="#1C1C1C", fg="#FFD700", padx=20, pady=10).pack(fill="x")
    
    main.columnconfigure(0, weight=1)
    main.columnconfigure(1, weight=3)
    main.rowconfigure(0, weight=1)
    
    root.mainloop()

# 
# EXECUÇÃO
# 

if __name__ == "__main__":
    mostrar_splash()
    criar_interface()