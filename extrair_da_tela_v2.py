"""
EXTRAIR PROCESSOS - ÁREA DE TRANSFERÊNCIA
Versão sem dependências externas (usa tkinter nativo)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re
import tkinter as tk

def extrair_processos_clipboard():
    """Extrai processos da área de transferência usando tkinter"""
    
    print("\n📋 Lendo área de transferência...")
    
    try:
        root = tk.Tk()
        root.withdraw()  # Esconder janela
        texto = root.clipboard_get()
        root.destroy()
    except Exception as e:
        print(f"❌ Erro ao ler área de transferência: {e}")
        return []
    
    if not texto:
        print("❌ Área de transferência vazia!")
        return []
    
    print(f"   ✅ {len(texto)} caracteres lidos")
    
    # Buscar processos no padrão TJSP
    print(f"\n🔍 Procurando processos...")
    
    pattern = r'\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}'
    matches = re.findall(pattern, texto)
    
    # Remover duplicatas mantendo ordem
    processos = []
    vistos = set()
    for proc in matches:
        if proc not in vistos:
            processos.append(proc)
            vistos.add(proc)
    
    print(f"   ✅ {len(processos)} processos únicos encontrados!")
    
    return processos

def criar_planilha(processos, arquivo):
    """Cria planilha Excel"""
    
    print(f"\n📊 Criando planilha Excel...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"
    
    ws['A1'] = 'Nº Processo'
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = header_alignment
    
    for row_idx, processo in enumerate(processos, 2):
        ws.cell(row=row_idx, column=1, value=processo)
    
    ws.column_dimensions['A'].width = 35
    ws.row_dimensions[1].height = 25
    ws.auto_filter.ref = f'A1:A{len(processos)+1}'
    
    wb.save(arquivo)
    
    print(f"   ✅ Salvo: {arquivo}")

if __name__ == "__main__":
    print("="*70)
    print("📋 EXTRAIR PROCESSOS DA ÁREA DE TRANSFERÊNCIA")
    print("="*70)
    
    print("\n📝 PASSO A PASSO:")
    print("="*70)
    print("   1. Abra o navegador e acesse o PUSH de Requisitórios")
    print("   2. Faça a busca: 01/11/2024 a 31/12/2025")
    print("   3. Aguarde os resultados carregarem na tela")
    print("   4. Selecione TODA a lista:")
    print("      → Ctrl+A (selecionar tudo)")
    print("      → OU arraste o mouse sobre a lista")
    print("   5. Copie: Ctrl+C")
    print("   6. Volte aqui e pressione ENTER")
    print("="*70)
    
    input("\n>>> ENTER após copiar a lista no navegador <<<\n")
    
    try:
        processos = extrair_processos_clipboard()
        
        if len(processos) == 0:
            print("\n❌ Nenhum processo encontrado!")
            print("\n💡 DICAS:")
            print("   1. Certifique-se de copiar a lista completa")
            print("   2. Os números devem estar no formato:")
            print("      XXXXXXX-XX.XXXX.X.XX.XXXX")
            print("   3. Tente copiar novamente")
        else:
            print(f"\n" + "="*70)
            print(f"✅ SUCESSO! {len(processos)} PROCESSOS EXTRAÍDOS!")
            print("="*70)
            
            # Preview
            print(f"\n📋 PREVIEW (primeiros 20 processos):")
            print(f"\n   {'#':<5} {'Nº Processo':<35}")
            print(f"   {'-'*5} {'-'*35}")
            
            for i, proc in enumerate(processos[:20], 1):
                print(f"   {i:<5} {proc:<35}")
            
            if len(processos) > 20:
                print(f"\n   ... e mais {len(processos) - 20} processos")
            
            # Criar planilha
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo = f"processos_push_{timestamp}.xlsx"
            
            criar_planilha(processos, arquivo)
            
            print("\n" + "="*70)
            print("🎉 PLANILHA CRIADA COM SUCESSO!")
            print("="*70)
            print(f"\n📁 Arquivo: {arquivo}")
            print(f"📊 Total: {len(processos)} processos")
            print(f"📂 Pasta: C:\TAX_MASTER_DEV\tax_master_static")
            
            # Verificar se atingiu meta
            if len(processos) >= 500:
                print(f"\n✅ Meta atingida! {len(processos)} processos (meta: 500)")
            elif len(processos) < 500:
                faltam = 500 - len(processos)
                print(f"\n⚠️  {len(processos)} de 500 processos (faltam {faltam})")
                print(f"\n💡 Para buscar mais:")
                print(f"   1. No navegador, vá para a próxima página")
                print(f"   2. Copie essa página também (Ctrl+A, Ctrl+C)")
                print(f"   3. Execute o script novamente")
                print(f"   4. Ele somará aos anteriores")
        
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nENTER para fechar...\n")
    
    print("\n✅ FIM!")
