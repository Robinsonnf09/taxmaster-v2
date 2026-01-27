"""
CADASTRO NO PUSH - TJSP (TESTE)
URL: https://esaj.tjsp.jus.br/push/index.do
Sistema: e-SAJ
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import os

class PushTJSP:
    
    def __init__(self):
        self.driver = None
        self.sucessos = []
        self.falhas = []
        self.url_push = "https://esaj.tjsp.jus.br/push/index.do"
    
    def iniciar(self):
        print("\n🌐 Iniciando Chrome...")
        
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--ignore-certificate-errors')
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        print("✅ Chrome iniciado!")
    
    def acessar_push(self):
        print(f"\n🔐 Acessando PUSH TJSP...")
        print(f"   URL: {self.url_push}")
        
        self.driver.get(self.url_push)
        time.sleep(5)
        
        print("\n" + "="*70)
        print("⚠️  VERIFICAÇÃO:")
        print("="*70)
        print("   - Sistema PUSH do TJSP carregou?")
        print("   - Você vê o formulário de cadastro?")
        print("   - Precisa fazer login? Se sim, faça agora")
        print("="*70)
        
        input("\n>>> ENTER quando estiver vendo a tela do PUSH <<<\n")
        
        print("✅ PUSH acessado!")
        return True
    
    def cadastrar_processo(self, numero_processo):
        try:
            print(f"\n📝 Cadastrando: {numero_processo}")
            
            # Procurar campo de número
            campo = None
            
            # Estratégias múltiplas
            seletores = [
                "//input[@type='text']",
                "//input[contains(@name, 'processo')]",
                "//input[contains(@id, 'processo')]",
                "//input[contains(@placeholder, 'processo')]",
                "//input[contains(@class, 'processo')]"
            ]
            
            for seletor in seletores:
                try:
                    campos = self.driver.find_elements(By.XPATH, seletor)
                    if campos:
                        # Pegar o primeiro visível
                        for c in campos:
                            if c.is_displayed():
                                campo = c
                                print(f"   ✅ Campo encontrado")
                                break
                        if campo:
                            break
                except:
                    continue
            
            if not campo:
                print(f"   ⚠️  Campo não encontrado automaticamente")
                print(f"   💡 Digite manualmente: {numero_processo}")
                input(f"   >>> ENTER após digitar <<<\n")
            else:
                # Preencher
                campo.clear()
                time.sleep(0.3)
                campo.send_keys(numero_processo)
                time.sleep(0.5)
                print(f"   ✅ Número digitado")
            
            # Procurar botão INCLUIR/CADASTRAR
            btn = None
            
            botoes = [
                "//button[contains(text(), 'INCLUIR')]",
                "//button[contains(text(), 'Incluir')]",
                "//button[contains(text(), 'CADASTRAR')]",
                "//button[contains(text(), 'Cadastrar')]",
                "//input[@value='INCLUIR']",
                "//input[@value='Incluir']",
                "//button[@type='submit']"
            ]
            
            for xpath in botoes:
                try:
                    btn = self.driver.find_element(By.XPATH, xpath)
                    if btn.is_displayed():
                        print(f"   ✅ Botão encontrado")
                        break
                except:
                    continue
            
            if not btn:
                print(f"   ⚠️  Botão não encontrado")
                input(f"   >>> Clique manualmente em INCLUIR <<<\n")
            else:
                btn.click()
                print(f"   ⏳ Aguardando...")
                time.sleep(4)
            
            # Verificar resultado
            page = self.driver.page_source.lower()
            
            if any(x in page for x in ['sucesso', 'incluído', 'cadastrado']):
                print(f"   ✅ {numero_processo} CADASTRADO!")
                self.sucessos.append(numero_processo)
                return True
            elif any(x in page for x in ['já cadastrado', 'já existe']):
                print(f"   ⚠️  {numero_processo} já estava cadastrado")
                self.sucessos.append(numero_processo)
                return True
            elif 'erro' in page or 'não encontrado' in page:
                print(f"   ❌ Erro ao cadastrar")
                self.falhas.append(numero_processo)
                return False
            else:
                print(f"   ⚠️  Status desconhecido")
                opcao = input(f"   >>> Cadastrou com sucesso? (s/n): ").lower()
                if opcao == 's':
                    self.sucessos.append(numero_processo)
                    return True
                else:
                    self.falhas.append(numero_processo)
                    return False
            
        except Exception as e:
            print(f"   ❌ Erro: {str(e)[:100]}")
            self.falhas.append(numero_processo)
            return False
    
    def processar_planilha(self, arquivo):
        print(f"\n📊 Processando: {arquivo}")
        
        if not os.path.exists(arquivo):
            print(f"❌ Arquivo não encontrado!")
            return
        
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        processos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                processos.append(str(row[0]).strip())
        
        total = len(processos)
        print(f"✅ {total} processos\n")
        
        for idx, numero in enumerate(processos, 1):
            print(f"\n{'='*70}")
            print(f"Processo {idx}/{total}")
            print(f"{'='*70}")
            
            self.cadastrar_processo(numero)
            
            if idx < total:
                time.sleep(2)
        
        # Resumo
        print("\n" + "="*70)
        print("📊 RESUMO FINAL:")
        print("="*70)
        print(f"   🏛️  Tribunal: TJSP")
        print(f"   📋 Total: {total}")
        print(f"   ✅ Sucessos: {len(self.sucessos)}")
        print(f"   ❌ Falhas: {len(self.falhas)}")
        if total > 0:
            print(f"   📊 Taxa: {(len(self.sucessos)/total*100):.1f}%")
        
        if self.falhas:
            print(f"\n❌ FALHAS:")
            for p in self.falhas[:20]:
                print(f"   - {p}")
        
        print("="*70)
    
    def fechar(self):
        if self.driver:
            self.driver.quit()
            print("\n🔒 Navegador fechado")

# MAIN
if __name__ == "__main__":
    print("="*70)
    print("🔔 CADASTRO NO PUSH - TJSP (TESTE)")
    print("="*70)
    print("\n🎯 Sistema: e-SAJ do TJSP")
    print("🔗 URL: https://esaj.tjsp.jus.br/push/index.do")
    
    push = PushTJSP()
    
    try:
        input("\nENTER para começar...\n")
        
        push.iniciar()
        
        if push.acessar_push():
            
            print("\n💡 MODO:")
            print("   1. Planilha completa")
            print("   2. Teste com 1 processo")
            print("   3. Teste com 5 processos")
            
            modo = input("\nDigite 1, 2 ou 3: ").strip()
            
            if modo == "1":
                arq = "processos_push_20260126_185045.xlsx"
                push.processar_planilha(arq)
            
            elif modo == "2":
                num = input("\nNúmero do processo: ").strip()
                push.cadastrar_processo(num)
            
            elif modo == "3":
                arq = "processos_push_20260126_185045.xlsx"
                print("\n📊 Carregando 5 primeiros processos...")
                
                wb = openpyxl.load_workbook(arq)
                ws = wb.active
                
                processos_teste = []
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                    if row[0] and i < 5:
                        processos_teste.append(str(row[0]).strip())
                
                print(f"✅ {len(processos_teste)} processos para teste\n")
                
                for idx, num in enumerate(processos_teste, 1):
                    print(f"\n{'='*70}")
                    print(f"Teste {idx}/{len(processos_teste)}")
                    print(f"{'='*70}")
                    push.cadastrar_processo(num)
                    if idx < len(processos_teste):
                        time.sleep(2)
                
                # Resumo
                print("\n" + "="*70)
                print("📊 RESUMO DO TESTE:")
                print("="*70)
                print(f"   Total testado: {len(processos_teste)}")
                print(f"   ✅ Sucessos: {len(push.sucessos)}")
                print(f"   ❌ Falhas: {len(push.falhas)}")
                print("="*70)
        
        input("\n\nENTER para fechar...\n")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Operação cancelada")
    
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        push.fechar()
    
    print("\n✅ CONCLUÍDO!")
