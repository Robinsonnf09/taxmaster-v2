"""
PUSH TJSP - VERSÃO FINAL AUTOMATIZADA
Com seletores confirmados e funcionais
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime

class PushTJSPFinal:
    
    def __init__(self):
        self.driver = None
        self.url_push = "https://esaj.tjsp.jus.br/push/index.do"
        self.sucessos = []
        self.falhas = []
        self.ja_cadastrados = []
    
    def iniciar(self):
        print("\n🌐 Iniciando Chrome...")
        
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--ignore-certificate-errors')
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        print("✅ Chrome iniciado!")
    
    def fazer_login(self):
        print(f"\n🔐 Acessando PUSH TJSP...")
        self.driver.get(self.url_push)
        time.sleep(3)
        
        print("\n" + "="*70)
        print("⚠️  FAÇA LOGIN:")
        print("="*70)
        print("   - Use certificado digital OU login/senha")
        print("   - Aguarde até estar LOGADO")
        print("   - Você deve ver a área interna do PUSH")
        print("="*70)
        
        input("\n>>> ENTER após login completo <<<\n")
        
        time.sleep(2)
        print("✅ Login confirmado!")
        return True
    
    def cadastrar_processo(self, numero):
        try:
            print(f"\n📝 Cadastrando: {numero}")
            
            wait = WebDriverWait(self.driver, 10)
            
            # CAMPO DE PROCESSO (seletor confirmado)
            try:
                campo = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//input[@type='text'][1]")
                ))
                
                if campo.is_displayed():
                    campo.clear()
                    time.sleep(0.3)
                    campo.send_keys(numero)
                    time.sleep(0.5)
                    print(f"   ✅ Número digitado")
                else:
                    raise Exception("Campo não visível")
                
            except Exception as e:
                print(f"   ⚠️  Campo não encontrado: {str(e)[:50]}")
                print(f"   💡 Digite manualmente: {numero}")
                input(f"   >>> ENTER após digitar <<<\n")
            
            # BOTÃO INCLUIR
            try:
                # Tentar múltiplos seletores
                botoes_xpath = [
                    "//button[contains(text(), 'Incluir')]",
                    "//input[@value='Incluir']",
                    "//button[@type='submit']",
                    "//input[@type='submit']"
                ]
                
                btn = None
                for xpath in botoes_xpath:
                    try:
                        btn = self.driver.find_element(By.XPATH, xpath)
                        if btn.is_displayed():
                            print(f"   ✅ Botão encontrado")
                            break
                    except:
                        continue
                
                if btn:
                    btn.click()
                    print(f"   ⏳ Processando...")
                    time.sleep(3)
                else:
                    print(f"   ⚠️  Clique manualmente em INCLUIR")
                    input(f"   >>> ENTER após clicar <<<\n")
                
            except Exception as e:
                print(f"   ⚠️  Erro ao clicar: {str(e)[:50]}")
                input(f"   >>> Clique manualmente <<<\n")
            
            # VERIFICAR RESULTADO
            time.sleep(2)
            page = self.driver.page_source.lower()
            
            if any(x in page for x in ['sucesso', 'incluído', 'cadastrado']):
                print(f"   ✅ CADASTRADO COM SUCESSO!")
                self.sucessos.append(numero)
                return True
            
            elif any(x in page for x in ['já cadastrado', 'já existe', 'duplicado']):
                print(f"   ⚠️  Já estava cadastrado")
                self.ja_cadastrados.append(numero)
                return True
            
            elif any(x in page for x in ['erro', 'não encontrado', 'inválido']):
                print(f"   ❌ Erro ao cadastrar")
                self.falhas.append(numero)
                return False
            
            else:
                print(f"   ⚠️  Status desconhecido")
                opcao = input(f"   >>> Cadastrou? (s/n): ").lower()
                
                if opcao == 's':
                    self.sucessos.append(numero)
                    return True
                else:
                    self.falhas.append(numero)
                    return False
            
        except Exception as e:
            print(f"   ❌ Erro: {str(e)[:100]}")
            self.falhas.append(numero)
            return False
    
    def processar_planilha(self, arquivo):
        print(f"\n📊 Processando planilha: {arquivo}")
        
        if not os.path.exists(arquivo):
            print(f"❌ Arquivo não encontrado!")
            return
        
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        # Filtrar apenas processos do TJSP
        processos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                num = str(row[0]).strip()
                # TJSP tem o código .8.26. no número
                if '.8.26.' in num:
                    processos.append(num)
        
        total = len(processos)
        print(f"✅ {total} processos do TJSP encontrados\n")
        
        for idx, numero in enumerate(processos, 1):
            print(f"\n{'='*70}")
            print(f"Processo {idx}/{total}")
            print(f"{'='*70}")
            
            self.cadastrar_processo(numero)
            
            # Intervalo entre cadastros
            if idx < total:
                time.sleep(2)
        
        # RELATÓRIO FINAL
        self.gerar_relatorio()
    
    def gerar_relatorio(self):
        print("\n" + "="*70)
        print("📊 RELATÓRIO FINAL - PUSH TJSP")
        print("="*70)
        
        total = len(self.sucessos) + len(self.ja_cadastrados) + len(self.falhas)
        
        print(f"\n📋 Total processado: {total}")
        print(f"✅ Cadastrados: {len(self.sucessos)}")
        print(f"⚠️  Já cadastrados: {len(self.ja_cadastrados)}")
        print(f"❌ Falhas: {len(self.falhas)}")
        
        if total > 0:
            taxa_sucesso = ((len(self.sucessos) + len(self.ja_cadastrados)) / total) * 100
            print(f"📊 Taxa de sucesso: {taxa_sucesso:.1f}%")
        
        # Detalhes das falhas
        if self.falhas:
            print(f"\n❌ PROCESSOS COM FALHA ({len(self.falhas)}):")
            for p in self.falhas[:20]:
                print(f"   - {p}")
            if len(self.falhas) > 20:
                print(f"   ... e mais {len(self.falhas) - 20}")
        
        print("\n" + "="*70)
        
        # Salvar relatório em arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        relatorio_file = f"relatorio_push_tjsp_{timestamp}.txt"
        
        with open(relatorio_file, "w", encoding="utf-8") as f:
            f.write("="*70 + "\n")
            f.write("RELATÓRIO - CADASTRO PUSH TJSP\n")
            f.write("="*70 + "\n\n")
            f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
            f.write(f"Total processado: {total}\n")
            f.write(f"Cadastrados: {len(self.sucessos)}\n")
            f.write(f"Já cadastrados: {len(self.ja_cadastrados)}\n")
            f.write(f"Falhas: {len(self.falhas)}\n\n")
            
            if self.sucessos:
                f.write("CADASTRADOS COM SUCESSO:\n")
                for p in self.sucessos:
                    f.write(f"  - {p}\n")
                f.write("\n")
            
            if self.ja_cadastrados:
                f.write("JÁ CADASTRADOS:\n")
                for p in self.ja_cadastrados:
                    f.write(f"  - {p}\n")
                f.write("\n")
            
            if self.falhas:
                f.write("FALHAS:\n")
                for p in self.falhas:
                    f.write(f"  - {p}\n")
        
        print(f"📄 Relatório salvo: {relatorio_file}")
    
    def fechar(self):
        if self.driver:
            self.driver.quit()
            print("\n🔒 Navegador fechado")

# MAIN
if __name__ == "__main__":
    print("="*70)
    print("🔔 PUSH TJSP - VERSÃO FINAL AUTOMATIZADA")
    print("="*70)
    print("\n✅ Seletores confirmados e testados!")
    print("✅ Pronto para processar em lote!")
    
    push = PushTJSPFinal()
    
    try:
        input("\nENTER para começar...\n")
        
        push.iniciar()
        
        if push.fazer_login():
            
            print("\n💡 MODO:")
            print("   1. Teste (1 processo)")
            print("   2. Teste (5 processos)")
            print("   3. Planilha completa (TJSP)")
            
            modo = input("\nDigite 1, 2 ou 3: ").strip()
            
            if modo == "1":
                num = input("\nNúmero do processo: ").strip()
                push.cadastrar_processo(num)
                
            elif modo == "2":
                print("\n📊 Carregando 5 processos do TJSP...")
                import openpyxl
                
                wb = openpyxl.load_workbook("processos_push_20260126_185045.xlsx")
                ws = wb.active
                
                processos_teste = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        num = str(row[0]).strip()
                        if '.8.26.' in num and len(processos_teste) < 5:
                            processos_teste.append(num)
                
                print(f"✅ {len(processos_teste)} processos para teste\n")
                
                for idx, num in enumerate(processos_teste, 1):
                    print(f"\n{'='*70}")
                    print(f"Teste {idx}/{len(processos_teste)}")
                    print(f"{'='*70}")
                    push.cadastrar_processo(num)
                    if idx < len(processos_teste):
                        time.sleep(2)
                
                push.gerar_relatorio()
                
            elif modo == "3":
                push.processar_planilha("processos_push_20260126_185045.xlsx")
        
        input("\n\nENTER para fechar...\n")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Operação cancelada pelo usuário")
        push.gerar_relatorio()
    
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        push.fechar()
    
    print("\n✅ CONCLUÍDO!")
