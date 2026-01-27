"""
REORGANIZAR OFÍCIOS - VERSÃO PDFPLUMBER (ROBUSTO)
Lê PDFs corrompidos e extrai órgão devedor
"""

import os
import shutil
import pdfplumber
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

class ReorganizadorRobusto:
    
    def __init__(self):
        self.pasta_origem = r"oficios_organizados\Erro na Leitura"
        self.pasta_destino = "oficios_organizados_final"
        self.estatisticas = defaultdict(lambda: {'quantidade': 0, 'arquivos': []})
        self.pdfs_corrompidos = []
        
        if not os.path.exists(self.pasta_destino):
            os.makedirs(self.pasta_destino)
    
    def extrair_orgao_devedor(self, pdf_path):
        """Extrai órgão usando pdfplumber (mais robusto)"""
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                
                # Verificar se tem páginas
                if len(pdf.pages) == 0:
                    return 'PDF_Vazio'
                
                # Extrair texto das primeiras 3 páginas
                texto = ""
                paginas_ler = min(3, len(pdf.pages))
                
                for i in range(paginas_ler):
                    try:
                        pagina_texto = pdf.pages[i].extract_text()
                        if pagina_texto:
                            texto += pagina_texto + "\n"
                    except:
                        continue
                
                if not texto or len(texto.strip()) < 30:
                    return 'PDF_Sem_Texto'
                
                texto_lower = texto.lower()
                
                # ESTADO DE SÃO PAULO
                if 'fazenda do estado' in texto_lower or 'secretaria da fazenda' in texto_lower:
                    return 'Estado_SP_Fazenda'
                
                if 'estado de são paulo' in texto_lower or 'estado de sao paulo' in texto_lower:
                    if 'educação' in texto_lower or 'educacao' in texto_lower:
                        return 'Estado_SP_Educacao'
                    elif 'saúde' in texto_lower or 'saude' in texto_lower:
                        return 'Estado_SP_Saude'
                    elif 'segurança' in texto_lower or 'seguranca' in texto_lower:
                        return 'Estado_SP_Seguranca'
                    elif 'desenvolvimento' in texto_lower:
                        return 'Estado_SP_Desenvolvimento'
                    else:
                        return 'Estado_SP_Outros'
                
                # AUTARQUIAS
                if 'ipesp' in texto_lower:
                    return 'Autarquia_IPESP'
                elif 'spprev' in texto_lower:
                    return 'Autarquia_SPPrev'
                elif 'sabesp' in texto_lower:
                    return 'Autarquia_SABESP'
                
                # UNIVERSIDADES
                if 'usp' in texto_lower or 'universidade de são paulo' in texto_lower:
                    return 'Universidade_USP'
                elif 'unesp' in texto_lower:
                    return 'Universidade_UNESP'
                elif 'unicamp' in texto_lower:
                    return 'Universidade_UNICAMP'
                
                # MUNICÍPIOS
                patterns_municipio = [
                    r'município\s+(?:de\s+)?([A-ZÀÁÂÃÇ][a-zàáâãçéêíóôõú\s]{2,30})',
                    r'prefeitura\s+(?:municipal\s+)?(?:de\s+)?([A-ZÀÁÂÃÇ][a-zàáâãçéêíóôõú\s]{2,30})',
                    r'MUNICÍPIO\s+DE\s+([A-ZÀÁÂÃÇ\s]{2,30})',
                ]
                
                for pattern in patterns_municipio:
                    match = re.search(pattern, texto, re.IGNORECASE)
                    if match:
                        municipio = match.group(1).strip().title()
                        municipio = ' '.join(municipio.split())
                        return f'Municipio_{municipio.replace(" ", "_")}'
                
                # OUTROS
                if len(texto.strip()) > 100:
                    return 'Outros_Orgaos'
                
                return 'Nao_Identificado'
                
        except Exception as e:
            erro_msg = str(e)
            
            if 'EOF' in erro_msg or 'marker' in erro_msg:
                return 'PDF_Corrompido'
            elif 'encrypted' in erro_msg.lower() or 'password' in erro_msg.lower():
                return 'PDF_Protegido'
            else:
                return 'Erro_Leitura'
    
    def reorganizar(self):
        """Reorganiza PDFs"""
        
        print(f"\n📂 Reorganizando com pdfplumber (robusto)...")
        print(f"   De: {self.pasta_origem}")
        print(f"   Para: {self.pasta_destino}")
        
        if not os.path.exists(self.pasta_origem):
            print(f"\n❌ Pasta não encontrada!")
            return False
        
        pdfs = [f for f in os.listdir(self.pasta_origem) if f.endswith('.pdf')]
        total = len(pdfs)
        
        print(f"\n   📄 Total: {total} PDFs")
        print(f"\n🔍 Processando...")
        
        for idx, pdf in enumerate(pdfs, 1):
            if idx % 50 == 0 or idx == 1:
                orgaos = len(self.estatisticas)
                print(f"   📋 {idx}/{total} ({idx/total*100:.1f}%) - {orgaos} categorias", end="\r")
            
            try:
                pdf_path = os.path.join(self.pasta_origem, pdf)
                
                # Extrair órgão
                orgao = self.extrair_orgao_devedor(pdf_path)
                
                # Registrar corrompidos
                if orgao == 'PDF_Corrompido':
                    self.pdfs_corrompidos.append(pdf)
                
                # Criar pasta
                pasta_orgao = os.path.join(self.pasta_destino, orgao)
                if not os.path.exists(pasta_orgao):
                    os.makedirs(pasta_orgao)
                
                # Copiar
                destino = os.path.join(pasta_orgao, pdf)
                
                if os.path.exists(destino):
                    base, ext = os.path.splitext(pdf)
                    contador = 1
                    while os.path.exists(os.path.join(pasta_orgao, f"{base}_{contador}{ext}")):
                        contador += 1
                    destino = os.path.join(pasta_orgao, f"{base}_{contador}{ext}")
                
                shutil.copy2(pdf_path, destino)
                
                # Estatísticas
                self.estatisticas[orgao]['quantidade'] += 1
                self.estatisticas[orgao]['arquivos'].append(pdf)
                
            except Exception as e:
                continue
        
        print(f"\n   ✅ {total} PDFs processados!")
        
        return True
    
    def gerar_relatorio(self):
        """Gera relatório"""
        
        print(f"\n📊 Gerando relatório...")
        
        ordenados = sorted(
            self.estatisticas.items(),
            key=lambda x: x[1]['quantidade'],
            reverse=True
        )
        
        print(f"\n" + "="*80)
        print(f"📊 DISTRIBUIÇÃO POR ÓRGÃO")
        print("="*80)
        
        total = sum(d['quantidade'] for _, d in ordenados)
        
        print(f"\n   #    Orgao                                          Qtd      %")
        print(f"   ---  ------------------------------------------------  ----  ------")
        
        for idx, (orgao, dados) in enumerate(ordenados, 1):
            perc = (dados['quantidade'] / total) * 100
            orgao_exibir = orgao.replace('_', ' ')
            print(f"   {idx:3d}  {orgao_exibir:48s}  {dados['quantidade']:4d}  {perc:5.1f}%")
        
        print(f"\n        TOTAL                                              {total:4d}  100.0%")
        
        # Alertas
        if self.pdfs_corrompidos:
            print(f"\n⚠️  PDFs corrompidos: {len(self.pdfs_corrompidos)}")
        
        # Planilha
        self.criar_planilha(ordenados)
    
    def criar_planilha(self, ordenados):
        """Cria planilha"""
        
        print(f"\n📊 Criando planilha...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Distribuição"
        
        headers = ['#', 'Órgão Devedor', 'Quantidade', '%']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
        
        total = sum(d['quantidade'] for _, d in ordenados)
        
        for idx, (orgao, dados) in enumerate(ordenados, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=orgao.replace('_', ' '))
            ws.cell(row=idx+1, column=3, value=dados['quantidade'])
            
            perc = (dados['quantidade'] / total) * 100
            ws.cell(row=idx+1, column=4, value=f"{perc:.1f}%")
        
        row_total = len(ordenados) + 2
        ws.cell(row=row_total, column=2, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row_total, column=3, value=total).font = Font(bold=True)
        ws.cell(row=row_total, column=4, value='100%').font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        
        ws.row_dimensions[1].height = 25
        ws.auto_filter.ref = f'A1:D{len(ordenados)+1}'
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo = f"distribuicao_final_{timestamp}.xlsx"
        
        wb.save(arquivo)
        
        print(f"   ✅ {arquivo}")

if __name__ == "__main__":
    print("="*80)
    print("📂 REORGANIZAR OFÍCIOS - VERSÃO ROBUSTA (PDFPLUMBER)")
    print("="*80)
    
    reorg = ReorganizadorRobusto()
    
    confirma = input(f"\nReorganizar com pdfplumber? (s/n): ").lower()
    
    if confirma != 's':
        print("\n❌ Cancelado")
        exit()
    
    try:
        inicio = datetime.now()
        
        if reorg.reorganizar():
            reorg.gerar_relatorio()
            
            fim = datetime.now()
            duracao = int((fim - inicio).total_seconds() / 60)
            
            print("\n" + "="*80)
            print("🎉 CONCLUÍDO!")
            print("="*80)
            print(f"\n⏱️  {duracao} min")
            print(f"📂 {reorg.pasta_destino}/")
            print(f"📊 {len(reorg.estatisticas)} categorias")
        
    except Exception as e:
        print(f"\n❌ {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nENTER...\n")
    
    print("\n✅ FIM!")
