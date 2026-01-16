import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import os
import sys
import requests
from datetime import datetime
from collections import Counter, defaultdict

# Importações matemáticas com fallback
try:
    import numpy as np
    from scipy import stats
    NUMPY_DISPONIVEL = True
except ImportError:
    NUMPY_DISPONIVEL = False
    print("⚠️ NumPy/SciPy não disponível")

try:
    from scipy.special import comb
except ImportError:
    import math
    def comb(n, k, exact=True):
        resultado = math.factorial(n) // (math.factorial(k) * math.factorial(n - k))
        return resultado if exact else float(resultado)

# ============================================================
# CONSTANTES
# ============================================================

TOTAL_DEZENAS = 25
TOTAL_COMBINACOES = 3268760

# Configurações de jogos
TIPOS_JOGOS = {
    15: {'dezenas': 15, 'apostas': 1, 'custo': 3.00, 'nome': 'Simples (15 números)'},
    16: {'dezenas': 16, 'apostas': 16, 'custo': 48.00, 'nome': 'Desdobramento (16 números)'},
    17: {'dezenas': 17, 'apostas': 68, 'custo': 204.00, 'nome': 'Desdobramento (17 números)'},
    18: {'dezenas': 18, 'apostas': 204, 'custo': 612.00, 'nome': 'Desdobramento (18 números)'}
}

PREMIOS_MEDIOS = {
    15: 1800000.00,
    14: 2000.00,
    13: 30.00,
    12: 12.00,
    11: 6.00
}

# ============================================================
# SPLASH SCREEN
# ============================================================

def mostrar_splash():
    """Splash screen premium"""
    splash = tk.Tk()
    splash.overrideredirect(True)
    splash.attributes('-topmost', True)
    
    largura = 700
    altura = 450
    x = (splash.winfo_screenwidth() - largura) // 2
    y = (splash.winfo_screenheight() - altura) // 2
    splash.geometry(f"{largura}x{altura}+{x}+{y}")
    
    frame = tk.Frame(splash, bg="#0D1117", bd=0)
    frame.pack(fill="both", expand=True)
    
    header = tk.Frame(frame, bg="#6A0DAD", height=120)
    header.pack(fill="x")
    header.pack_propagate(False)
    
    tk.Label(header, text="🎲", font=("Segoe UI Emoji", 70), bg="#6A0DAD", fg="#FFD700").pack(pady=20)
    
    tk.Label(frame, text="LOTOFÁCIL QUANTUM", font=("Segoe UI", 38, "bold"), bg="#0D1117", fg="#FFD700").pack(pady=20)
    tk.Label(frame, text="Sistema Matemático Hiper-Avançado", font=("Segoe UI", 13), bg="#0D1117", fg="#B0B0B0").pack()
    tk.Label(frame, text="v5.0 ULTIMATE - Desdobramentos 15/16/17/18", font=("Segoe UI", 10, "bold"), bg="#0D1117", fg="#4CAF50").pack(pady=5)
    tk.Label(frame, text="Robinson Tax Master", font=("Segoe UI", 9), bg="#0D1117", fg="#707070").pack()
    
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("Quantum.Horizontal.TProgressbar", troughcolor='#1C1C1C', background='#FFD700', bordercolor='#0D1117', lightcolor='#FFE55C', darkcolor='#FFD700', thickness=22)
    
    progress = ttk.Progressbar(frame, length=600, mode='determinate', style="Quantum.Horizontal.TProgressbar")
    progress.pack(pady=30)
    
    label_status = tk.Label(frame, text="Inicializando...", font=("Segoe UI", 11), bg="#0D1117", fg="#FFFFFF")
    label_status.pack(pady=10)
    
    etapas = [
        (0, 15, "🔬 Inicializando núcleo quântico..."),
        (15, 30, "📊 Carregando algoritmos preditivos..."),
        (30, 50, "🧮 Configurando desdobramentos..."),
        (50, 70, "🎯 Calibrando probabilidades..."),
        (70, 90, "💎 Preparando análise avançada..."),
        (90, 100, "✅ Sistema ULTIMATE ativo!")
    ]
    
    for inicio, fim, texto in etapas:
        label_status.config(text=texto)
        for i in range(inicio, fim + 1):
            progress['value'] = i
            splash.update()
            time.sleep(0.01)
    
    time.sleep(0.4)
    splash.destroy()

# ============================================================
# CALCULADORA DE PROBABILIDADES
# ============================================================

class CalculadoraProbabilidades:
    """Calcula probabilidades para jogos simples e desdobramentos"""
    
    @staticmethod
    def calcular_probabilidade_acertos_simples(acertos):
        """Calcula P(X=k) para jogo simples de 15 números"""
        if acertos < 11 or acertos > 15:
            return 0.0
        
        comb_acertos = comb(15, acertos, exact=True)
        comb_erros = comb(10, 15 - acertos, exact=True)
        numerador = comb_acertos * comb_erros
        
        return numerador / TOTAL_COMBINACOES
    
    @staticmethod
    def calcular_probabilidade_desdobramento(num_dezenas, acertos):
        """
        Calcula probabilidade de ter pelo menos 1 aposta com K acertos
        em um desdobramento de N dezenas
        """
        if acertos < 11 or acertos > 15:
            return 0.0
        
        # Número de apostas no desdobramento
        num_apostas = comb(num_dezenas, 15, exact=True)
        
        # Probabilidade de uma aposta específica acertar exatamente K números
        prob_uma_aposta = CalculadoraProbabilidades.calcular_probabilidade_acertos_simples(acertos)
        
        # Para desdobramentos, calculamos a probabilidade de PELO MENOS uma aposta acertar
        # P(pelo menos 1) = 1 - P(nenhuma acertar)
        # Aproximação: para prêmios maiores (14, 15) é muito improvável ter mais de 1 hit
        
        if acertos >= 14:
            # Para 14 e 15, a chance de ter múltiplos hits é desprezível
            # Então P(pelo menos 1) ≈ num_apostas × P(uma aposta)
            prob_ajustada = min(num_apostas * prob_uma_aposta, 1.0)
        else:
            # Para 11, 12, 13, usamos fórmula mais precisa
            # P(nenhuma) = (1 - p)^n
            prob_nenhuma = (1 - prob_uma_aposta) ** num_apostas
            prob_ajustada = 1 - prob_nenhuma
        
        return prob_ajustada
    
    @staticmethod
    def calcular_todas_probabilidades(jogo, tipo_jogo=15):
        """Calcula todas as probabilidades baseado no tipo de jogo"""
        num_dezenas = len(jogo)
        probs = {}
        
        if tipo_jogo == 15 or num_dezenas == 15:
            # Jogo simples
            for acertos in range(11, 16):
                probs[acertos] = CalculadoraProbabilidades.calcular_probabilidade_acertos_simples(acertos)
        else:
            # Desdobramento
            for acertos in range(11, 16):
                probs[acertos] = CalculadoraProbabilidades.calcular_probabilidade_desdobramento(num_dezenas, acertos)
        
        return probs
    
    @staticmethod
    def calcular_valor_esperado(jogo, tipo_jogo=15):
        """Calcula EV do jogo"""
        probs = CalculadoraProbabilidades.calcular_todas_probabilidades(jogo, tipo_jogo)
        custo = TIPOS_JOGOS[tipo_jogo]['custo']
        
        ev = sum(PREMIOS_MEDIOS.get(acertos, 0) * probs[acertos] for acertos in range(11, 16))
        
        return ev, ev - custo
    
    @staticmethod
    def calcular_roi(jogo, tipo_jogo=15):
        """Calcula ROI"""
        _, ev_liquido = CalculadoraProbabilidades.calcular_valor_esperado(jogo, tipo_jogo)
        custo = TIPOS_JOGOS[tipo_jogo]['custo']
        return (ev_liquido / custo) * 100

# ============================================================
# ANALISADOR MATEMÁTICO
# ============================================================

class AnalisadorMatematico:
    """Análise matemática avançada"""
    
    def __init__(self):
        self.historico = []
        self.matriz_frequencia = [0] * (TOTAL_DEZENAS + 1)
        self.matriz_gaps = defaultdict(list)
        self.matriz_pares = defaultdict(int)
        self.ultima_atualizacao = None
    
    def carregar_historico_completo(self, max_concursos=120):
        """Carrega histórico da API"""
        try:
            url = "https://servicebus2.caixa.gov.br/portaldeloterias/api/lotofacil/"
            headers = {'User-Agent': 'Mozilla/5.0'}
            
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                return False
            
            dados = response.json()
            ultimo = dados.get('numero', 0)
            
            print(f"📥 Baixando {max_concursos} concursos...")
            
            for i in range(min(max_concursos, ultimo)):
                num = ultimo - i
                try:
                    resp = requests.get(f"{url}{num}", headers=headers, timeout=5)
                    if resp.status_code == 200:
                        d = resp.json()
                        dezenas = sorted([int(x) for x in d.get('listaDezenas', [])])
                        self.historico.append({
                            'concurso': num,
                            'data': d.get('dataApuracao'),
                            'dezenas': dezenas
                        })
                        if (i + 1) % 15 == 0:
                            print(f"✓ {i + 1} concursos...")
                except:
                    continue
            
            self.historico.reverse()
            self.ultima_atualizacao = datetime.now()
            self._calcular_estatisticas()
            return True
        except Exception as e:
            print(f"Erro: {e}")
            return False
    
    def _calcular_estatisticas(self):
        """Calcula estatísticas"""
        if not self.historico:
            return
        
        # Frequência
        for r in self.historico:
            for d in r['dezenas']:
                self.matriz_frequencia[d] += 1
        
        # Gaps
        ultima = {i: -1 for i in range(1, TOTAL_DEZENAS + 1)}
        for idx, r in enumerate(self.historico):
            for d in range(1, TOTAL_DEZENAS + 1):
                if d in r['dezenas']:
                    if ultima[d] >= 0:
                        self.matriz_gaps[d].append(idx - ultima[d])
                    ultima[d] = idx
        
        # Pares
        for r in self.historico:
            dez = r['dezenas']
            for i in range(len(dez)):
                for j in range(i + 1, len(dez)):
                    self.matriz_pares[tuple(sorted([dez[i], dez[j]]))] += 1
    
    def calcular_score(self, dezena):
        """Score avançado com 6 critérios"""
        score = 0.0
        
        # 1. Frequência (28%)
        total = sum(self.matriz_frequencia)
        if total > 0:
            freq = self.matriz_frequencia[dezena] / total
            score += (freq / (1/TOTAL_DEZENAS)) * 28
        
        # 2. Tendência recente (25%)
        ultimos = [r['dezenas'] for r in self.historico[-25:]]
        freq_rec = sum(1 for j in ultimos if dezena in j)
        score += (freq_rec / 25) * 100 * 0.25
        
        # 3. Gap analysis (22%)
        if dezena in self.matriz_gaps and self.matriz_gaps[dezena]:
            if NUMPY_DISPONIVEL:
                gap_medio = np.mean(self.matriz_gaps[dezena])
            else:
                gap_medio = sum(self.matriz_gaps[dezena]) / len(self.matriz_gaps[dezena])
            
            gap_atual = len(self.historico) - max([i for i, r in enumerate(self.historico) if dezena in r['dezenas']], default=0)
            if gap_atual > gap_medio:
                score += min((gap_atual / gap_medio) * 22, 32)
        
        # 4. Posição geométrica (12%)
        score += (1 - abs(dezena - 13) / 12) * 12
        
        # 5. Paridade (8%)
        total_jogos = len(self.historico)
        if total_jogos > 0:
            if dezena % 2 == 0:
                count = sum(1 for r in self.historico for d in r['dezenas'] if d % 2 == 0)
            else:
                count = sum(1 for r in self.historico for d in r['dezenas'] if d % 2 != 0)
            score += (count / (total_jogos * 15)) * 8
        
        # 6. Momentum (5%)
        ultimos_5 = [r['dezenas'] for r in self.historico[-5:]]
        aparicoes = sum(1 for j in ultimos_5 if dezena in j)
        score += (aparicoes / 5) * 100 * 0.05
        
        return round(score, 2)
    
    def calcular_scores_todas(self):
        """Scores de todas"""
        return {d: self.calcular_score(d) for d in range(1, TOTAL_DEZENAS + 1)}
    
    def gerar_jogo_otimizado(self, scores, num_dezenas, fixas=None, descartadas=None):
        """Gera jogo otimizado com N dezenas"""
        if fixas is None:
            fixas = set()
        if descartadas is None:
            descartadas = set()
        
        jogo = set(fixas)
        disponiveis = {d: s for d, s in scores.items() if d not in fixas and d not in descartadas}
        ordenadas = sorted(disponiveis.items(), key=lambda x: x[1], reverse=True)
        
        while len(jogo) < num_dezenas:
            melhor = None
            melhor_score = -1
            
            for dez, score_base in ordenadas:
                if dez in jogo:
                    continue
                
                score_total = score_base
                
                # Compatibilidade com dezenas já no jogo
                for d_jogo in jogo:
                    par = tuple(sorted([dez, d_jogo]))
                    score_total += (self.matriz_pares.get(par, 0) / len(self.historico)) * 12
                
                # Balanceamento pares/ímpares (mais flexível para desdobramentos)
                pares = sum(1 for d in jogo if d % 2 == 0)
                if dez % 2 == 0:
                    pares += 1
                prop = pares / (len(jogo) + 1)
                
                # Para desdobramentos, balanceamento é menos crítico
                if num_dezenas == 15:
                    if prop < 0.35 or prop > 0.65:
                        score_total *= 0.75
                else:
                    if prop < 0.25 or prop > 0.75:
                        score_total *= 0.90
                
                if score_total > melhor_score:
                    melhor_score = score_total
                    melhor = dez
            
            if melhor:
                jogo.add(melhor)
            else:
                break
        
        return sorted(list(jogo))
    
    def gerar_multiplos(self, qtd, scores, tipo_jogo, fixas=None, descartadas=None):
        """Gera múltiplos jogos"""
        num_dezenas = TIPOS_JOGOS[tipo_jogo]['dezenas']
        jogos = []
        jogos_set = set()
        
        max_tentativas = qtd * 80
        tentativas = 0
        
        while len(jogos) < qtd and tentativas < max_tentativas:
            tentativas += 1
            
            # Variação controlada nos scores
            diversidade = 0.25 if tipo_jogo == 15 else 0.20
            scores_var = {d: s * random.uniform(1 - diversidade, 1 + diversidade) for d, s in scores.items()}
            
            jogo = self.gerar_jogo_otimizado(scores_var, num_dezenas, fixas, descartadas)
            
            if len(jogo) == num_dezenas:
                t = tuple(jogo)
                if t not in jogos_set:
                    jogos.append(jogo)
                    jogos_set.add(t)
        
        return jogos

# ============================================================
# INSTÂNCIAS GLOBAIS
# ============================================================

analisador = AnalisadorMatematico()
calc_prob = CalculadoraProbabilidades()

# ============================================================
# FUNÇÕES DE INTERFACE
# ============================================================

def atualizar_info_tipo_jogo(*args):
    """Atualiza informações do tipo de jogo selecionado"""
    tipo = combo_tipo.get()
    tipo_num = int(tipo.split()[0])
    
    info = TIPOS_JOGOS[tipo_num]
    
    texto_info_jogo.set(
        f"💰 Custo: R$ {info['custo']:.2f} | "
        f"🎲 {info['apostas']} aposta{'s' if info['apostas'] > 1 else ''} | "
        f"📊 {info['dezenas']} números"
    )

def buscar_analisar():
    """Busca e analisa"""
    try:
        texto_status.set("🔬 Analisando histórico completo...")
        root.update()
        
        if not analisador.carregar_historico_completo(120):
            messagebox.showerror("Erro", "Falha ao carregar histórico.")
            texto_status.set("❌ Erro ao carregar.")
            return
        
        scores = analisador.calcular_scores_todas()
        exibir_analise(scores)
        
        texto_status.set(f"✅ {len(analisador.historico)} concursos analisados com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro na análise:\n{str(e)}")
        texto_status.set("❌ Erro na análise.")

def exibir_analise(scores):
    """Exibe análise completa"""
    janela = tk.Toplevel(root)
    janela.title("🔬 Análise Quantum Completa")
    janela.geometry("900x650")
    
    # Título
    frame_titulo = tk.Frame(janela, bg="#6A0DAD", height=60)
    frame_titulo.pack(fill="x")
    frame_titulo.pack_propagate(False)
    
    tk.Label(
        frame_titulo,
        text="🏆 RANKING QUANTUM DE DEZENAS",
        font=("Segoe UI", 16, "bold"),
        bg="#6A0DAD",
        fg="white"
    ).pack(pady=15)
    
    # Frame principal
    frame_main = tk.Frame(janela, bg="white")
    frame_main.pack(fill="both", expand=True, padx=15, pady=15)
    
    # Info
    frame_info = tk.Frame(frame_main, bg="#F0F0F0", relief="solid", bd=1)
    frame_info.pack(fill="x", pady=(0, 10))
    
    tk.Label(
        frame_info,
        text=f"📊 Base: {len(analisador.historico)} concursos | "
             f"🕐 Atualizado: {analisador.ultima_atualizacao.strftime('%d/%m/%Y %H:%M')}",
        font=("Segoe UI", 9),
        bg="#F0F0F0",
        fg="#333"
    ).pack(pady=8)
    
    # Text widget
    text = scrolledtext.ScrolledText(frame_main, font=("Courier New", 9), bg="#FAFAFA", wrap="word", relief="sunken", bd=2)
    text.pack(fill="both", expand=True)
    
    ranking = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    
    # Cabeçalho
    text.insert("end", f"{'#':<5} {'DEZ':<7} {'SCORE':<12} {'VISUAL':<38} {'STATUS':<18} {'FREQ'}\n", "header")
    text.insert("end", "="*100 + "\n\n", "separator")
    
    for idx, (dez, score) in enumerate(ranking, 1):
        barra = "█" * int((score / 100) * 32)
        freq = analisador.matriz_frequencia[dez]
        
        if score >= 80:
            status = "🔥 ULTRA QUENTE"
            tag = "ultra"
        elif score >= 70:
            status = "🌡️ MUITO QUENTE"
            tag = "muito_quente"
        elif score >= 60:
            status = "🔸 QUENTE"
            tag = "quente"
        elif score >= 45:
            status = "➡️ NEUTRO"
            tag = "neutro"
        elif score >= 30:
            status = "❄️ FRIO"
            tag = "frio"
        else:
            status = "🧊 CONGELADO"
            tag = "congelado"
        
        linha = f"{idx:<5} {dez:02d}{'':<5} {score:<12.2f} {barra:<38} {status:<18} {freq}x\n"
        text.insert("end", linha, tag)
    
    # Tags
    text.tag_config("header", foreground="#6A0DAD", font=("Courier New", 10, "bold"))
    text.tag_config("separator", foreground="#CCCCCC")
    text.tag_config("ultra", foreground="#FF0000", font=("Courier New", 9, "bold"))
    text.tag_config("muito_quente", foreground="#FF5722", font=("Courier New", 9, "bold"))
    text.tag_config("quente", foreground="#FF9800", font=("Courier New", 9))
    text.tag_config("neutro", foreground="#4CAF50", font=("Courier New", 9))
    text.tag_config("frio", foreground="#2196F3", font=("Courier New", 9))
    text.tag_config("congelado", foreground="#9E9E9E", font=("Courier New", 9))
    
    text.config(state="disabled")
    
    # Botões
    frame_botoes = tk.Frame(frame_main, bg="white")
    frame_botoes.pack(pady=10)
    
    tk.Button(
        frame_botoes,
        text="🎯 Usar Top 8 Dezenas",
        command=lambda: usar_top([d for d, s in ranking[:8]]),
        font=("Segoe UI", 11, "bold"),
        bg="#4CAF50",
        fg="white",
        padx=20,
        pady=10,
        cursor="hand2",
        relief="raised",
        bd=2
    ).pack(side="left", padx=5)
    
    tk.Button(
        frame_botoes,
        text="📋 Usar Top 10 Dezenas",
        command=lambda: usar_top([d for d, s in ranking[:10]]),
        font=("Segoe UI", 11, "bold"),
        bg="#2196F3",
        fg="white",
        padx=20,
        pady=10,
        cursor="hand2",
        relief="raised",
        bd=2
    ).pack(side="left", padx=5)
    
    tk.Button(
        frame_botoes,
        text="❌ Fechar",
        command=janela.destroy,
        font=("Segoe UI", 10),
        bg="#FF5722",
        fg="white",
        padx=20,
        pady=10,
        cursor="hand2",
        relief="raised",
        bd=2
    ).pack(side="left", padx=5)

def usar_top(dezenas):
    """Usa top dezenas"""
    entry_fixas.delete(0, "end")
    entry_fixas.insert(0, " ".join(map(str, dezenas)))
    messagebox.showinfo(
        "✅ Dezenas Aplicadas",
        f"As {len(dezenas)} dezenas com maior score foram configuradas!\n\n"
        f"Dezenas: {', '.join(map(str, dezenas))}\n\n"
        f"Agora clique em 'Gerar Jogos Quantum' para criar seus jogos."
    )
    texto_status.set(f"✨ Top {len(dezenas)} dezenas configuradas!")

def gerar_jogos():
    """Gera jogos quantum"""
    try:
        if not analisador.historico:
            if messagebox.askyesno(
                "Análise Necessária",
                "Para gerar jogos otimizados, é necessário carregar o histórico.\n\n"
                "Deseja carregar agora? (~40 segundos)"
            ):
                buscar_analisar()
                return
            else:
                messagebox.showinfo("Info", "Gerando sem otimização avançada.")
                return
        
        # Validar quantidade
        qtd = int(entry_qtd.get())
        if qtd <= 0 or qtd > 50:
            messagebox.showerror("Erro", "Quantidade deve estar entre 1 e 50.")
            return
        
        # Tipo de jogo
        tipo_str = combo_tipo.get()
        tipo_jogo = int(tipo_str.split()[0])
        
        # Validar fixas e descartadas
        fixas = set()
        if entry_fixas.get().strip():
            try:
                fixas = set(map(int, entry_fixas.get().split()))
                if not all(1 <= d <= 25 for d in fixas):
                    messagebox.showerror("Erro", "Dezenas fixas devem estar entre 1 e 25.")
                    return
                if len(fixas) > TIPOS_JOGOS[tipo_jogo]['dezenas']:
                    messagebox.showerror("Erro", f"Máximo de {TIPOS_JOGOS[tipo_jogo]['dezenas']} dezenas fixas para este tipo de jogo.")
                    return
            except ValueError:
                messagebox.showerror("Erro", "Formato inválido nas dezenas fixas. Use: 1 5 12")
                return
        
        descartadas = set()
        if entry_descartadas.get().strip():
            try:
                descartadas = set(map(int, entry_descartadas.get().split()))
                if not all(1 <= d <= 25 for d in descartadas):
                    messagebox.showerror("Erro", "Dezenas descartadas devem estar entre 1 e 25.")
                    return
            except ValueError:
                messagebox.showerror("Erro", "Formato inválido nas dezenas descartadas. Use: 2 7 20")
                return
        
        # Verificar conflito
        if fixas & descartadas:
            messagebox.showerror("Erro", "Há dezenas que são fixas e descartadas ao mesmo tempo!")
            return
        
        # Gerar
        texto_status.set(f"🎲 Gerando {qtd} jogos quantum de {tipo_jogo} números...")
        root.update()
        
        scores = analisador.calcular_scores_todas()
        jogos = analisador.gerar_multiplos(qtd, scores, tipo_jogo, fixas, descartadas)
        
        if not jogos:
            messagebox.showwarning("Aviso", "Não foi possível gerar jogos com os critérios especificados.")
            texto_status.set("❌ Nenhum jogo gerado.")
            return
        
        # Calcular probabilidades
        jogos_info = []
        for jogo in jogos:
            probs = calc_prob.calcular_todas_probabilidades(jogo, tipo_jogo)
            ev, ev_liq = calc_prob.calcular_valor_esperado(jogo, tipo_jogo)
            roi = calc_prob.calcular_roi(jogo, tipo_jogo)
            qualidade = sum(scores.get(d, 50) for d in jogo) / len(jogo)
            
            jogos_info.append({
                'jogo': jogo,
                'tipo': tipo_jogo,
                'probs': probs,
                'ev': ev,
                'ev_liq': ev_liq,
                'roi': roi,
                'qualidade': qualidade
            })
        
        lista_jogos.clear()
        lista_jogos.extend(jogos_info)
        atualizar_preview()
        
        # Estatísticas gerais
        custo_total = len(jogos) * TIPOS_JOGOS[tipo_jogo]['custo']
        apostas_totais = len(jogos) * TIPOS_JOGOS[tipo_jogo]['apostas']
        
        texto_status.set(f"✅ {len(jogos)} jogos gerados | Custo total: R$ {custo_total:,.2f}")
        
        messagebox.showinfo(
            "✅ Sucesso Quantum",
            f"🎉 {len(jogos)} jogos quantum gerados!\n\n"
            f"📊 Tipo: {TIPOS_JOGOS[tipo_jogo]['nome']}\n"
            f"🎲 Total de apostas: {apostas_totais:,}\n"
            f"💰 Custo total: R$ {custo_total:,.2f}\n"
            f"🔬 Método: Algoritmo Quantum Hiper-Avançado\n"
            f"📈 Base: {len(analisador.historico)} concursos\n"
            f"🎯 Análise probabilística individual incluída!"
        )
        
    except ValueError:
        messagebox.showerror("Erro", "Preencha a quantidade com um número válido.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao gerar jogos:\n{str(e)}")

def atualizar_preview():
    """Atualiza preview com probabilidades"""
    text_preview.config(state="normal")
    text_preview.delete("1.0", "end")
    
    if not lista_jogos:
        text_preview.insert("1.0", 
            "═══════════════════════════════════════════════════════\n"
            "        Nenhum jogo gerado ainda\n\n"
            "  1. Clique em '🔬 Análise Quantum' (recomendado)\n"
            "  2. Configure o tipo de jogo (15/16/17/18 números)\n"
            "  3. Clique em '🎲 Gerar Jogos Quantum'\n"
            "═══════════════════════════════════════════════════════"
        )
        text_preview.config(state="disabled")
        return
    
    # Cabeçalho
    text_preview.insert("end", "="*130 + "\n", "sep")
    text_preview.insert("end", "                    JOGOS QUANTUM GERADOS - ANÁLISE PROBABILÍSTICA COMPLETA\n", "header")
    text_preview.insert("end", "="*130 + "\n\n", "sep")
    
    # Info geral
    primeiro_jogo = lista_jogos[0]
    tipo = primeiro_jogo['tipo']
    info_tipo = TIPOS_JOGOS[tipo]
    
    text_preview.insert("end", f"📊 Tipo: {info_tipo['nome']} | ", "info")
    text_preview.insert("end", f"💰 Custo unitário: R$ {info_tipo['custo']:.2f} | ", "info")
    text_preview.insert("end", f"🎲 {info_tipo['apostas']} aposta(s) por jogo\n", "info")
    text_preview.insert("end", f"💵 Custo total: R$ {len(lista_jogos) * info_tipo['custo']:,.2f}\n\n", "info_destaque")
    text_preview.insert("end", "="*130 + "\n\n", "sep")
    
    # Jogos
    preview_count = min(20, len(lista_jogos))
    
    for i, info in enumerate(lista_jogos[:preview_count], 1):
        jogo = info['jogo']
        probs = info['probs']
        roi = info['roi']
        qual = info['qualidade']
        
        dez_str = ' - '.join(f'{d:02d}' for d in jogo)
        pares = sum(1 for d in jogo if d % 2 == 0)
        impares = len(jogo) - pares
        soma = sum(jogo)
        
        # Linha do jogo
        text_preview.insert("end", f"Jogo {i:02d}: {dez_str}\n", "jogo")
        
        # Estatísticas
        text_preview.insert("end", 
            f"         Pares: {pares} | Ímpares: {impares} | Soma: {soma} | "
            f"Qualidade: {qual:.1f}/100\n", 
            "stats"
        )
        
        # Probabilidades
        text_preview.insert("end", "         Probabilidades: ", "label")
        text_preview.insert("end", f"15pts: {probs[15]*100:.6f}% ", "prob15")
        text_preview.insert("end", f"| 14pts: {probs[14]*100:.4f}% ", "prob14")
        text_preview.insert("end", f"| 13pts: {probs[13]*100:.3f}% ", "prob13")
        text_preview.insert("end", f"| 12pts: {probs[12]*100:.2f}% ", "prob12")
        text_preview.insert("end", f"| 11pts: {probs[11]*100:.2f}%\n", "prob11")
        
        # ROI e EV
        roi_tag = "roi_positivo" if roi > -50 else "roi_negativo"
        text_preview.insert("end", 
            f"         ROI: {roi:+.2f}% | EV Líquido: R$ {info['ev_liq']:+.2f}\n", 
            roi_tag
        )
        
        text_preview.insert("end", "\n", "normal")
    
    if len(lista_jogos) > preview_count:
        text_preview.insert("end", 
            f"\n... e mais {len(lista_jogos) - preview_count} jogos com análise completa.\n", 
            "info_extra"
        )
    
    # Configurar tags
    text_preview.tag_config("header", foreground="#6A0DAD", font=("Courier New", 11, "bold"))
    text_preview.tag_config("sep", foreground="#CCCCCC")
    text_preview.tag_config("info", foreground="#333333", font=("Courier New", 9))
    text_preview.tag_config("info_destaque", foreground="#4CAF50", font=("Courier New", 9, "bold"))
    text_preview.tag_config("jogo", foreground="#2196F3", font=("Courier New", 10, "bold"))
    text_preview.tag_config("stats", foreground="#4CAF50", font=("Courier New", 9))
    text_preview.tag_config("label", foreground="#000000", font=("Courier New", 9))
    text_preview.tag_config("prob15", foreground="#FFD700", font=("Courier New", 9, "bold"))
    text_preview.tag_config("prob14", foreground="#FF5722", font=("Courier New", 9, "bold"))
    text_preview.tag_config("prob13", foreground="#FF9800", font=("Courier New", 9))
    text_preview.tag_config("prob12", foreground="#4CAF50", font=("Courier New", 9))
    text_preview.tag_config("prob11", foreground="#2196F3", font=("Courier New", 9))
    text_preview.tag_config("roi_positivo", foreground="#4CAF50", font=("Courier New", 9, "bold"))
    text_preview.tag_config("roi_negativo", foreground="#FF5722", font=("Courier New", 9))
    text_preview.tag_config("info_extra", foreground="#757575", font=("Courier New", 9, "italic"))
    
    text_preview.config(state="disabled")

def salvar_excel():
    """Salva Excel com probabilidades"""
    if not lista_jogos:
        messagebox.showwarning("Aviso", "Nenhum jogo para salvar.")
        return
    
    try:
        caminho = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"Lotofacil_Quantum_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not caminho:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Jogos Quantum"
        
        # Info do tipo de jogo
        primeiro = lista_jogos[0]
        tipo = primeiro['tipo']
        info_tipo = TIPOS_JOGOS[tipo]
        
        # Cabeçalhos
        max_dez = info_tipo['dezenas']
        headers = ["#"] + [f"D{i}" for i in range(1, max_dez + 1)] + [
            "Pares", "Ímpares", "Soma", "Qualidade", 
            "P(15)", "P(14)", "P(13)", "P(12)", "P(11)", 
            "EV", "EV Líq", "ROI%", "Custo"
        ]
        ws.append(headers)
        
        # Estilo cabeçalho
        header_fill = PatternFill(start_color="6A0DAD", end_color="6A0DAD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        for i, info in enumerate(lista_jogos, start=1):
            jogo = info['jogo']
            probs = info['probs']
            
            # Preencher jogo com espaços vazios se necessário
            jogo_completo = jogo + [None] * (max_dez - len(jogo))
            
            pares = sum(1 for d in jogo if d % 2 == 0)
            impares = len(jogo) - pares
            soma = sum(jogo)
            
            linha = [i] + jogo_completo + [
                pares,
                impares,
                soma,
                round(info['qualidade'], 2),
                f"{probs[15]*100:.8f}%",
                f"{probs[14]*100:.6f}%",
                f"{probs[13]*100:.4f}%",
                f"{probs[12]*100:.3f}%",
                f"{probs[11]*100:.3f}%",
                f"R$ {info['ev']:.2f}",
                f"R$ {info['ev_liq']:+.2f}",
                f"{info['roi']:+.2f}%",
                f"R$ {info_tipo['custo']:.2f}"
            ]
            ws.append(linha)
        
        # Linha de resumo
        ws.append([])
        custo_total = len(lista_jogos) * info_tipo['custo']
        apostas_totais = len(lista_jogos) * info_tipo['apostas']
        
        resumo_linha = ws.max_row + 1
        ws[f"A{resumo_linha}"] = "RESUMO"
        ws[f"A{resumo_linha}"].font = Font(bold=True, size=11)
        
        ws[f"B{resumo_linha}"] = f"Total de jogos: {len(lista_jogos)}"
        ws[f"D{resumo_linha}"] = f"Total de apostas: {apostas_totais:,}"
        ws[f"F{resumo_linha}"] = f"Custo total: R$ {custo_total:,.2f}"
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 18)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(caminho)
        messagebox.showinfo(
            "✅ Sucesso",
            f"Excel Quantum salvo com sucesso!\n\n"
            f"📁 {caminho}\n\n"
            f"📊 {len(lista_jogos)} jogos exportados"
        )
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar Excel:\n{str(e)}")

def salvar_pdf():
    """Salva PDF com probabilidades"""
    if not lista_jogos:
        messagebox.showwarning("Aviso", "Nenhum jogo para salvar.")
        return
    
    try:
        caminho = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=f"Lotofacil_Quantum_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        )
        if not caminho:
            return
        
        c = canvas.Canvas(caminho, pagesize=A4)
        w, h = A4
        y = h - 40
        
        # Título
        c.setFont("Helvetica-Bold", 18)
        c.drawString(50, y, "JOGOS QUANTUM - LOTOFÁCIL")
        y -= 22
        
        # Info
        primeiro = lista_jogos[0]
        tipo = primeiro['tipo']
        info_tipo = TIPOS_JOGOS[tipo]
        custo_total = len(lista_jogos) * info_tipo['custo']
        
        c.setFont("Helvetica", 9)
        c.drawString(50, y, f"Tipo: {info_tipo['nome']} | Total: {len(lista_jogos)} jogos | Custo: R$ {custo_total:,.2f}")
        y -= 15
        c.drawString(50, y, f"Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 25
        
        # Jogos
        c.setFont("Courier", 7)
        
        for i, info in enumerate(lista_jogos, 1):
            jogo = info['jogo']
            probs = info['probs']
            
            dez_str = ' - '.join(f'{d:02d}' for d in jogo)
            pares = sum(1 for d in jogo if d % 2 == 0)
            soma = sum(jogo)
            
            # Jogo
            c.drawString(50, y, f"Jogo {i:02d}: {dez_str}")
            y -= 10
            
            # Stats e probs
            c.drawString(65, y, 
                f"P:{pares} I:{len(jogo)-pares} S:{soma} Q:{info['qualidade']:.1f} | "
                f"P(15):{probs[15]*100:.6f}% P(14):{probs[14]*100:.4f}% "
                f"P(13):{probs[13]*100:.3f}% | ROI:{info['roi']:+.1f}%"
            )
            y -= 15
            
            if y < 60:
                c.showPage()
                y = h - 40
                c.setFont("Courier", 7)
        
        c.save()
        messagebox.showinfo(
            "✅ Sucesso",
            f"PDF Quantum salvo com sucesso!\n\n"
            f"📁 {caminho}\n\n"
            f"📄 {len(lista_jogos)} jogos exportados"
        )
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao salvar PDF:\n{str(e)}")

def limpar():
    """Limpa tudo"""
    if lista_jogos and not messagebox.askyesno("Confirmar", "Deseja realmente limpar todos os jogos?"):
        return
    lista_jogos.clear()
    atualizar_preview()
    texto_status.set("✨ Sistema pronto para nova geração!")

# ============================================================
# INTERFACE PRINCIPAL
# ============================================================

def criar_interface():
    """Interface principal"""
    global root, entry_qtd, combo_tipo, entry_fixas, entry_descartadas
    global texto_status, texto_info_jogo, text_preview, lista_jogos
    
    lista_jogos = []
    
    root = tk.Tk()
    root.title("🎲 Lotofácil QUANTUM v5.0 ULTIMATE - Desdobramentos 15/16/17/18")
    root.geometry("1200x820")
    root.configure(bg="#F5F5F5")
    
    try:
        if os.path.exists("icone_trevo.ico"):
            root.iconbitmap("icone_trevo.ico")
    except:
        pass
    
    # Frame principal
    main = tk.Frame(root, bg="#F5F5F5")
    main.pack(fill="both", expand=True, padx=10, pady=10)
    
    # ===== CONFIGURAÇÕES =====
    config = tk.LabelFrame(
        main,
        text="⚙️ Configuração Quantum",
        font=("Segoe UI", 12, "bold"),
        bg="white",
        fg="#6A0DAD",
        padx=20,
        pady=20,
        relief="solid",
        bd=2
    )
    config.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
    
    # Tipo de jogo
    tk.Label(config, text="🎲 Tipo de jogo:", bg="white", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", pady=8)
    
    tipos_opcoes = [f"{k} números - {v['nome']}" for k, v in TIPOS_JOGOS.items()]
    combo_tipo = ttk.Combobox(config, values=tipos_opcoes, state="readonly", font=("Segoe UI", 10), width=32)
    combo_tipo.set(tipos_opcoes[0])
    combo_tipo.grid(row=0, column=1, pady=8, padx=5, sticky="ew")
    combo_tipo.bind("<<ComboboxSelected>>", atualizar_info_tipo_jogo)
    
    # Info do tipo
    texto_info_jogo = tk.StringVar()
    label_info = tk.Label(
        config,
        textvariable=texto_info_jogo,
        bg="#F0F0F0",
        font=("Segoe UI", 9),
        fg="#333",
        relief="solid",
        bd=1,
        padx=8,
        pady=5
    )
    label_info.grid(row=1, column=0, columnspan=2, pady=8, sticky="ew")
    atualizar_info_tipo_jogo()
    
    # Quantidade
    tk.Label(config, text="📊 Quantidade de jogos:", bg="white", font=("Segoe UI", 10, "bold")).grid(row=2, column=0, sticky="w", pady=8)
    entry_qtd = tk.Entry(config, font=("Segoe UI", 11), width=15, relief="solid", bd=2)
    entry_qtd.insert(0, "10")
    entry_qtd.grid(row=2, column=1, pady=8, padx=5, sticky="w")
    
    # Fixas
    tk.Label(config, text="🎯 Dezenas fixas:", bg="white", font=("Segoe UI", 10, "bold")).grid(row=3, column=0, sticky="w", pady=8)
    tk.Label(config, text="(ex: 1 7 13 21)", bg="white", font=("Segoe UI", 8), fg="gray").grid(row=3, column=1, sticky="w")
    entry_fixas = tk.Entry(config, font=("Segoe UI", 10), relief="solid", bd=2)
    entry_fixas.grid(row=4, column=0, columnspan=2, pady=5, padx=5, sticky="ew")
    
    # Descartadas
    tk.Label(config, text="🚫 Dezenas descartadas:", bg="white", font=("Segoe UI", 10, "bold")).grid(row=5, column=0, sticky="w", pady=8)
    tk.Label(config, text="(ex: 2 8 24)", bg="white", font=("Segoe UI", 8), fg="gray").grid(row=5, column=1, sticky="w")
    entry_descartadas = tk.Entry(config, font=("Segoe UI", 10), relief="solid", bd=2)
    entry_descartadas.grid(row=6, column=0, columnspan=2, pady=5, padx=5, sticky="ew")
    
    # Botões
    botoes = tk.Frame(config, bg="white")
    botoes.grid(row=7, column=0, columnspan=2, pady=20)
    
    tk.Button(
        botoes,
        text="🔬 Análise Quantum Completa",
        command=buscar_analisar,
        bg="#FF9800",
        fg="white",
        font=("Segoe UI", 11, "bold"),
        padx=15,
        pady=12,
        cursor="hand2",
        relief="raised",
        bd=3
    ).pack(fill="x", pady=4)
    
    tk.Button(
        botoes,
        text="🎲 Gerar Jogos Quantum",
        command=gerar_jogos,
        bg="#6A0DAD",
        fg="white",
        font=("Segoe UI", 12, "bold"),
        padx=20,
        pady=14,
        cursor="hand2",
        relief="raised",
        bd=3
    ).pack(fill="x", pady=4)
    
    tk.Button(
        botoes,
        text="🗑️ Limpar Tudo",
        command=limpar,
        bg="#FF6B6B",
        fg="white",
        font=("Segoe UI", 10),
        padx=20,
        pady=10,
        cursor="hand2",
        relief="raised",
        bd=2
    ).pack(fill="x", pady=4)
    
    tk.Button(
        botoes,
        text="📊 Salvar Excel",
        command=salvar_excel,
        bg="#4CAF50",
        fg="white",
        font=("Segoe UI", 10),
        padx=20,
        pady=10,
        cursor="hand2",
        relief="raised",
        bd=2
    ).pack(fill="x", pady=4)
    
    tk.Button(
        botoes,
        text="📄 Salvar PDF",
        command=salvar_pdf,
        bg="#2196F3",
        fg="white",
        font=("Segoe UI", 10),
        padx=20,
        pady=10,
        cursor="hand2",
        relief="raised",
        bd=2
    ).pack(fill="x", pady=4)
    
    # ===== PREVIEW =====
    preview = tk.LabelFrame(
        main,
        text="📋 Jogos Quantum Gerados (Análise Probabilística Completa)",
        font=("Segoe UI", 12, "bold"),
        bg="white",
        fg="#6A0DAD",
        padx=20,
        pady=20,
        relief="solid",
        bd=2
    )
    preview.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
    
    text_preview = scrolledtext.ScrolledText(
        preview,
        font=("Courier New", 8),
        bg="#FAFAFA",
        wrap="none",
        relief="sunken",
        bd=2
    )
    text_preview.pack(fill="both", expand=True)
    atualizar_preview()
    
    # ===== STATUS =====
    status_frame = tk.Frame(main, bg="#1C1C1C", bd=2, relief="raised")
    status_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(12, 0))
    
    texto_status = tk.StringVar()
    texto_status.set("✨ Sistema QUANTUM pronto! Clique em '🔬 Análise Quantum' para começar.")
    
    tk.Label(
        status_frame,
        textvariable=texto_status,
        font=("Segoe UI", 11, "bold"),
        bg="#1C1C1C",
        fg="#FFD700",
        padx=20,
        pady=10
    ).pack(fill="x")
    
    # Grid weights
    main.columnconfigure(0, weight=1)
    main.columnconfigure(1, weight=3)
    main.rowconfigure(0, weight=1)
    
    root.mainloop()

# ============================================================
# EXECUÇÃO
# ============================================================

if __name__ == "__main__":
    mostrar_splash()
    criar_interface()