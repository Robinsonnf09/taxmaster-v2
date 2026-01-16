import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import time
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import os
import sys
import requests
from datetime import datetime
from collections import Counter, defaultdict
import numpy as np
from scipy import stats
from scipy.special import comb
import json
import pickle
from itertools import combinations
import math

# ============================================================
# CONSTANTES MATEMÁTICAS
# ============================================================

TOTAL_DEZENAS = 25
DEZENAS_POR_JOGO = 15
TOTAL_COMBINACOES = 3268760  # C(25,15)

# Valores dos prêmios (médias históricas em R$)
PREMIOS_MEDIOS = {
    15: 1800000.00,  # 15 acertos
    14: 2000.00,     # 14 acertos
    13: 30.00,       # 13 acertos
    12: 12.00,       # 12 acertos
    11: 6.00         # 11 acertos
}

CUSTO_JOGO = 3.00  # Custo de um jogo simples

# ============================================================
# SPLASH SCREEN PREMIUM
# ============================================================

def mostrar_splash():
    """Splash screen profissional"""
    splash = tk.Tk()
    splash.overrideredirect(True)
    splash.attributes('-topmost', True)
    
    largura_splash = 650
    altura_splash = 420
    largura_tela = splash.winfo_screenwidth()
    altura_tela = splash.winfo_screenheight()
    x = (largura_tela - largura_splash) // 2
    y = (altura_tela - altura_splash) // 2
    splash.geometry(f"{largura_splash}x{altura_splash}+{x}+{y}")
    
    frame_principal = tk.Frame(splash, bg="#0D1117", bd=0)
    frame_principal.pack(fill="both", expand=True)
    
    header = tk.Frame(frame_principal, bg="#6A0DAD", height=110)
    header.pack(fill="x")
    header.pack_propagate(False)
    
    label_icone = tk.Label(header, text="🎲", font=("Segoe UI Emoji", 65), bg="#6A0DAD", fg="#FFD700")
    label_icone.pack(pady=15)
    
    label_titulo = tk.Label(frame_principal, text="LOTOFÁCIL QUANTUM", font=("Segoe UI", 36, "bold"), bg="#0D1117", fg="#FFD700")
    label_titulo.pack(pady=20)
    
    label_subtitulo = tk.Label(frame_principal, text="Sistema de Predição Matemática Hiper-Avançada", font=("Segoe UI", 12), bg="#0D1117", fg="#B0B0B0")
    label_subtitulo.pack()
    
    label_versao = tk.Label(frame_principal, text="v4.0 QUANTUM - Análise Probabilística Individual", font=("Segoe UI", 9), bg="#0D1117", fg="#707070")
    label_versao.pack(pady=5)
    
    style = ttk.Style()
    style.theme_use('clam')
    style.configure("Quantum.Horizontal.TProgressbar", troughcolor='#1C1C1C', background='#FFD700', bordercolor='#0D1117', lightcolor='#FFE55C', darkcolor='#FFD700', thickness=20)
    
    progress = ttk.Progressbar(frame_principal, length=550, mode='determinate', style="Quantum.Horizontal.TProgressbar")
    progress.pack(pady=30)
    
    label_status = tk.Label(frame_principal, text="Inicializando...", font=("Segoe UI", 11), bg="#0D1117", fg="#FFFFFF")
    label_status.pack(pady=10)
    
    def animar_progresso():
        etapas = [
            (0, 15, "🔬 Inicializando núcleo quântico..."),
            (15, 30, "📊 Carregando algoritmos preditivos..."),
            (30, 50, "🧮 Configurando cálculo probabilístico..."),
            (50, 70, "🎯 Preparando simulação Monte Carlo..."),
            (70, 90, "💎 Calibrando análise de valor esperado..."),
            (90, 100, "✅ Sistema Quantum ativo!")
        ]
        
        for inicio, fim, texto in etapas:
            label_status.config(text=texto)
            for i in range(inicio, fim + 1):
                progress['value'] = i
                splash.update()
                time.sleep(0.01)
        
        time.sleep(0.4)
        splash.destroy()
    
    animar_progresso()

# ============================================================
# CLASSE DE CÁLCULO DE PROBABILIDADES
# ============================================================

class CalculadoraProbabilidades:
    """Calcula probabilidades individuais para cada jogo"""
    
    @staticmethod
    def calcular_probabilidade_acertos(jogo_apostado, acertos):
        """
        Calcula a probabilidade de acertar exatamente N números
        
        Fórmula: P(X=k) = [C(15,k) * C(10, 15-k)] / C(25,15)
        
        Onde:
        - C(15,k) = combinações de k acertos entre as 15 dezenas apostadas
        - C(10, 15-k) = combinações de (15-k) acertos entre as 10 dezenas não apostadas
        - C(25,15) = total de combinações possíveis
        """
        
        # Dezenas apostadas: 15
        # Dezenas sorteadas: 15
        # Dezenas não apostadas: 10
        
        if acertos < 11 or acertos > 15:
            return 0.0
        
        # C(15, acertos) - maneiras de acertar 'acertos' números das 15 apostadas
        comb_acertos = comb(15, acertos, exact=True)
        
        # C(10, 15-acertos) - maneiras de pegar (15-acertos) números das 10 não apostadas
        comb_erros = comb(10, 15 - acertos, exact=True)
        
        # Total de combinações
        numerador = comb_acertos * comb_erros
        denominador = TOTAL_COMBINACOES
        
        probabilidade = numerador / denominador
        
        return probabilidade
    
    @staticmethod
    def calcular_todas_probabilidades(jogo):
        """Calcula probabilidades de 11 a 15 acertos"""
        probs = {}
        for acertos in range(11, 16):
            probs[acertos] = CalculadoraProbabilidades.calcular_probabilidade_acertos(jogo, acertos)
        return probs
    
    @staticmethod
    def calcular_valor_esperado(jogo, premios=PREMIOS_MEDIOS, custo=CUSTO_JOGO):
        """Calcula o valor esperado (Expected Value) do jogo"""
        probs = CalculadoraProbabilidades.calcular_todas_probabilidades(jogo)
        
        ev = 0.0
        for acertos in range(11, 16):
            premio = premios.get(acertos, 0)
            prob = probs[acertos]
            ev += premio * prob
        
        # Subtrair o custo
        ev_liquido = ev - custo
        
        return ev, ev_liquido
    
    @staticmethod
    def calcular_roi(jogo, premios=PREMIOS_MEDIOS, custo=CUSTO_JOGO):
        """Calcula o ROI (Return on Investment) esperado"""
        ev, ev_liquido = CalculadoraProbabilidades.calcular_valor_esperado(jogo, premios, custo)
        roi = (ev_liquido / custo) * 100
        return roi
    
    @staticmethod
    def calcular_indice_qualidade(jogo, scores_dezenas):
        """Calcula índice de qualidade do jogo baseado nos scores"""
        if not scores_dezenas:
            return 50.0
        
        score_medio = np.mean([scores_dezenas.get(d, 50) for d in jogo])
        return score_medio

# ============================================================
# CLASSE DE ANÁLISE MATEMÁTICA AVANÇADA
# ============================================================

class AnalisadorMatematico:
    """Análise matemática e estatística hiper-avançada"""
    
    def __init__(self):
        self.historico = []
        self.matriz_frequencia = np.zeros(TOTAL_DEZENAS + 1)
        self.matriz_gaps = defaultdict(list)
        self.matriz_pares = defaultdict(int)
        self.matriz_sequencias = defaultdict(int)
        self.padroes_geometricos = defaultdict(int)
        self.ultima_atualizacao = None
        
    def carregar_historico_completo(self, max_concursos=150):
        """Carrega histórico completo"""
        try:
            url = "https://servicebus2.caixa.gov.br/portaldeloterias/api/lotofacil/"
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
            
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code != 200:
                return False
            
            dados = response.json()
            ultimo_concurso = dados.get('numero', 0)
            
            print(f"📥 Baixando últimos {max_concursos} concursos...")
            
            for i in range(min(max_concursos, ultimo_concurso)):
                concurso_num = ultimo_concurso - i
                
                try:
                    url_concurso = f"https://servicebus2.caixa.gov.br/portaldeloterias/api/lotofacil/{concurso_num}"
                    resp = requests.get(url_concurso, headers=headers, timeout=5)
                    
                    if resp.status_code == 200:
                        dados_concurso = resp.json()
                        dezenas = sorted([int(d) for d in dados_concurso.get('listaDezenas', [])])
                        
                        self.historico.append({
                            'concurso': concurso_num,
                            'data': dados_concurso.get('dataApuracao'),
                            'dezenas': dezenas
                        })
                        
                        if (i + 1) % 10 == 0:
                            print(f"✓ {i + 1} concursos carregados...")
                        
                except Exception as e:
                    continue
            
            self.historico.reverse()
            self.ultima_atualizacao = datetime.now()
            self._calcular_estatisticas_avancadas()
            return True
            
        except Exception as e:
            print(f"Erro: {e}")
            return False
    
    def _calcular_estatisticas_avancadas(self):
        """Calcula estatísticas avançadas"""
        if not self.historico:
            return
        
        # 1. Frequência absoluta
        for resultado in self.historico:
            for dezena in resultado['dezenas']:
                self.matriz_frequencia[dezena] += 1
        
        # 2. Análise de gaps
        ultima_aparicao = {i: -1 for i in range(1, TOTAL_DEZENAS + 1)}
        
        for idx, resultado in enumerate(self.historico):
            for dezena in range(1, TOTAL_DEZENAS + 1):
                if dezena in resultado['dezenas']:
                    if ultima_aparicao[dezena] >= 0:
                        gap = idx - ultima_aparicao[dezena]
                        self.matriz_gaps[dezena].append(gap)
                    ultima_aparicao[dezena] = idx
        
        # 3. Análise de pares (correlações)
        for resultado in self.historico:
            dezenas = resultado['dezenas']
            for i in range(len(dezenas)):
                for j in range(i + 1, len(dezenas)):
                    par = tuple(sorted([dezenas[i], dezenas[j]]))
                    self.matriz_pares[par] += 1
        
        # 4. Análise de sequências
        for resultado in self.historico:
            dezenas = sorted(resultado['dezenas'])
            for i in range(len(dezenas) - 1):
                if dezenas[i+1] - dezenas[i] == 1:
                    seq = (dezenas[i], dezenas[i+1])
                    self.matriz_sequencias[seq] += 1
        
        # 5. Padrões geométricos (linhas, colunas, diagonais no volante 5x5)
        for resultado in self.historico:
            padrao = self._identificar_padrao_geometrico(resultado['dezenas'])
            self.padroes_geometricos[padrao] += 1
    
    def _identificar_padrao_geometrico(self, dezenas):
        """Identifica padrão geométrico no volante 5x5"""
        # Volante: 
        # 01 02 03 04 05
        # 06 07 08 09 10
        # 11 12 13 14 15
        # 16 17 18 19 20
        # 21 22 23 24 25
        
        linhas = [0] * 5
        colunas = [0] * 5
        
        for dez in dezenas:
            linha = (dez - 1) // 5
            coluna = (dez - 1) % 5
            linhas[linha] += 1
            colunas[coluna] += 1
        
        # Classificar padrão
        max_linha = max(linhas)
        max_coluna = max(colunas)
        
        if max_linha >= 5:
            return "linha_forte"
        elif max_coluna >= 5:
            return "coluna_forte"
        elif max_linha >= 4 and max_coluna >= 4:
            return "cruz"
        else:
            return "disperso"
    
    def calcular_score_hiperavancado(self, dezena):
        """Score hiper-avançado com 8 critérios ponderados"""
        score = 0.0
        
        # 1. FREQUÊNCIA HISTÓRICA (20%)
        freq_total = np.sum(self.matriz_frequencia)
        if freq_total > 0:
            freq_relativa = self.matriz_frequencia[dezena] / freq_total
            freq_esperada = 1 / TOTAL_DEZENAS
            score_freq = (freq_relativa / freq_esperada) * 20
            score += score_freq
        
        # 2. TENDÊNCIA RECENTE (18%)
        ultimos_20 = [r['dezenas'] for r in self.historico[-20:]]
        freq_recente = sum(1 for jogo in ultimos_20 if dezena in jogo)
        score_tendencia = (freq_recente / 20) * 100 * 0.18
        score += score_tendencia
        
        # 3. ANÁLISE DE GAPS (17%)
        if dezena in self.matriz_gaps and self.matriz_gaps[dezena]:
            gaps = self.matriz_gaps[dezena]
            gap_medio = np.mean(gaps)
            gap_atual = len(self.historico) - max([idx for idx, r in enumerate(self.historico) if dezena in r['dezenas']], default=0)
            
            if gap_atual > gap_medio:
                score_gap = min((gap_atual / gap_medio) * 17, 25)
                score += score_gap
        
        # 4. DISTRIBUIÇÃO ESTATÍSTICA (15%)
        if dezena in self.matriz_gaps and len(self.matriz_gaps[dezena]) > 5:
            gaps = self.matriz_gaps[dezena]
            cv = np.std(gaps) / np.mean(gaps) if np.mean(gaps) > 0 else 0
            score_dist = (1 - min(cv, 1)) * 15
            score += score_dist
        
        # 5. POSIÇÃO GEOMÉTRICA (10%)
        # Centro do volante (13) tem vantagem
        distancia_centro = abs(dezena - 13)
        score_posicao = (1 - distancia_centro / 12) * 10
        score += score_posicao
        
        # 6. PARIDADE (10%)
        total_jogos = len(self.historico)
        if total_jogos > 0:
            if dezena % 2 == 0:
                pares_historico = sum(1 for r in self.historico for d in r['dezenas'] if d % 2 == 0)
                proporcao = pares_historico / (total_jogos * DEZENAS_POR_JOGO)
            else:
                impares_historico = sum(1 for r in self.historico for d in r['dezenas'] if d % 2 != 0)
                proporcao = impares_historico / (total_jogos * DEZENAS_POR_JOGO)
            
            score_paridade = proporcao * 10
            score += score_paridade
        
        # 7. PARTICIPAÇÃO EM SEQUÊNCIAS (5%)
        seq_count = 0
        for seq, freq in self.matriz_sequencias.items():
            if dezena in seq:
                seq_count += freq
        score_seq = min((seq_count / total_jogos) * 100 * 0.05, 5)
        score += score_seq
        
        # 8. ANÁLISE DE MOMENTUM (5%)
        ultimos_5 = [r['dezenas'] for r in self.historico[-5:]]
        aparicoes_recentes = sum(1 for jogo in ultimos_5 if dezena in jogo)
        score_momentum = (aparicoes_recentes / 5) * 100 * 0.05
        score += score_momentum
        
        return round(score, 2)
    
    def calcular_scores_todas_dezenas(self):
        """Calcula scores para todas"""
        scores = {}
        for dezena in range(1, TOTAL_DEZENAS + 1):
            scores[dezena] = self.calcular_score_hiperavancado(dezena)
        return scores
    
    def calcular_compatibilidade_par(self, dez1, dez2):
        """Compatibilidade entre duas dezenas"""
        par = tuple(sorted([dez1, dez2]))
        freq = self.matriz_pares.get(par, 0)
        total_jogos = len(self.historico)
        
        if total_jogos == 0:
            return 0.5
        
        return freq / total_jogos
    
    def gerar_jogo_quantum(self, scores, fixas=None, descartadas=None):
        """Gera jogo usando algoritmo quântico"""
        if fixas is None:
            fixas = set()
        if descartadas is None:
            descartadas = set()
        
        jogo = set(fixas)
        disponiveis = {d: s for d, s in scores.items() if d not in fixas and d not in descartadas}
        
        dezenas_ordenadas = sorted(disponiveis.items(), key=lambda x: x[1], reverse=True)
        
        while len(jogo) < DEZENAS_POR_JOGO:
            melhor_dezena = None
            melhor_score_total = -1
            
            for dezena, score_base in dezenas_ordenadas:
                if dezena in jogo:
                    continue
                
                score_total = score_base
                
                # Compatibilidade com dezenas no jogo
                for d_jogo in jogo:
                    compat = self.calcular_compatibilidade_par(dezena, d_jogo)
                    score_total += compat * 15
                
                # Balanceamento pares/ímpares
                pares_atuais = sum(1 for d in jogo if d % 2 == 0)
                if dezena % 2 == 0:
                    pares_projetados = pares_atuais + 1
                else:
                    pares_projetados = pares_atuais
                
                dezenas_restantes = DEZENAS_POR_JOGO - len(jogo)
                if dezenas_restantes > 0:
                    proporcao_pares = pares_projetados / (len(jogo) + 1)
                    if proporcao_pares < 0.35 or proporcao_pares > 0.65:
                        score_total *= 0.75
                
                # Evitar muitas sequências
                sequencias = 0
                jogo_sorted = sorted(list(jogo))
                for d in jogo_sorted:
                    if dezena == d + 1 or dezena == d - 1:
                        sequencias += 1
                
                if sequencias > 2:
                    score_total *= 0.85
                
                if score_total > melhor_score_total:
                    melhor_score_total = score_total
                    melhor_dezena = dezena
            
            if melhor_dezena:
                jogo.add(melhor_dezena)
            else:
                break
        
        return sorted(list(jogo))
    
    def gerar_multiplos_jogos_quantum(self, quantidade, scores, fixas=None, descartadas=None, diversidade=0.25):
        """Gera múltiplos jogos quantum"""
        jogos = []
        jogos_set = set()
        
        tentativas = 0
        max_tentativas = quantidade * 150
        
        while len(jogos) < quantidade and tentativas < max_tentativas:
            tentativas += 1
            
            scores_variados = {}
            for dezena, score in scores.items():
                variacao = random.uniform(1 - diversidade, 1 + diversidade)
                scores_variados[dezena] = score * variacao
            
            jogo = self.gerar_jogo_quantum(scores_variados, fixas, descartadas)
            
            jogo_tuple = tuple(jogo)
            if jogo_tuple not in jogos_set and len(jogo) == DEZENAS_POR_JOGO:
                jogos.append(jogo)
                jogos_set.add(jogo_tuple)
        
        return jogos

# ============================================================
# INSTÂNCIA GLOBAL
# ============================================================

analisador = AnalisadorMatematico()
calc_prob = CalculadoraProbabilidades()

# ============================================================
# FUNÇÕES DE INTERFACE
# ============================================================

def buscar_e_analisar():
    """Busca e analisa histórico"""
    try:
        texto_status.set("🔬 Iniciando análise quantum completa...")
        root.update()
        
        sucesso = analisador.carregar_historico_completo(max_concursos=150)
        
        if not sucesso:
            messagebox.showerror("Erro", "Não foi possível carregar o histórico.")
            texto_status.set("❌ Erro ao carregar histórico.")
            return
        
        scores = analisador.calcular_scores_todas_dezenas()
        exibir_analise_quantum(scores)
        
        texto_status.set(f"✅ Análise Quantum completa! {len(analisador.historico)} concursos processados.")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro:\n{str(e)}")
        texto_status.set("❌ Erro na análise.")

def exibir_analise_quantum(scores):
    """Exibe análise quantum"""
    janela = tk.Toplevel(root)
    janela.title("🔬 Análise Quantum Completa")
    janela.geometry("950x750")
    
    notebook = ttk.Notebook(janela)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
    # ===== ABA: RANKING =====
    frame_ranking = tk.Frame(notebook, bg="white")
    notebook.add(frame_ranking, text="🏆 Ranking Quantum")
    
    tk.Label(frame_ranking, text="🎯 Ranking Preditivo Quantum", font=("Segoe UI", 15, "bold"), bg="white", fg="#6A0DAD").pack(pady=15)
    
    text_ranking = scrolledtext.ScrolledText(frame_ranking, font=("Courier New", 9), bg="#F5F5F5", wrap="word")
    text_ranking.pack(fill="both", expand=True, padx=15, pady=10)
    
    ranking = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    
    text_ranking.insert(tk.END, f"{'#':<4} {'DEZ':<6} {'SCORE':<10} {'VISUAL':<35} {'STATUS':<20} {'FREQ'}\n", "header")
    text_ranking.insert(tk.END, "="*95 + "\n\n", "separator")
    
    for idx, (dezena, score) in enumerate(ranking, 1):
        barra_tamanho = int((score / 100) * 30)
        barra = "█" * barra_tamanho
        
        freq = int(analisador.matriz_frequencia[dezena])
        
        if score >= 80:
            status = "🔥 ULTRA QUENTE"
            tag = "ultra"
        elif score >= 70:
            status = "🌡️ MUITO QUENTE"
            tag = "muito_quente"
        elif score >= 60:
            status = "🔸 QUENTE"
            tag = "quente"
        elif score >= 45:
            status = "➡️ NEUTRO"
            tag = "neutro"
        elif score >= 30:
            status = "❄️ FRIO"
            tag = "frio"
        else:
            status = "🧊 CONGELADO"
            tag = "congelado"
        
        linha = f"{idx:<4} {dezena:02d}{'':<4} {score:<10.2f} {barra:<35} {status:<20} {freq}x\n"
        text_ranking.insert(tk.END, linha, tag)
    
    text_ranking.tag_config("header", foreground="#6A0DAD", font=("Courier New", 10, "bold"))
    text_ranking.tag_config("separator", foreground="#CCCCCC")
    text_ranking.tag_config("ultra", foreground="#FF0000", font=("Courier New", 9, "bold"))
    text_ranking.tag_config("muito_quente", foreground="#FF5722", font=("Courier New", 9, "bold"))
    text_ranking.tag_config("quente", foreground="#FF9800")
    text_ranking.tag_config("neutro", foreground="#4CAF50")
    text_ranking.tag_config("frio", foreground="#2196F3")
    text_ranking.tag_config("congelado", foreground="#9E9E9E")
    
    text_ranking.config(state="disabled")
    
    # ===== ABA: PROBABILIDADES =====
    frame_prob = tk.Frame(notebook, bg="white")
    notebook.add(frame_prob, text="📊 Probabilidades")
    
    tk.Label(frame_prob, text="🎲 Tabela de Probabilidades", font=("Segoe UI", 15, "bold"), bg="white", fg="#6A0DAD").pack(pady=15)
    
    text_prob = scrolledtext.ScrolledText(frame_prob, font=("Courier New", 10), bg="#F5F5F5", wrap="word")
    text_prob.pack(fill="both", expand=True, padx=15, pady=10)
    
    text_prob.insert(tk.END, "="*80 + "\n", "sep")
    text_prob.insert(tk.END, "PROBABILIDADES MATEMÁTICAS EXATAS\n", "titulo")
    text_prob.insert(tk.END, "="*80 + "\n\n", "sep")
    
    jogo_exemplo = list(range(1, 16))
    probs_exemplo = calc_prob.calcular_todas_probabilidades(jogo_exemplo)
    
    text_prob.insert(tk.END, f"{'ACERTOS':<12} {'PROBABILIDADE':<25} {'1 EM':<20} {'PRÊMIO MÉDIO'}\n", "header")
    text_prob.insert(tk.END, "-"*80 + "\n", "sep")
    
    for acertos in range(15, 10, -1):
        prob = probs_exemplo[acertos]
        prob_perc = prob * 100
        um_em = int(1 / prob) if prob > 0 else 0
        premio = PREMIOS_MEDIOS.get(acertos, 0)
        
        linha = f"{acertos} pontos{'':<3} {prob_perc:.8f}%{'':<10} 1 em {um_em:,}{'':<8} R$ {premio:,.2f}\n"
        
        if acertos == 15:
            text_prob.insert(tk.END, linha, "acerto15")
        elif acertos == 14:
            text_prob.insert(tk.END, linha, "acerto14")
        elif acertos == 13:
            text_prob.insert(tk.END, linha, "acerto13")
        else:
            text_prob.insert(tk.END, linha, "normal")
    
    text_prob.insert(tk.END, "\n" + "="*80 + "\n", "sep")
    text_prob.insert(tk.END, "VALOR ESPERADO E ROI\n", "titulo")
    text_prob.insert(tk.END, "="*80 + "\n\n", "sep")
    
    ev, ev_liq = calc_prob.calcular_valor_esperado(jogo_exemplo)
    roi = calc_prob.calcular_roi(jogo_exemplo)
    
    text_prob.insert(tk.END, f"Valor Esperado (EV): R$ {ev:.2f}\n")
    text_prob.insert(tk.END, f"EV Líquido (após custo): R$ {ev_liq:.2f}\n")
    text_prob.insert(tk.END, f"ROI Esperado: {roi:.2f}%\n\n")
    
    text_prob.insert(tk.END, "⚠️ NOTA: Estes valores são baseados em prêmios médios históricos.\n", "nota")
    text_prob.insert(tk.END, "Os prêmios reais variam conforme arrecadação e número de ganhadores.\n", "nota")
    
    text_prob.tag_config("titulo", foreground="#6A0DAD", font=("Courier New", 11, "bold"))
    text_prob.tag_config("header", foreground="#6A0DAD", font=("Courier New", 10, "bold"))
    text_prob.tag_config("sep", foreground="#CCCCCC")
    text_prob.tag_config("acerto15", foreground="#FFD700", font=("Courier New", 10, "bold"))
    text_prob.tag_config("acerto14", foreground="#FF5722", font=("Courier New", 10, "bold"))
    text_prob.tag_config("acerto13", foreground="#4CAF50", font=("Courier New", 10, "bold"))
    text_prob.tag_config("nota", foreground="#FF5722", font=("Courier New", 9, "italic"))
    
    text_prob.config(state="disabled")
    
    # Botão usar
    btn_usar = tk.Button(
        frame_ranking,
        text="🎯 Aplicar Top 8 Dezenas",
        command=lambda: aplicar_top_dezenas([d for d, s in ranking[:8]]),
        font=("Segoe UI", 11, "bold"),
        bg="#4CAF50",
        fg="white",
        padx=20,
        pady=10,
        cursor="hand2"
    )
    btn_usar.pack(pady=10)

def aplicar_top_dezenas(dezenas):
    """Aplica top dezenas"""
    entry_fixas.delete(0, tk.END)
    entry_fixas.insert(0, " ".join([str(d) for d in dezenas]))
    
    messagebox.showinfo(
        "Top Dezenas Aplicadas",
        f"✅ As {len(dezenas)} dezenas de maior score foram configuradas!\n\n"
        f"Dezenas: {', '.join([str(d) for d in dezenas])}\n\n"
        f"Clique em 'Gerar Jogos Quantum' para criar."
    )
    
    texto_status.set("✨ Top dezenas aplicadas! Pronto para gerar.")

def gerar_jogos_quantum():
    """Gera jogos quantum com análise probabilística"""
    try:
        if not analisador.historico:
            resposta = messagebox.askyesno(
                "Análise Necessária",
                "Para gerar jogos quantum, é necessário análise completa.\n\n"
                "Carregar agora? (≈40 segundos)"
            )
            if resposta:
                buscar_e_analisar()
                return
            else:
                messagebox.showinfo("Info", "Gerando com método padrão.")
                return
        
        qtd = int(entry_qtd.get())
        
        if qtd <= 0 or qtd > 100:
            messagebox.showerror("Erro", "Quantidade: 1-100 para modo quantum.")
            return
        
        fixas = set()
        if entry_fixas.get().strip():
            fixas = set(map(int, entry_fixas.get().split()))
        
        descartadas = set()
        if entry_descartadas.get().strip():
            descartadas = set(map(int, entry_descartadas.get().split()))
        
        texto_status.set("🔬 Gerando jogos quantum...")
        root.update()
        
        scores = analisador.calcular_scores_todas_dezenas()
        
        jogos = analisador.gerar_multiplos_jogos_quantum(
            quantidade=qtd,
            scores=scores,
            fixas=fixas,
            descartadas=descartadas,
            diversidade=0.2
        )
        
        if not jogos:
            messagebox.showwarning("Aviso", "Não foi possível gerar jogos.")
            return
        
        # Calcular probabilidades para cada jogo
        jogos_com_prob = []
        for jogo in jogos:
            probs = calc_prob.calcular_todas_probabilidades(jogo)
            ev, ev_liq = calc_prob.calcular_valor_esperado(jogo)
            roi = calc_prob.calcular_roi(jogo)
            qualidade = calc_prob.calcular_indice_qualidade(jogo, scores)
            
            jogos_com_prob.append({
                'jogo': jogo,
                'probabilidades': probs,
                'ev': ev,
                'ev_liquido': ev_liq,
                'roi': roi,
                'qualidade': qualidade
            })
        
        lista_jogos.clear()
        lista_jogos.extend(jogos_com_prob)
        atualizar_preview_quantum()
        
        texto_status.set(f"✅ {len(jogos)} jogos quantum gerados!")
        
        messagebox.showinfo(
            "Sucesso Quantum",
            f"✅ {len(jogos)} jogos quantum gerados!\n\n"
            f"🔬 Método: Algoritmo Quantum Hiper-Avançado\n"
            f"📊 Base: {len(analisador.historico)} concursos\n"
            f"🎯 Análise probabilística individual incluída"
        )
        
    except ValueError:
        messagebox.showerror("Erro", "Preencha quantidade válida.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro:\n{str(e)}")

def atualizar_preview_quantum():
    """Atualiza preview com probabilidades"""
    text_preview.config(state="normal")
    text_preview.delete(1.0, tk.END)
    
    if not lista_jogos:
        text_preview.insert(1.0, "Nenhum jogo gerado.\n\nClique em 'Gerar Jogos Quantum'.")
        text_preview.config(state="disabled")
        return
    
    # Cabeçalho
    text_preview.insert(tk.END, "="*120 + "\n", "sep")
    text_preview.insert(tk.END, "JOGOS QUANTUM GERADOS - ANÁLISE PROBABILÍSTICA INDIVIDUAL\n", "header")
    text_preview.insert(tk.END, "="*120 + "\n\n", "sep")
    
    preview_count = min(20, len(lista_jogos))
    
    for i, jogo_info in enumerate(lista_jogos[:preview_count], start=1):
        jogo = jogo_info['jogo']
        probs = jogo_info['probabilidades']
        roi = jogo_info['roi']
        qualidade = jogo_info['qualidade']
        
        # Linha do jogo
        dezenas_str = ' - '.join(f'{d:02d}' for d in jogo)
        pares = sum(1 for d in jogo if d % 2 == 0)
        soma = sum(jogo)
        
        text_preview.insert(tk.END, f"Jogo {i:03d}: {dezenas_str}\n", "jogo")
        text_preview.insert(tk.END, f"          Pares: {pares} | Ímpares: {15-pares} | Soma: {soma} | Qualidade: {qualidade:.1f}/100\n", "stats")
        
        # Probabilidades
        text_preview.insert(tk.END, f"          Probabilidades: ", "label")
        text_preview.insert(tk.END, f"15pts: {probs[15]*100:.6f}% ", "prob15")
        text_preview.insert(tk.END, f"| 14pts: {probs[14]*100:.4f}% ", "prob14")
        text_preview.insert(tk.END, f"| 13pts: {probs[13]*100:.3f}% ", "prob13")
        text_preview.insert(tk.END, f"| 12pts: {probs[12]*100:.2f}% ", "prob12")
        text_preview.insert(tk.END, f"| 11pts: {probs[11]*100:.2f}%\n", "prob11")
        
        # ROI
        roi_cor = "roi_positivo" if roi > -50 else "roi_negativo"
        text_preview.insert(tk.END, f"          ROI Esperado: {roi:+.2f}%\n", roi_cor)
        text_preview.insert(tk.END, "\n", "normal")
    
    if len(lista_jogos) > preview_count:
        text_preview.insert(tk.END, f"\n... e mais {len(lista_jogos) - preview_count} jogos com análise completa.\n", "info")
    
    # Configurar tags
    text_preview.tag_config("header", foreground="#6A0DAD", font=("Courier New", 11, "bold"))
    text_preview.tag_config("sep", foreground="#CCCCCC")
    text_preview.tag_config("jogo", foreground="#2196F3", font=("Courier New", 10, "bold"))
    text_preview.tag_config("stats", foreground="#4CAF50", font=("Courier New", 9))
    text_preview.tag_config("label", foreground="#000000", font=("Courier New", 9))
    text_preview.tag_config("prob15", foreground="#FFD700", font=("Courier New", 9, "bold"))
    text_preview.tag_config("prob14", foreground="#FF5722", font=("Courier New", 9, "bold"))
    text_preview.tag_config("prob13", foreground="#FF9800", font=("Courier New", 9))
    text_preview.tag_config("prob12", foreground="#4CAF50", font=("Courier New", 9))
    text_preview.tag_config("prob11", foreground="#2196F3", font=("Courier New", 9))
    text_preview.tag_config("roi_positivo", foreground="#4CAF50", font=("Courier New", 9, "bold"))
    text_preview.tag_config("roi_negativo", foreground="#FF5722", font=("Courier New", 9))
    text_preview.tag_config("info", foreground="#757575", font=("Courier New", 9, "italic"))
    
    text_preview.config(state="disabled")

def salvar_excel_quantum():
    """Salva Excel com probabilidades"""
    if not lista_jogos:
        messagebox.showwarning("Aviso", "Nenhum jogo para salvar.")
        return
    
    try:
        caminho = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not caminho:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Jogos Quantum"
        
        # Cabeçalhos
        headers = ["#"] + [f"D{i}" for i in range(1, 16)] + ["Pares", "Soma", "Qualidade", 
                   "P(15)", "P(14)", "P(13)", "P(12)", "P(11)", "ROI%"]
        ws.append(headers)
        
        # Estilo cabeçalho
        header_fill = PatternFill(start_color="6A0DAD", end_color="6A0DAD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        for i, jogo_info in enumerate(lista_jogos, start=1):
            jogo = jogo_info['jogo']
            probs = jogo_info['probabilidades']
            roi = jogo_info['roi']
            qualidade = jogo_info['qualidade']
            
            pares = sum(1 for d in jogo if d % 2 == 0)
            soma = sum(jogo)
            
            linha = [i] + jogo + [
                pares, 
                soma, 
                round(qualidade, 2),
                f"{probs[15]*100:.8f}%",
                f"{probs[14]*100:.6f}%",
                f"{probs[13]*100:.4f}%",
                f"{probs[12]*100:.3f}%",
                f"{probs[11]*100:.3f}%",
                f"{roi:+.2f}%"
            ]
            ws.append(linha)
        
        # Auto-ajustar colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(caminho)
        messagebox.showinfo("Sucesso", f"✅ Excel Quantum salvo!\n\n📁 {caminho}")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro:\n{str(e)}")

def salvar_pdf_quantum():
    """Salva PDF com probabilidades"""
    if not lista_jogos:
        messagebox.showwarning("Aviso", "Nenhum jogo para salvar.")
        return
    
    try:
        caminho = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if not caminho:
            return
        
        c = canvas.Canvas(caminho, pagesize=A4)
        largura, altura = A4
        y = altura - 40
        
        # Título
        c.setFont("Helvetica-Bold", 18)
        c.drawString(50, y, "JOGOS QUANTUM - ANÁLISE PROBABILÍSTICA")
        y -= 20
        
        c.setFont("Helvetica", 9)
        c.drawString(50, y, f"Total: {len(lista_jogos)} jogos | Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        y -= 25
        
        c.setFont("Courier", 7)
        
        for i, jogo_info in enumerate(lista_jogos, start=1):
            jogo = jogo_info['jogo']
            probs = jogo_info['probabilidades']
            roi = jogo_info['roi']
            qualidade = jogo_info['qualidade']
            
            dezenas_str = ' - '.join(f'{d:02d}' for d in jogo)
            pares = sum(1 for d in jogo if d % 2 == 0)
            soma = sum(jogo)
            
            # Jogo
            c.drawString(50, y, f"Jogo {i:03d}: {dezenas_str}")
            y -= 10
            
            # Stats
            c.drawString(70, y, f"P:{pares} I:{15-pares} S:{soma} Q:{qualidade:.1f} | P(15):{probs[15]*100:.6f}% P(14):{probs[14]*100:.4f}% P(13):{probs[13]*100:.3f}% | ROI:{roi:+.1f}%")
            y -= 15
            
            if y < 50:
                c.showPage()
                y = altura - 40
                c.setFont("Courier", 7)
        
        c.save()
        messagebox.showinfo("Sucesso", f"✅ PDF Quantum salvo!\n\n📁 {caminho}")
        
    except Exception as e:
        messagebox.showerror("Erro", f"Erro:\n{str(e)}")

def limpar_tudo():
    """Limpa tudo"""
    if lista_jogos and not messagebox.askyesno("Confirmar", "Limpar?"):
        return
    lista_jogos.clear()
    atualizar_preview_quantum()
    texto_status.set("Sistema pronto.")

# ============================================================
# INTERFACE PRINCIPAL
# ============================================================

def criar_interface():
    """Interface principal"""
    global root, entry_qtd, entry_fixas, entry_descartadas
    global texto_status, text_preview, lista_jogos
    
    lista_jogos = []
    
    root = tk.Tk()
    root.title("🎲 Lotofácil QUANTUM - Sistema Matemático Hiper-Avançado v4.0")
    root.geometry("1100x800")
    
    try:
        if os.path.exists("icone_trevo.ico"):
            root.iconbitmap("icone_trevo.ico")
    except:
        pass
    
    main_frame = tk.Frame(root, bg="#F0F0F0")
    main_frame.pack(fill="both", expand=True, padx=10, pady=10)
    
    # Frame config
    frame_config = tk.LabelFrame(main_frame, text="⚙️ Configuração Quantum", font=("Segoe UI", 12, "bold"), bg="white", padx=15, pady=15)
    frame_config.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
    
    tk.Label(frame_config, text="Quantidade de jogos:", bg="white", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=8)
    entry_qtd = tk.Entry(frame_config, font=("Segoe UI", 11), width=15)
    entry_qtd.insert(0, "10")
    entry_qtd.grid(row=0, column=1, pady=8, padx=5)
    
    tk.Label(frame_config, text="Dezenas fixas:", bg="white", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=8)
    tk.Label(frame_config, text="(ex: 1 7 13 21)", bg="white", font=("Segoe UI", 8), fg="gray").grid(row=1, column=1, sticky="w")
    entry_fixas = tk.Entry(frame_config, font=("Segoe UI", 10), width=30)
    entry_fixas.grid(row=2, column=0, columnspan=2, pady=5, padx=5, sticky="ew")
    
    tk.Label(frame_config, text="Dezenas descartadas:", bg="white", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="w", pady=8)
    tk.Label(frame_config, text="(ex: 2 8 24)", bg="white", font=("Segoe UI", 8), fg="gray").grid(row=3, column=1, sticky="w")
    entry_descartadas = tk.Entry(frame_config, font=("Segoe UI", 10), width=30)
    entry_descartadas.grid(row=4, column=0, columnspan=2, pady=5, padx=5, sticky="ew")
    
    frame_botoes = tk.Frame(frame_config, bg="white")
    frame_botoes.grid(row=5, column=0, columnspan=2, pady=20)
    
    tk.Button(frame_botoes, text="🔬 Análise Quantum Completa", command=buscar_e_analisar, font=("Segoe UI", 11, "bold"), bg="#FF9800", fg="white", padx=15, pady=12, cursor="hand2", relief="raised", bd=3).pack(fill="x", pady=5)
    tk.Button(frame_botoes, text="🎲 Gerar Jogos Quantum", command=gerar_jogos_quantum, font=("Segoe UI", 12, "bold"), bg="#6A0DAD", fg="white", padx=20, pady=12, cursor="hand2", relief="raised", bd=3).pack(fill="x", pady=5)
    tk.Button(frame_botoes, text="🗑️ Limpar Tudo", command=limpar_tudo, font=("Segoe UI", 10), bg="#FF6B6B", fg="white", padx=20, pady=10, cursor="hand2", relief="raised", bd=2).pack(fill="x", pady=5)
    tk.Button(frame_botoes, text="📊 Salvar Excel Quantum", command=salvar_excel_quantum, font=("Segoe UI", 10), bg="#4CAF50", fg="white", padx=20, pady=10, cursor="hand2", relief="raised", bd=2).pack(fill="x", pady=5)
    tk.Button(frame_botoes, text="📄 Salvar PDF Quantum", command=salvar_pdf_quantum, font=("Segoe UI", 10), bg="#2196F3", fg="white", padx=20, pady=10, cursor="hand2", relief="raised", bd=2).pack(fill="x", pady=5)
    
    # Frame preview
    frame_preview = tk.LabelFrame(main_frame, text="📋 Jogos Quantum Gerados (com Análise Probabilística)", font=("Segoe UI", 12, "bold"), bg="white", padx=15, pady=15)
    frame_preview.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
    
    text_preview = scrolledtext.ScrolledText(frame_preview, font=("Courier New", 8), bg="#FAFAFA", wrap="none", relief="sunken", bd=2)
    text_preview.pack(fill="both", expand=True)
    atualizar_preview_quantum()
    
    # Status
    frame_status = tk.Frame(main_frame, bg="#1C1C1C", relief="raised", bd=2)
    frame_status.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))
    
    texto_status = tk.StringVar()
    texto_status.set("✨ Sistema Quantum pronto! Clique em 'Análise Quantum Completa' para começar.")
    
    tk.Label(frame_status, textvariable=texto_status, font=("Segoe UI", 11, "bold"), bg="#1C1C1C", fg="#FFD700", anchor="w", padx=15, pady=8).pack(fill="x")
    
    main_frame.columnconfigure(0, weight=1)
    main_frame.columnconfigure(1, weight=3)
    main_frame.rowconfigure(0, weight=1)
    
    root.mainloop()

# ============================================================
# EXECUÇÃO
# ============================================================

if __name__ == "__main__":
    mostrar_splash()
    criar_interface()