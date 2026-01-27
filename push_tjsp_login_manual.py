"""
PUSH TJSP - COM LOGIN MANUAL COMPLETO
Você faz TODO o login manualmente, depois script cadastra
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from datetime import datetime
import os

class PushTJSPLoginManual:
    
    def __init__(self):
        self.driver = None
        self.url_push = "https://esaj.tjsp.jus.br/push/index.do"
        self.sucessos = []
        self.falhas = []
        self.ja_cadastrados = []
    
    def iniciar(self):
        print("\n🌐 Iniciando Chrome...")
        
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--ignore-certificate-errors')
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        print("✅ Chrome iniciado!")
    
    def login_manual_completo(self):
        print(f"\n🔐 Acessando PUSH TJSP...")
        self.driver.get(self.url_push)
        time.sleep(3)
        
        print("\n" + "="*70)
        print("👉 LOGIN MANUAL COMPLETO:")
        print("="*70)
        print("   1. Faça login com certificado OU login/senha")
        print("   2. Se pedir Web Signer:")
        print("      - Clique em CANCELAR")
        print("      - Use login/senha ao invés de certificado")
        print("   3. NAVEGUE até a área de CADASTRO de processos")
        print("   4. Você deve ver:")
        print("      - Campo para digitar número do processo")
        print("      - Botão 'Incluir'")
        print("   5. Volte aqui e pressione ENTER")
        print("="*70)
        
        input("\n>>> ENTER quando estiver na tela de CADASTRO <<<\n")
        
        url_atual = self.driver.current_url
        print(f"✅ URL atual: {url_atual}")
        print("✅ Pronto para cadastrar!")
        
        return True
    
    def cadastrar_processo(self, numero):
        try:
            print(f"\n📝 Cadastrando: {numero}")
            
            wait = WebDriverWait(self.driver, 10)
            
            # Campo de processo
            try:
                campo = wait.until(EC.presence_of_element_located(
                    (By.XPATH, "//input[@type='text'][1]")
                ))
                
                if campo.is_displayed():
                    campo.clear()
                    time.sleep(0.2)
                    campo.send_keys(numero)
                    time.sleep(0.3)
                    print(f"   ✅ Número digitado")
                else:
                    raise Exception("Campo não visível")
                
            except:
                print(f"   ⚠️  Campo não encontrado")
                print(f"   💡 Digite manualmente: {numero}")
                input(f"   >>> ENTER após digitar <<<\n")
            
            # Botão Incluir
            try:
                botoes = [
                    "//button[contains(text(), 'Incluir')]",
                    "//input[@value='Incluir']",
                    "//button[@type='submit']",
                    "//input[@type='submit']"
                ]
                
                btn = None
                for xpath in botoes:
                    try:
                        btn = self.driver.find_element(By.XPATH, xpath)
                        if btn.is_displayed():
                            break
                    except:
                        continue
                
                if btn:
                    btn.click()
                    print(f"   ⏳ Processando...")
                    time.sleep(2)
                else:
                    print(f"   ⚠️  Clique em INCLUIR")
                    input(f"   >>> ENTER após clicar <<<\n")
                
            except:
                input(f"   >>> Clique em INCLUIR manualmente <<<\n")
            
            # Verificar resultado
            time.sleep(1)
            page = self.driver.page_source.lower()
            
            if any(x in page for x in ['sucesso', 'incluído', 'cadastrado']):
                print(f"   ✅ CADASTRADO!")
                self.sucessos.append(numero)
                return True
            elif any(x in page for x in ['já cadastrado', 'já existe']):
                print(f"   ⚠️  Já cadastrado")
                self.ja_cadastrados.append(numero)
                return True
            elif any(x in page for x in ['erro', 'inválido']):
                print(f"   ❌ Erro")
                self.falhas.append(numero)
                return False
            else:
                opcao = input(f"   >>> Cadastrou? (s/n): ").lower()
                if opcao == 's':
                    self.sucessos.append(numero)
                    return True
                else:
                    self.falhas.append(numero)
                    return False
            
        except Exception as e:
            print(f"   ❌ Erro: {str(e)[:80]}")
            self.falhas.append(numero)
            return False
    
    def processar_planilha(self, arquivo):
        print(f"\n📊 Processando: {arquivo}")
        
        if not os.path.exists(arquivo):
            print(f"❌ Arquivo não encontrado!")
            return
        
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        processos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                num = str(row[0]).strip()
                if '.8.26.' in num:
                    processos.append(num)
        
        total = len(processos)
        print(f"✅ {total} processos do TJSP\n")
        
        for idx, numero in enumerate(processos, 1):
            print(f"\n{'='*70}")
            print(f"Processo {idx}/{total}")
            print(f"{'='*70}")
            
            self.cadastrar_processo(numero)
            
            if idx < total:
                time.sleep(1.5)
        
        self.gerar_relatorio()
    
    def gerar_relatorio(self):
        print("\n" + "="*70)
        print("📊 RELATÓRIO FINAL")
        print("="*70)
        
        total = len(self.sucessos) + len(self.ja_cadastrados) + len(self.falhas)
        
        print(f"\n📋 Total: {total}")
        print(f"✅ Cadastrados: {len(self.sucessos)}")
        print(f"⚠️  Já cadastrados: {len(self.ja_cadastrados)}")
        print(f"❌ Falhas: {len(self.falhas)}")
        
        if total > 0:
            taxa = ((len(self.sucessos) + len(self.ja_cadastrados)) / total) * 100
            print(f"📊 Taxa: {taxa:.1f}%")
        
        if self.falhas:
            print(f"\n❌ FALHAS:")
            for p in self.falhas[:15]:
                print(f"   - {p}")
        
        print("="*70)
        
        # Salvar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo = f"relatorio_push_tjsp_{timestamp}.txt"
        
        with open(arquivo, "w", encoding="utf-8") as f:
            f.write("RELATÓRIO PUSH TJSP\n")
            f.write("="*70 + "\n\n")
            f.write(f"Total: {total}\n")
            f.write(f"Cadastrados: {len(self.sucessos)}\n")
            f.write(f"Já cadastrados: {len(self.ja_cadastrados)}\n")
            f.write(f"Falhas: {len(self.falhas)}\n\n")
            
            if self.sucessos:
                f.write("CADASTRADOS:\n")
                for p in self.sucessos:
                    f.write(f"  {p}\n")
            
            if self.ja_cadastrados:
                f.write("\nJÁ CADASTRADOS:\n")
                for p in self.ja_cadastrados:
                    f.write(f"  {p}\n")
            
            if self.falhas:
                f.write("\nFALHAS:\n")
                for p in self.falhas:
                    f.write(f"  {p}\n")
        
        print(f"\n📄 Relatório: {arquivo}")
    
    def fechar(self):
        if self.driver:
            self.driver.quit()

# MAIN
if __name__ == "__main__":
    print("="*70)
    print("🔔 PUSH TJSP - LOGIN MANUAL")
    print("="*70)
    print("\n💡 Use esta versão se:")
    print("   - Web Signer não está instalado")
    print("   - Quer usar login/senha ao invés de certificado")
    
    push = PushTJSPLoginManual()
    
    try:
        input("\nENTER para começar...\n")
        
        push.iniciar()
        
        if push.login_manual_completo():
            
            print("\n💡 MODO:")
            print("   1. Teste (1 processo)")
            print("   2. Teste (5 processos)")
            print("   3. Completo (todos)")
            
            modo = input("\nDigite: ").strip()
            
            if modo == "1":
                num = input("\nProcesso: ").strip()
                push.cadastrar_processo(num)
                
            elif modo == "2":
                wb = openpyxl.load_workbook("processos_push_20260126_185045.xlsx")
                ws = wb.active
                
                teste = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        num = str(row[0]).strip()
                        if '.8.26.' in num and len(teste) < 5:
                            teste.append(num)
                
                print(f"\n✅ {len(teste)} processos\n")
                
                for idx, num in enumerate(teste, 1):
                    print(f"\n{'='*70}")
                    print(f"Teste {idx}/{len(teste)}")
                    print(f"{'='*70}")
                    push.cadastrar_processo(num)
                    if idx < len(teste):
                        time.sleep(1.5)
                
                push.gerar_relatorio()
                
            elif modo == "3":
                push.processar_planilha("processos_push_20260126_185045.xlsx")
        
        input("\n\nENTER para fechar...\n")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Cancelado")
        push.gerar_relatorio()
    
    except Exception as e:
        print(f"\n❌ Erro: {e}")
    
    finally:
        push.fechar()
    
    print("\n✅ FIM!")
