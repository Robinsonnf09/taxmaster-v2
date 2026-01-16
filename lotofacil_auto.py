import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================
# CONFIGURAÇÕES AUTOMÁTICAS
# ==========================

QTDE_DEZENAS_TOTAL = 25
QTDE_DEZENAS_POR_JOGO = 15

# Configurações padrão
QTD_JOGOS = 20
MIN_PARES = 5
MAX_PARES = 10
SOMA_MIN = 150
SOMA_MAX = 250


# ==========================
# FUNÇÕES PRINCIPAIS
# ==========================

def gerar_jogo_lotofacil():
    return sorted(random.sample(range(1, QTDE_DEZENAS_TOTAL + 1), QTDE_DEZENAS_POR_JOGO))


def contar_pares_impares(jogo):
    pares = sum(1 for n in jogo if n % 2 == 0)
    impares = len(jogo) - pares
    return pares, impares


def jogo_valido(jogo, min_pares, max_pares, soma_min, soma_max):
    pares, _ = contar_pares_impares(jogo)
    soma = sum(jogo)

    if not (min_pares <= pares <= max_pares):
        return False
    if not (soma_min <= soma <= soma_max):
        return False

    return True


def gerar_jogos_filtrados(qtd_jogos, min_pares, max_pares, soma_min, soma_max):
    jogos = []
    tentativas = 0

    while len(jogos) < qtd_jogos:
        tentativas += 1
        jogo = gerar_jogo_lotofacil()

        if jogo_valido(jogo, min_pares, max_pares, soma_min, soma_max):
            if jogo not in jogos:
                jogos.append(jogo)

    return jogos, tentativas


# ==========================
# EXCEL PROFISSIONAL
# ==========================

def salvar_em_excel(jogos):
    data = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    nome_arquivo = f"jogos_lotofacil_{data}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Jogos"

    # Cabeçalho
    cabecalho = ["Jogo", "Dezenas", "Pares", "Ímpares", "Soma"]
    ws.append(cabecalho)

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(cabecalho) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Bordas finas
    thin = Side(border_style="thin", color="000000")

    # Inserir jogos
    for i, jogo in enumerate(jogos, start=1):
        pares, impares = contar_pares_impares(jogo)
        soma = sum(jogo)
        dezenas_formatadas = " ".join(f"{n:02d}" for n in jogo)

        linha = [i, dezenas_formatadas, pares, impares, soma]
        ws.append(linha)

        # Aplicar bordas e alinhamento
        for col in range(1, 6):
            cell = ws.cell(row=i + 1, column=col)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = center

        # Destacar somas fora do padrão
        if soma < SOMA_MIN or soma > SOMA_MAX:
            for col in range(1, 6):
                ws.cell(row=i + 1, column=col).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

    # Ajustar largura das colunas
    larguras = [10, 40, 10, 10, 10]
    for col, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura

    # Filtro automático
    ws.auto_filter.ref = "A1:E1"

    wb.save(nome_arquivo)
    print(f"\nArquivo Excel profissional salvo como: {nome_arquivo}")


# ==========================
# EXECUÇÃO AUTOMÁTICA
# ==========================

def main():
    print("=" * 50)
    print(" GERADOR AUTOMÁTICO LOTOFÁCIL ")
    print("=" * 50)

    print("\nConfigurações automáticas:")
    print(f"- Quantidade de jogos: {QTD_JOGOS}")
    print(f"- Pares entre {MIN_PARES} e {MAX_PARES}")
    print(f"- Soma entre {SOMA_MIN} e {SOMA_MAX}")

    print("\nGerando jogos, aguarde...")

    jogos, tentativas = gerar_jogos_filtrados(
        QTD_JOGOS, MIN_PARES, MAX_PARES, SOMA_MIN, SOMA_MAX
    )

    print(f"\nTentativas realizadas: {tentativas}")
    print(f"Jogos gerados: {len(jogos)}\n")

    for i, jogo in enumerate(jogos, start=1):
        dezenas_formatadas = " ".join(f"{n:02d}" for n in jogo)
        pares, impares = contar_pares_impares(jogo)
        soma = sum(jogo)
        print(f"Jogo {i:02d}: {dezenas_formatadas} | Pares: {pares} | Ímpares: {impares} | Soma: {soma}")

    salvar_em_excel(jogos)


if __name__ == "__main__":
    main()