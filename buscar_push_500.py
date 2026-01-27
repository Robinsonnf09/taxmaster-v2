"""
BUSCAR 500 PROCESSOS - PUSH REQUISITÓRIOS TJSP
Busca diretamente na página de requisitórios PUSH
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re

class BuscadorProcessosPush:
    
    def __init__(self):
        self.driver = None
        self.processos = []
    
    def iniciar(self):
        print("\n🌐 Iniciando Chrome...")
        
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
        print("✅ Chrome iniciado!")
    
    def fazer_login(self):
        print(f"\n🔐 Login...")
        self.driver.get("https://esaj.tjsp.jus.br")
        time.sleep(3)
        
        print("\n" + "="*70)
        print("⚠️  FAÇA LOGIN:")
        print("="*70)
        print("   ❌ Se aparecer 'WEB SIGNER' → CANCELAR")
        print("   ✅ Use LOGIN e SENHA")
        print("="*70)
        
        input("\n>>> ENTER após login <<<\n")
        
        print("✅ Login efetuado!")
        return True
    
    def buscar_processos_push(self, limite=500):
        """Busca processos direto do PUSH"""
        
        print(f"\n🔍 Acessando PUSH de Requisitórios...")
        
        # Acessar diretamente o PUSH
        url_push = "https://esaj.tjsp.jus.br/cpopg/abrirPasta.do?gateway=true"
        
        self.driver.get(url_push)
        time.sleep(3)
        
        print("\n" + "="*70)
        print("⚠️  INSTRUÇÕES:")
        print("="*70)
        print("   1. No menu lateral, clique em 'PUSH'")
        print("   2. Clique em 'Requisitórios'")
        print("   3. Configure os filtros:")
        print("      📅 Data Protocolo De: 01/11/2024")
        print("      📅 Data Protocolo Até: 31/12/2025")
        print("      📋 Natureza: Alimentares (se disponível)")
        print("   4. Clique em PESQUISAR")
        print("   5. Aguarde a lista carregar")
        print("="*70)
        
        input("\n>>> ENTER após carregar a lista <<<\n")
        
        print(f"\n📋 Extraindo processos...")
        
        processos_encontrados = set()
        pagina = 1
        tentativas_vazias = 0
        
        while len(processos_encontrados) < limite and tentativas_vazias < 3:
            print(f"\n   📄 Página {pagina}...", end=" ", flush=True)
            
            time.sleep(2)
            
            try:
                # Script melhorado para extrair processos
                script = r"""
                let processos = new Set();
                
                // Padrão de processo TJSP
                let pattern = /\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}/g;
                
                // Buscar em todo o body
                let texto = document.body.innerText;
                let matches = texto.match(pattern);
                
                if (matches) {
                    matches.forEach(p => processos.add(p));
                }
                
                // Buscar em tabelas especificamente
                document.querySelectorAll('table td').forEach(td => {
                    let m = td.textContent.match(pattern);
                    if (m) {
                        m.forEach(p => processos.add(p));
                    }
                });
                
                // Buscar em links
                document.querySelectorAll('a').forEach(a => {
                    let m = a.textContent.match(pattern);
                    if (m) {
                        m.forEach(p => processos.add(p));
                    }
                    
                    // Também no href
                    if (a.href) {
                        let m2 = a.href.match(pattern);
                        if (m2) {
                            m2.forEach(p => processos.add(p));
                        }
                    }
                });
                
                return Array.from(processos);
                """
                
                processos_pagina = self.driver.execute_script(script)
                
                if processos_pagina and len(processos_pagina) > 0:
                    novos = 0
                    for proc in processos_pagina:
                        if proc not in processos_encontrados:
                            processos_encontrados.add(proc)
                            novos += 1
                            if len(processos_encontrados) >= limite:
                                break
                    
                    print(f"✅ {novos} novos (total: {len(processos_encontrados)})")
                    tentativas_vazias = 0
                else:
                    print(f"⚠️  Nenhum processo")
                    tentativas_vazias += 1
                
                # Verificar se atingiu o limite
                if len(processos_encontrados) >= limite:
                    print(f"\n   ✅ Limite de {limite} atingido!")
                    break
                
                # Tentar próxima página
                try:
                    # Procurar botão "Próxima" ou similar
                    proxima_buttons = self.driver.find_elements(By.XPATH, 
                        "//a[contains(text(), 'Próxima') or contains(text(), 'próxima') or contains(@title, 'Próxima')]"
                    )
                    
                    if proxima_buttons:
                        proxima_buttons[0].click()
                        time.sleep(2)
                        pagina += 1
                    else:
                        print(f"\n   ℹ️  Não há mais páginas")
                        break
                        
                except Exception as e:
                    print(f"\n   ℹ️  Fim da paginação")
                    break
                
            except Exception as e:
                print(f"❌ Erro: {str(e)[:50]}")
                tentativas_vazias += 1
        
        self.processos = list(processos_encontrados)[:limite]
        
        print(f"\n" + "="*70)
        print(f"✅ TOTAL ENCONTRADO: {len(self.processos)} processos")
        print("="*70)
        
        return self.processos
    
    def criar_planilha(self, arquivo_saida):
        """Cria planilha Excel"""
        
        print(f"\n📊 Criando planilha...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Processos"
        
        ws['A1'] = 'Nº Processo'
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = header_alignment
        
        for row_idx, processo in enumerate(self.processos, 2):
            ws.cell(row=row_idx, column=1, value=processo)
        
        ws.column_dimensions['A'].width = 35
        ws.row_dimensions[1].height = 25
        ws.auto_filter.ref = f'A1:A{len(self.processos)+1}'
        
        wb.save(arquivo_saida)
        
        print(f"   ✅ Salvo: {arquivo_saida}")
    
    def fechar(self):
        if self.driver:
            self.driver.quit()

if __name__ == "__main__":
    print("="*70)
    print("🔍 BUSCAR PROCESSOS - PUSH TJSP (11/2024 a 12/2025)")
    print("="*70)
    
    buscador = BuscadorProcessosPush()
    
    try:
        input("\nENTER...\n")
        
        buscador.iniciar()
        
        if buscador.fazer_login():
            processos = buscador.buscar_processos_push(limite=500)
            
            if len(processos) > 0:
                print(f"\n📋 PREVIEW (primeiros 15):")
                for i, proc in enumerate(processos[:15], 1):
                    print(f"   {i:3}. {proc}")
                
                if len(processos) > 15:
                    print(f"\n   ... e mais {len(processos) - 15}")
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                arquivo = f"processos_push_{timestamp}.xlsx"
                
                buscador.criar_planilha(arquivo)
                
                print("\n" + "="*70)
                print("🎉 CONCLUÍDO!")
                print("="*70)
                print(f"\n📁 {arquivo}")
                print(f"📊 {len(processos)} processos")
            else:
                print("\n⚠️  Nenhum processo encontrado!")
                print("   Verifique se:")
                print("   1. Fez a busca no PUSH")
                print("   2. Há resultados na tela")
                print("   3. Aguardou carregar completamente")
        
        input("\n\nENTER para fechar...\n")
        
    except Exception as e:
        print(f"\n❌ {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        buscador.fechar()
    
    print("\n✅ FIM!")
