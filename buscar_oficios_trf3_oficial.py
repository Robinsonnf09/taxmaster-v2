"""
BUSCADOR DE OFÍCIOS REQUISITÓRIOS - TRF3 OFICIAL
URL: https://web.trf3.jus.br/consultas/Internet/ConsultaReqPag
Baseado na documentação oficial do TRF3
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
import os
from datetime import datetime

class BuscadorOficiosTRF3Oficial:
    
    def __init__(self):
        self.driver = None
        self.sucessos = []
        self.falhas = []
        self.url_consulta = "https://web.trf3.jus.br/consultas/Internet/ConsultaReqPag"
        
        self.pasta_oficios = "oficios_trf3_oficial"
        if not os.path.exists(self.pasta_oficios):
            os.makedirs(self.pasta_oficios)
        
        # Relatório
        self.relatorio = []
    
    def iniciar(self):
        print("\n🌐 Iniciando Chrome...")
        
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--ignore-certificate-errors')
        
        prefs = {
            "download.default_directory": os.path.abspath(self.pasta_oficios),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        print("✅ Chrome iniciado!")
    
    def acessar_sistema(self):
        print(f"\n🔐 Acessando sistema oficial TRF3...")
        print(f"   URL: {self.url_consulta}")
        
        self.driver.get(self.url_consulta)
        time.sleep(5)
        
        print("\n" + "="*70)
        print("⚠️  VERIFICAÇÃO MANUAL:")
        print("="*70)
        print("   - Se aparecer CAPTCHA, resolva manualmente")
        print("   - Aguarde a página carregar completamente")
        print("   - Você deve ver o formulário de consulta")
        print("="*70)
        
        input("\n>>> ENTER quando estiver vendo o formulário de consulta <<<\n")
        
        print("✅ Sistema acessado!")
        return True
    
    def buscar_processo(self, numero_processo, cpf_cnpj=None):
        try:
            print(f"\n📝 Buscando: {numero_processo}")
            
            wait = WebDriverWait(self.driver, 15)
            
            # CAMPO CPF/CNPJ (obrigatório segundo documentação)
            if cpf_cnpj:
                campo_cpf = None
                try:
                    campo_cpf = self.driver.find_element(By.XPATH, 
                        "//input[contains(@name, 'cpf') or contains(@name, 'cnpj') or contains(@id, 'cpf') or contains(@id, 'cnpj')]")
                    campo_cpf.clear()
                    campo_cpf.send_keys(cpf_cnpj)
                    print(f"   ✅ CPF/CNPJ digitado: {cpf_cnpj}")
                except:
                    print(f"   ⚠️  Campo CPF/CNPJ não encontrado")
            
            # CAMPO PROCESSO ORIGEM
            campo_processo = None
            try:
                # Tentar localizar campo de processo
                campos_possiveis = [
                    "//input[contains(@name, 'processo')]",
                    "//input[contains(@id, 'processo')]",
                    "//input[contains(@placeholder, 'processo')]"
                ]
                
                for xpath in campos_possiveis:
                    try:
                        campo_processo = self.driver.find_element(By.XPATH, xpath)
                        break
                    except:
                        continue
                
                if campo_processo:
                    campo_processo.clear()
                    campo_processo.send_keys(numero_processo)
                    print(f"   ✅ Processo digitado")
                else:
                    print(f"   ⚠️  Campo não encontrado automaticamente")
                    print(f"   💡 Digite manualmente: {numero_processo}")
                    if cpf_cnpj:
                        print(f"   💡 E o CPF/CNPJ: {cpf_cnpj}")
                    input(f"   >>> ENTER após preencher <<<\n")
            
            except Exception as e:
                print(f"   ⚠️  Erro ao preencher: {str(e)[:50]}")
                print(f"   💡 Preencha manualmente e pressione ENTER")
                input(f"   >>> <<<\n")
            
            # BOTÃO CONSULTAR
            try:
                btn = self.driver.find_element(By.XPATH, 
                    "//button[contains(text(), 'Consultar') or contains(text(), 'Pesquisar') or contains(text(), 'Buscar')] | //input[@type='submit']")
                btn.click()
                print(f"   ⏳ Consultando...")
                time.sleep(5)
            except:
                print(f"   ⚠️  Clique manualmente em CONSULTAR")
                input(f"   >>> ENTER após clicar <<<\n")
            
            # ANALISAR RESULTADO
            page = self.driver.page_source.lower()
            
            if "não encontrado" in page or "nenhum resultado" in page:
                print(f"   ❌ Processo não encontrado")
                self.falhas.append(numero_processo)
                self.relatorio.append({
                    "processo": numero_processo,
                    "status": "Não encontrado",
                    "oficio": "Não",
                    "proposta": "-",
                    "banco": "-"
                })
                return False
            
            # EXTRAIR INFORMAÇÕES
            print(f"   🔍 Extraindo informações...")
            
            info = {
                "processo": numero_processo,
                "status": "Encontrado",
                "oficio": "Não identificado",
                "proposta": "-",
                "banco": "-",
                "valor": "-"
            }
            
            # Procurar informações na página
            try:
                # Status
                if "pago - comunicado" in page:
                    info["status"] = "PAGO"
                    print(f"   ✅ Status: PAGO")
                elif "proposta orçamentária" in page or "po " in page:
                    info["status"] = "Em Proposta Orçamentária"
                    print(f"   ⚠️  Status: Em PO")
                
                # Banco
                if "banco do brasil" in page or "bb" in page:
                    info["banco"] = "Banco do Brasil"
                    print(f"   🏦 Banco: BB")
                elif "caixa" in page or "cef" in page:
                    info["banco"] = "Caixa Econômica Federal"
                    print(f"   🏦 Banco: CEF")
                
                # Ofício requisitório
                links = self.driver.find_elements(By.TAG_NAME, "a")
                oficios_encontrados = []
                
                for link in links:
                    texto = link.text.lower()
                    if any(x in texto for x in ['ofício', 'requisitório', 'or', 'pdf', 'download']):
                        oficios_encontrados.append(link)
                
                if oficios_encontrados:
                    info["oficio"] = f"{len(oficios_encontrados)} ofício(s)"
                    print(f"   📄 Ofícios: {len(oficios_encontrados)}")
                    
                    # Baixar ofícios
                    for idx, link in enumerate(oficios_encontrados, 1):
                        try:
                            print(f"   📥 Baixando ofício {idx}...")
                            link.click()
                            time.sleep(3)
                            
                            # Fechar aba se abriu
                            if len(self.driver.window_handles) > 1:
                                self.driver.switch_to.window(self.driver.window_handles[-1])
                                time.sleep(1)
                                self.driver.close()
                                self.driver.switch_to.window(self.driver.window_handles[0])
                            
                            print(f"   ✅ Ofício {idx} baixado!")
                        except:
                            print(f"   ⚠️  Erro ao baixar ofício {idx}")
                else:
                    print(f"   ⚠️  Nenhum link de ofício encontrado")
                    print(f"   💡 Há ofício visível na página?")
                    opcao = input(f"   >>> (s/n): ").lower()
                    
                    if opcao == 's':
                        info["oficio"] = "Sim (manual)"
                        input(f"   >>> Baixe manualmente e pressione ENTER <<<\n")
            
            except Exception as e:
                print(f"   ⚠️  Erro ao extrair: {str(e)[:50]}")
            
            self.relatorio.append(info)
            self.sucessos.append(numero_processo)
            
            return True
            
        except Exception as e:
            print(f"   ❌ Erro: {str(e)[:100]}")
            self.falhas.append(numero_processo)
            self.relatorio.append({
                "processo": numero_processo,
                "status": f"Erro: {str(e)[:50]}",
                "oficio": "-",
                "proposta": "-",
                "banco": "-"
            })
            return False
    
    def processar_planilha(self, arquivo, coluna_cpf=None):
        print(f"\n📊 Processando planilha: {arquivo}")
        
        if not os.path.exists(arquivo):
            print(f"❌ Arquivo não encontrado!")
            return
        
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        processos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                cpf = row[coluna_cpf] if coluna_cpf and len(row) > coluna_cpf else None
                processos.append({
                    "numero": str(row[0]).strip(),
                    "cpf_cnpj": str(cpf).strip() if cpf else None
                })
        
        total = len(processos)
        print(f"✅ {total} processos")
        print(f"📁 Salvando em: {os.path.abspath(self.pasta_oficios)}\n")
        
        for idx, processo in enumerate(processos, 1):
            print(f"\n{'='*70}")
            print(f"Processo {idx}/{total}")
            print(f"{'='*70}")
            
            self.buscar_processo(processo["numero"], processo["cpf_cnpj"])
            
            # Voltar para nova busca
            if idx < total:
                print("\n   ⏪ Voltando para nova busca...")
                try:
                    self.driver.back()
                    time.sleep(2)
                except:
                    print("   ⚠️  Navegue para nova busca")
                    input("   >>> ENTER quando pronto <<<\n")
                
                time.sleep(2)
        
        # Gerar relatório Excel
        self.gerar_relatorio_excel()
        
        # Resumo
        print("\n" + "="*70)
        print("📊 RESUMO FINAL:")
        print("="*70)
        print(f"   🏛️  Sistema: TRF3 Oficial")
        print(f"   📋 Total: {total}")
        print(f"   ✅ Sucessos: {len(self.sucessos)}")
        print(f"   ❌ Falhas: {len(self.falhas)}")
        if total > 0:
            print(f"   📊 Taxa: {(len(self.sucessos)/total*100):.1f}%")
        
        print(f"\n📁 Ofícios em: {os.path.abspath(self.pasta_oficios)}")
        print(f"📄 Relatório: relatorio_oficios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        print("="*70)
    
    def gerar_relatorio_excel(self):
        print("\n📄 Gerando relatório Excel...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório Ofícios"
        
        # Cabeçalho
        ws.append(["Processo", "Status", "Ofício", "Proposta Orçamentária", "Banco", "Valor"])
        
        # Dados
        for item in self.relatorio:
            ws.append([
                item.get("processo", "-"),
                item.get("status", "-"),
                item.get("oficio", "-"),
                item.get("proposta", "-"),
                item.get("banco", "-"),
                item.get("valor", "-")
            ])
        
        arquivo = f"relatorio_oficios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(arquivo)
        
        print(f"✅ Relatório salvo: {arquivo}")
    
    def fechar(self):
        if self.driver:
            self.driver.quit()
            print("\n🔒 Navegador fechado")

# MAIN
if __name__ == "__main__":
    print("="*70)
    print("🔔 BUSCADOR DE OFÍCIOS REQUISITÓRIOS - TRF3 OFICIAL")
    print("="*70)
    print("\n📋 URL Oficial: https://web.trf3.jus.br/consultas/Internet/ConsultaReqPag")
    
    buscador = BuscadorOficiosTRF3Oficial()
    
    try:
        input("\nENTER para começar...\n")
        
        buscador.iniciar()
        
        if buscador.acessar_sistema():
            
            print("\n💡 MODO:")
            print("   1. Planilha completa (215 processos)")
            print("   2. Teste individual")
            
            modo = input("\nDigite 1 ou 2: ").strip()
            
            if modo == "1":
                arq = "processos_push_20260126_185045.xlsx"
                
                print("\n❓ Sua planilha tem CPF/CNPJ dos beneficiários?")
                tem_cpf = input("   (s/n): ").lower()
                
                coluna_cpf = None
                if tem_cpf == 's':
                    coluna_cpf = int(input("   Número da coluna do CPF (ex: 1, 2, 3...): ").strip()) - 1
                
                buscador.processar_planilha(arq, coluna_cpf)
            
            elif modo == "2":
                num = input("\nNúmero do processo: ").strip()
                cpf = input("CPF/CNPJ (ou ENTER para pular): ").strip()
                buscador.buscar_processo(num, cpf if cpf else None)
        
        input("\n\n>>> ENTER para fechar <<<\n")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Operação cancelada")
    
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        buscador.fechar()
    
    print("\n✅ CONCLUÍDO!")

