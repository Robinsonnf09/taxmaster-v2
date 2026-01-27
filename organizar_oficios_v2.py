"""
ORGANIZAR OFÍCIOS - VERSÃO MELHORADA
Usa múltiplos métodos de extração + análise dos nomes dos arquivos
"""

import os
import shutil
import PyPDF2
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

class OrganizadorMelhorado:
    
    def __init__(self, pasta_origem):
        self.pasta_origem = pasta_origem
        self.pasta_organizada = "oficios_por_orgao"
        self.estatisticas = defaultdict(lambda: {'quantidade': 0, 'arquivos': []})
        self.erros = []
        
        if not os.path.exists(self.pasta_organizada):
            os.makedirs(self.pasta_organizada)
    
    def extrair_orgao_do_nome_arquivo(self, nome_arquivo):
        """Extrai órgão do nome do arquivo (backup)"""
        
        # Padrões comuns em nomes de arquivo
        if 'estado' in nome_arquivo.lower():
            return 'Estado de São Paulo'
        elif 'municipio' in nome_arquivo.lower() or 'prefeitura' in nome_arquivo.lower():
            return 'Municípios'
        elif 'fazenda' in nome_arquivo.lower():
            return 'Fazenda do Estado de SP'
        
        return None
    
    def extrair_orgao_devedor(self, pdf_path, nome_arquivo):
        """Extrai órgão devedor usando múltiplos métodos"""
        
        # MÉTODO 1: Do nome do arquivo
        orgao_nome = self.extrair_orgao_do_nome_arquivo(nome_arquivo)
        if orgao_nome:
            return orgao_nome
        
        # MÉTODO 2: Tentar ler o PDF
        try:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                
                # Tentar várias páginas se necessário
                texto = ""
                paginas_tentar = min(3, len(reader.pages))
                
                for i in range(paginas_tentar):
                    try:
                        texto += reader.pages[i].extract_text() + "\n"
                    except:
                        continue
                
                if not texto or len(texto.strip()) < 50:
                    return 'PDF Protegido ou Vazio'
                
                # Procurar padrões
                texto_lower = texto.lower()
                
                # Estado de São Paulo
                if any(x in texto_lower for x in ['estado de são paulo', 'estado de sao paulo', 'governo do estado']):
                    if 'fazenda' in texto_lower:
                        return 'Fazenda do Estado de SP'
                    return 'Estado de São Paulo'
                
                # Municípios
                municipios = re.findall(r'(?:município|prefeitura)\s+(?:municipal\s+)?(?:de\s+)?([a-zà-ú\s]{3,30})', texto, re.IGNORECASE)
                if municipios:
                    municipio = municipios[0].strip().title()
                    return f'Município de {municipio}'
                
                # Autarquias e outros órgãos
                if 'sabesp' in texto_lower:
                    return 'SABESP'
                elif 'ipesp' in texto_lower:
                    return 'IPESP'
                elif 'spprev' in texto_lower:
                    return 'SPPrev'
                elif any(x in texto_lower for x in ['universidade', 'usp', 'unesp', 'unicamp']):
                    return 'Universidades Estaduais'
                elif 'desenvolvimento social' in texto_lower:
                    return 'Secretaria de Desenvolvimento Social'
                elif 'educação' in texto_lower or 'educacao' in texto_lower:
                    return 'Secretaria de Educação'
                elif 'saúde' in texto_lower or 'saude' in texto_lower:
                    return 'Secretaria de Saúde'
                elif 'segurança' in texto_lower:
                    return 'Secretaria de Segurança Pública'
                
                # Padrão genérico para secretarias
                secretarias = re.findall(r'secretaria\s+(?:de\s+)?([a-zà-ú\s]{3,40})', texto, re.IGNORECASE)
                if secretarias:
                    sec = secretarias[0].strip().title()
                    return f'Secretaria de {sec}'
                
                # Se chegou aqui e tem texto, mas não identificou
                if len(texto.strip()) > 100:
                    return 'Outros Órgãos Estaduais'
                
                return 'Não Identificado'
                
        except Exception as e:
            self.erros.append({
                'arquivo': nome_arquivo,
                'erro': str(e)[:100]
            })
            return 'Erro na Leitura'
    
    def criar_nome_pasta_seguro(self, nome):
        """Cria nome de pasta válido"""
        
        nome = re.sub(r'[<>:"/\|?*]', '', nome)
        nome = nome.strip()
        
        if len(nome) > 80:
            nome = nome[:80]
        
        return nome if nome else 'Sem_Nome'
    
    def organizar_oficios(self):
        """Organiza todos os ofícios"""
        
        print(f"\n📂 Organizando ofícios (versão melhorada)...")
        print(f"   Origem: {self.pasta_origem}")
        print(f"   Destino: {self.pasta_organizada}")
        
        pdfs = [f for f in os.listdir(self.pasta_origem) if f.endswith('.pdf')]
        total = len(pdfs)
        
        print(f"\n   📄 Total: {total} PDFs")
        print(f"\n🔍 Processando...")
        
        for idx, pdf in enumerate(pdfs, 1):
            if idx % 50 == 0 or idx == 1:
                print(f"   📋 {idx}/{total} ({idx/total*100:.1f}%) - {len(self.estatisticas)} órgãos identificados", end="\r")
            
            try:
                pdf_path = os.path.join(self.pasta_origem, pdf)
                
                # Extrair órgão (com múltiplos métodos)
                orgao = self.extrair_orgao_devedor(pdf_path, pdf)
                
                # Criar pasta
                nome_pasta = self.criar_nome_pasta_seguro(orgao)
                pasta_orgao = os.path.join(self.pasta_organizada, nome_pasta)
                
                if not os.path.exists(pasta_orgao):
                    os.makedirs(pasta_orgao)
                
                # Copiar
                destino = os.path.join(pasta_orgao, pdf)
                
                # Se já existe, adicionar sufixo
                if os.path.exists(destino):
                    base, ext = os.path.splitext(pdf)
                    contador = 1
                    while os.path.exists(os.path.join(pasta_orgao, f"{base}_{contador}{ext}")):
                        contador += 1
                    destino = os.path.join(pasta_orgao, f"{base}_{contador}{ext}")
                
                shutil.copy2(pdf_path, destino)
                
                # Estatísticas
                self.estatisticas[orgao]['quantidade'] += 1
                self.estatisticas[orgao]['arquivos'].append(pdf)
                
            except Exception as e:
                print(f"\n   ❌ Erro em {pdf}: {str(e)[:50]}")
                self.erros.append({
                    'arquivo': pdf,
                    'erro': str(e)[:100]
                })
                continue
        
        print(f"\n   ✅ {total} PDFs processados!")
        
        return True
    
    def gerar_relatorio(self):
        """Gera relatório detalhado"""
        
        print(f"\n📊 Gerando relatório...")
        
        orgaos_ordenados = sorted(
            self.estatisticas.items(),
            key=lambda x: x[1]['quantidade'],
            reverse=True
        )
        
        print(f"\n" + "="*70)
        print(f"📊 ESTATÍSTICAS POR ÓRGÃO DEVEDOR")
        print("="*70)
        
        print(f"\n   {'#':>3} {'Órgão Devedor':<45} {'Qtd':>8} {'%':>6}")
        print(f"   {'-'*3} {'-'*45} {'-'*8} {'-'*6}")
        
        total_geral = sum(d['quantidade'] for _, d in orgaos_ordenados)
        
        for idx, (orgao, dados) in enumerate(orgaos_ordenados, 1):
            perc = (dados['quantidade'] / total_geral) * 100
            print(f"   {idx:>3} {orgao:<45} {dados['quantidade']:>8} {perc:>5.1f}%")
        
        print(f"\n   {'':>3} {'TOTAL':<45} {total_geral:>8} {'100.0%':>6}")
        
        # Relatório de erros
        if self.erros:
            print(f"\n⚠️  Erros encontrados: {len(self.erros)}")
        
        # Criar planilha
        self.criar_planilha_relatorio(orgaos_ordenados)
    
    def criar_planilha_relatorio(self, orgaos_ordenados):
        """Cria planilha Excel"""
        
        print(f"\n📊 Criando planilha...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Por Órgão"
        
        # Cabeçalhos
        headers = ['#', 'Órgão Devedor', 'Quantidade', 'Percentual']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        total_geral = sum(d['quantidade'] for _, d in orgaos_ordenados)
        
        for row_idx, (idx_org, (orgao, dados)) in enumerate(enumerate(orgaos_ordenados, 1), 2):
            ws.cell(row=row_idx, column=1, value=idx_org)
            ws.cell(row=row_idx, column=2, value=orgao)
            ws.cell(row=row_idx, column=3, value=dados['quantidade'])
            
            perc = (dados['quantidade'] / total_geral) * 100
            ws.cell(row=row_idx, column=4, value=f"{perc:.1f}%")
        
        # Total
        row_total = len(orgaos_ordenados) + 2
        ws.cell(row=row_total, column=2, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row_total, column=3, value=total_geral).font = Font(bold=True)
        ws.cell(row=row_total, column=4, value='100%').font = Font(bold=True)
        
        # Larguras
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        
        ws.row_dimensions[1].height = 25
        
        ws.auto_filter.ref = f'A1:D{len(orgaos_ordenados)+1}'
        
        # Salvar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo = f"relatorio_orgaos_{timestamp}.xlsx"
        
        wb.save(arquivo)
        
        print(f"   ✅ Salvo: {arquivo}")

if __name__ == "__main__":
    print("="*70)
    print("📂 ORGANIZAR OFÍCIOS POR ÓRGÃO - VERSÃO MELHORADA")
    print("="*70)
    
    pasta_origem = "oficios_requisitorios_tjsp"
    
    if not os.path.exists(pasta_origem):
        print(f"\n❌ Pasta não encontrada: {pasta_origem}")
        input("\nENTER...")
        exit()
    
    pdfs = [f for f in os.listdir(pasta_origem) if f.endswith('.pdf')]
    print(f"\n📊 Encontrados: {len(pdfs)} PDFs")
    
    confirma = input(f"\nOrganizar? (s/n): ").lower()
    
    if confirma != 's':
        print("\n❌ Cancelado")
        exit()
    
    try:
        inicio = datetime.now()
        
        org = OrganizadorMelhorado(pasta_origem)
        
        if org.organizar_oficios():
            org.gerar_relatorio()
            
            fim = datetime.now()
            duracao = fim - inicio
            
            print("\n" + "="*70)
            print("🎉 CONCLUÍDO!")
            print("="*70)
            print(f"\n⏱️  {int(duracao.total_seconds()/60)} min")
            print(f"📂 {org.pasta_organizada}/")
            print(f"📊 {len(org.estatisticas)} órgãos")
            
            if org.erros:
                print(f"⚠️  {len(org.erros)} erros")
        
    except Exception as e:
        print(f"\n❌ {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nENTER...\n")
    
    print("\n✅ FIM!")
