"""
GERAR PLANILHA DE PRECATÓRIOS DO PDF
Extrai todos os precatórios e cria Excel
"""

import PyPDF2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from datetime import datetime

def extrair_precatorios_pdf(arquivo_pdf):
    """Extrai todos os precatórios do PDF"""
    
    print(f"\n📄 Lendo PDF: {arquivo_pdf}")
    
    precatorios = []
    
    with open(arquivo_pdf, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        
        print(f"   Total de páginas: {len(reader.pages)}")
        
        texto_completo = ""
        for pagina in reader.pages:
            texto_completo += pagina.extract_text()
        
        # Dividir por "Devedora:" (início de cada registro)
        blocos = texto_completo.split("Devedora:")
        
        print(f"   Blocos encontrados: {len(blocos)-1}")
        
        for bloco in blocos[1:]:  # Pular primeiro vazio
            try:
                # Extrair campos usando regex
                linhas = bloco.strip().split('\n')
                
                precatorio = {}
                
                # Primeira linha é a devedora
                precatorio['Devedora'] = linhas[0].strip()
                
                # Percorrer linhas procurando campos
                texto_bloco = '\n'.join(linhas)
                
                # Ordem de Pagamento
                match = re.search(r'Ordem de Pagamento:\s*(\d+)', texto_bloco)
                if match:
                    precatorio['Ordem de Pagamento'] = match.group(1)
                
                # Nº Processo DEPRE
                match = re.search(r'Nº Processo DEPRE:\s*([\d\-\.]+)', texto_bloco)
                if match:
                    precatorio['Nº Processo DEPRE'] = match.group(1)
                
                # Natureza
                match = re.search(r'Natureza:\s*(\w+)', texto_bloco)
                if match:
                    precatorio['Natureza'] = match.group(1)
                
                # Nº de autos
                match = re.search(r'Nº de autos:\s*([\d\-\.]+)', texto_bloco)
                if match:
                    precatorio['Nº de autos'] = match.group(1)
                
                # Ordem Orçamentária
                match = re.search(r'Ordem Orçamentária:\s*([\d/]+)', texto_bloco)
                if match:
                    precatorio['Ordem Orçamentária'] = match.group(1)
                
                # Suspenso?
                match = re.search(r'Suspenso\?:\s*([SN])', texto_bloco)
                if match:
                    precatorio['Suspenso?'] = match.group(1)
                
                # Data do Protocolo
                match = re.search(r'Data do Protocolo:\s*([\d/:\s\.]+)', texto_bloco)
                if match:
                    precatorio['Data do Protocolo'] = match.group(1).strip()
                
                # Advogado(s)
                match = re.search(r'Advogado\(s\):\s*(.+?)(?=Devedora:|$)', texto_bloco, re.DOTALL)
                if match:
                    advogados = match.group(1).strip()
                    # Limpar quebras de linha extras
                    advogados = ' '.join(advogados.split())
                    precatorio['Advogado(s)'] = advogados
                
                if precatorio.get('Ordem de Pagamento'):
                    precatorios.append(precatorio)
                    
            except Exception as e:
                print(f"   ⚠️  Erro ao processar bloco: {str(e)[:50]}")
                continue
    
    return precatorios

def criar_planilha_excel(precatorios, arquivo_saida):
    """Cria planilha Excel formatada"""
    
    print(f"\n📊 Criando planilha Excel...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Precatórios"
    
    # Cabeçalhos
    headers = [
        'Devedora',
        'Ordem de Pagamento',
        'Nº Processo DEPRE',
        'Natureza',
        'Nº de autos',
        'Ordem Orçamentária',
        'Suspenso?',
        'Data do Protocolo',
        'Advogado(s)'
    ]
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Escrever cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Escrever dados
    for row_idx, prec in enumerate(precatorios, 2):
        ws.cell(row=row_idx, column=1, value=prec.get('Devedora', ''))
        ws.cell(row=row_idx, column=2, value=prec.get('Ordem de Pagamento', ''))
        ws.cell(row=row_idx, column=3, value=prec.get('Nº Processo DEPRE', ''))
        ws.cell(row=row_idx, column=4, value=prec.get('Natureza', ''))
        ws.cell(row=row_idx, column=5, value=prec.get('Nº de autos', ''))
        ws.cell(row=row_idx, column=6, value=prec.get('Ordem Orçamentária', ''))
        ws.cell(row=row_idx, column=7, value=prec.get('Suspenso?', ''))
        ws.cell(row=row_idx, column=8, value=prec.get('Data do Protocolo', ''))
        ws.cell(row=row_idx, column=9, value=prec.get('Advogado(s)', ''))
    
    # Ajustar larguras
    larguras = {
        'A': 45,  # Devedora
        'B': 18,  # Ordem de Pagamento
        'C': 30,  # Nº Processo DEPRE
        'D': 15,  # Natureza
        'E': 30,  # Nº de autos
        'F': 20,  # Ordem Orçamentária
        'G': 12,  # Suspenso?
        'H': 22,  # Data do Protocolo
        'I': 35   # Advogado(s)
    }
    
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura
    
    # Altura da primeira linha
    ws.row_dimensions[1].height = 30
    
    # Filtros
    ws.auto_filter.ref = ws.dimensions
    
    # Salvar
    wb.save(arquivo_saida)
    
    print(f"   ✅ Planilha salva: {arquivo_saida}")
    print(f"   📊 Total de precatórios: {len(precatorios)}")

if __name__ == "__main__":
    print("="*70)
    print("📊 GERAR PLANILHA DE PRECATÓRIOS")
    print("="*70)
    
    arquivo_pdf = "099--Lista-de-Prectorios--22000_22300.pdf"
    
    try:
        # Extrair precatórios
        precatorios = extrair_precatorios_pdf(arquivo_pdf)
        
        if len(precatorios) == 0:
            print("\n❌ Nenhum precatório encontrado!")
        else:
            print(f"\n✅ {len(precatorios)} precatórios extraídos")
            
            # Gerar nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo_saida = f"planilha_precatorios_{timestamp}.xlsx"
            
            # Criar planilha
            criar_planilha_excel(precatorios, arquivo_saida)
            
            print("\n" + "="*70)
            print("🎉 PLANILHA CRIADA COM SUCESSO!")
            print("="*70)
            
            # Mostrar preview
            print(f"\n📋 PREVIEW (primeiros 5):")
            for i, prec in enumerate(precatorios[:5], 1):
                print(f"\n   {i}. OP: {prec.get('Ordem de Pagamento')}")
                print(f"      DEPRE: {prec.get('Nº Processo DEPRE')}")
                print(f"      Devedora: {prec.get('Devedora', '')[:50]}")
            
            if len(precatorios) > 5:
                print(f"\n   ... e mais {len(precatorios) - 5}")
            
            print(f"\n📁 Arquivo: {arquivo_saida}")
        
    except FileNotFoundError:
        print(f"\n❌ PDF não encontrado: {arquivo_pdf}")
    
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n✅ FIM!")
