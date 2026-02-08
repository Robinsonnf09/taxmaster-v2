"""
EXTRATOR SIMPLES - Nº PROCESSO DEPRE
"""

import PyPDF2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from datetime import datetime
import os

def extrair_depres_pdf(arquivo_pdf):
    """Extrai todos os números DEPRE"""
    
    print(f"\n📄 Lendo PDF: {arquivo_pdf}")
    
    with open(arquivo_pdf, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        
        total_paginas = len(reader.pages)
        print(f"   📋 Páginas: {total_paginas}")
        
        texto_completo = ""
        for idx, pagina in enumerate(reader.pages, 1):
            print(f"   📖 Página {idx}/{total_paginas}...", end="\r")
            texto_completo += pagina.extract_text() + "\n"
        
        print(f"\n   ✅ Texto extraído!")
        
        print(f"   🔍 Procurando DEPRE...")
        
        pattern = r'Nº Processo DEPRE:\s*([\d\-\.]+)'
        matches = re.findall(pattern, texto_completo)
        
        depres = []
        vistos = set()
        for depre in matches:
            if depre not in vistos:
                depres.append(depre)
                vistos.add(depre)
        
        print(f"   ✅ {len(depres)} encontrados!")
    
    return depres

def criar_planilha(depres, arquivo_saida):
    """Cria planilha Excel"""
    
    print(f"\n📊 Criando Excel...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DEPRE"
    
    ws['A1'] = 'Nº Processo DEPRE'
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = header_alignment
    
    for row_idx, depre in enumerate(depres, 2):
        ws.cell(row=row_idx, column=1, value=depre)
    
    ws.column_dimensions['A'].width = 40
    ws.row_dimensions[1].height = 25
    ws.auto_filter.ref = f'A1:A{len(depres)+1}'
    
    wb.save(arquivo_saida)
    
    print(f"   ✅ Salvo: {arquivo_saida}")

if __name__ == "__main__":
    print("="*70)
    print("📊 EXTRATOR DE PROCESSOS DEPRE")
    print("="*70)
    
    # Procurar PDF na pasta
    import glob
    pdfs = glob.glob("*.pdf")
    
    if not pdfs:
        print("\n❌ Nenhum PDF encontrado na pasta!")
        print(f"   Pasta atual: {os.getcwd()}")
        input("\nENTER para sair...")
        exit()
    
    print(f"\n📁 PDFs encontrados:")
    for i, pdf in enumerate(pdfs, 1):
        print(f"   {i}. {pdf}")
    
    # Usar o primeiro PDF ou perguntar
    if len(pdfs) == 1:
        arquivo_pdf = pdfs[0]
        print(f"\n✅ Usando: {arquivo_pdf}")
    else:
        escolha = input(f"\nEscolha o número do PDF (1-{len(pdfs)}): ")
        try:
            idx = int(escolha) - 1
            arquivo_pdf = pdfs[idx]
        except:
            arquivo_pdf = pdfs[0]
    
    try:
        depres = extrair_depres_pdf(arquivo_pdf)
        
        if len(depres) == 0:
            print("\n❌ Nenhum DEPRE encontrado!")
        else:
            print(f"\n" + "="*70)
            print(f"✅ TOTAL: {len(depres)} PROCESSOS DEPRE")
            print("="*70)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo_saida = f"lista_depre_{timestamp}.xlsx"
            
            criar_planilha(depres, arquivo_saida)
            
            print("\n" + "="*70)
            print("🎉 PLANILHA CRIADA!")
            print("="*70)
            
            print(f"\n📋 PREVIEW (primeiros 20):")
            for i, depre in enumerate(depres[:20], 1):
                print(f"   {i:3}. {depre}")
            
            if len(depres) > 20:
                print(f"\n   ... e mais {len(depres) - 20}")
            
            print(f"\n📁 Arquivo: {arquivo_saida}")
            print(f"📊 Total: {len(depres)} processos")
            
            print(f"\n📂 Pasta: {os.path.abspath(arquivo_saida)}")
        
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nENTER para fechar...")
    
    print("\n✅ FIM!")
