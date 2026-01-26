"""
Gera planilha modelo para cadastro no PUSH
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def criar_planilha_push():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos PUSH"
    
    # Cabeçalho
    header_fill = PatternFill(start_color="2C8FA0", end_color="2C8FA0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws['A1'] = "numero_processo"
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['B1'] = "tribunal"
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Exemplos (FICTÍCIOS - substitua por reais)
    exemplos = [
        ["0000001-00.2020.4.01.3800", "TRF1"],
        ["0000002-00.2021.4.01.3800", "TRF1"],
        ["0000003-00.2022.4.01.3800", "TRF1"],
    ]
    
    for idx, (processo, tribunal) in enumerate(exemplos, 2):
        ws[f'A{idx}'] = processo
        ws[f'B{idx}'] = tribunal
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    
    filename = f'processos_push_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    
    print(f"✅ Planilha criada: {filename}")
    return filename

if __name__ == "__main__":
    criar_planilha_push()
