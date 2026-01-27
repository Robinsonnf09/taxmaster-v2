"""
BUSCAR OFÍCIOS REQUISITÓRIOS - PROCESSOS DEPRE DO PDF
Lê a planilha de precatórios e busca todos os ofícios
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
import time
import os
import openpyxl
from datetime import datetime
import requests

class BuscadorOficiosPrecatorios:
    
    def __init__(self):
        self.driver = None
        self.session = None
        self.pasta_oficios = "oficios_precatorios_pdf"
        
        if not os.path.exists(self.pasta_oficios):
            os.makedirs(self.pasta_oficios)
        
        self.sucessos = []
        self.falhas = []
        self.sem_oficio = []
    
    def iniciar(self):
        print("\n🌐 Iniciando Chrome...")
        
        chrome_options = Options()
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-popup-blocking')
        
        prefs = {
            "download.default_directory": os.path.abspath(self.pasta_oficios),
            "download.prompt_for_download": False,
            "plugins.always_open_pdf_externally": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        service = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        
        self.session = requests.Session()
        
        print("✅ Chrome iniciado!")
        print(f"📁 PDFs em: {os.path.abspath(self.pasta_oficios)}")
    
    def fazer_login_unico(self):
        print(f"\n🔐 Login único...")
        self.driver.get("https://esaj.tjsp.jus.br")
        time.sleep(3)
        
        print("\n" + "="*70)
        print("⚠️  FAÇA LOGIN (1 VEZ):")
        print("="*70)
        print("   ❌ Se aparecer popup 'WEB SIGNER' → CANCELAR")
        print("   ✅ Use LOGIN e SENHA (não certificado)")
        print("="*70)
        
        input("\n>>> ENTER após login <<<\n")
        
        for cookie in self.driver.get_cookies():
            self.session.cookies.set(cookie['name'], cookie['value'])
        
        print("✅ Login salvo!")
        return True
    
    def buscar_processo_automatico(self, numero_processo):
        try:
            url_consulta = "https://esaj.tjsp.jus.br/cpopg/open.do"
            self.driver.get(url_consulta)
            time.sleep(1.5)
            
            wait = WebDriverWait(self.driver, 10)
            
            # Clicar no radio "Número Antigo"
            try:
                radio = wait.until(EC.element_to_be_clickable(
                    (By.ID, "radioNumeroAntigo")
                ))
                radio.click()
                time.sleep(0.5)
            except:
                pass
            
            # Preencher campo
            try:
                campo = wait.until(EC.visibility_of_element_located(
                    (By.ID, "nuProcessoAntigoFormatado")
                ))
            except:
                return None
            
            campo.clear()
            time.sleep(0.2)
            campo.send_keys(numero_processo)
            time.sleep(0.3)
            campo.send_keys(Keys.RETURN)
            
            time.sleep(2.5)
            
            url_atual = self.driver.current_url
            
            if "processo.codigo=" in url_atual:
                codigo = url_atual.split("processo.codigo=")[1].split("&")[0]
                return codigo
            
            return None
            
        except:
            return None
    
    def extrair_foro(self, numero_processo):
        foro_completo = numero_processo.split(".")[-1]
        foro = foro_completo.lstrip('0')
        return foro if foro else "0"
    
    def baixar_pdf_automatico(self, url_pdf, nome_arquivo):
        try:
            for cookie in self.driver.get_cookies():
                self.session.cookies.set(cookie['name'], cookie['value'])
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Referer': 'https://esaj.tjsp.jus.br/'
            }
            
            response = self.session.get(url_pdf, headers=headers, timeout=30)
            
            if response.status_code == 200 and len(response.content) > 1000:
                caminho = os.path.join(self.pasta_oficios, nome_arquivo)
                
                with open(caminho, 'wb') as f:
                    f.write(response.content)
                
                return True, len(response.content)
            
            return False, 0
            
        except:
            return False, 0
    
    def processar_processo_automatico(self, numero_processo, idx, total, ordem_pagamento=""):
        try:
            print(f"\n{'='*70}")
            print(f"⚡ [{idx}/{total}] {numero_processo}")
            if ordem_pagamento:
                print(f"   📋 OP: {ordem_pagamento}")
            print(f"{'='*70}")
            
            print(f"   🔍 Código...", end=" ", flush=True)
            
            codigo = self.buscar_processo_automatico(numero_processo)
            
            if not codigo:
                print(f"❌")
                self.falhas.append(numero_processo)
                return False
            
            print(f"✅ {codigo}")
            
            foro = self.extrair_foro(numero_processo)
            
            url_requisitorios = (
                f"https://esaj.tjsp.jus.br/cpopg/show.do?"
                f"processo.codigo={codigo}&"
                f"processo.foro={foro}&"
                f"processo.numero={numero_processo}&"
                f"consultaDeRequisitorios=true"
            )
            
            print(f"   🎯 Requisitórios...", end=" ", flush=True)
            self.driver.get(url_requisitorios)
            time.sleep(2)
            print(f"✅")
            
            print(f"   🔍 Ofícios...", end=" ", flush=True)
            
            script = """
            let links = [];
            document.querySelectorAll('a').forEach(a => {
                let texto = a.textContent.toLowerCase();
                let href = a.href;
                
                if ((texto.includes('ofício') || texto.includes('requisitório') || 
                     texto.includes('or') || texto.includes('depre')) 
                    && href && href.length > 0 && !href.includes('javascript')) {
                    links.push({
                        url: href,
                        texto: a.textContent.trim()
                    });
                }
            });
            return links;
            """
            
            oficios = self.driver.execute_script(script)
            
            if not oficios or len(oficios) == 0:
                print(f"⚠️")
                self.sem_oficio.append(numero_processo)
                return False
            
            print(f"✅ {len(oficios)}")
            
            baixados = 0
            
            for idx_of, oficio in enumerate(oficios, 1):
                # Nome com ordem de pagamento se disponível
                nome_limpo = numero_processo.replace('-','').replace('.','')
                if ordem_pagamento:
                    nome_arquivo = f"OP{ordem_pagamento}_{nome_limpo}_of{idx_of}.pdf"
                else:
                    nome_arquivo = f"{nome_limpo}_of{idx_of}.pdf"
                
                print(f"   📥 {idx_of}/{len(oficios)}...", end=" ", flush=True)
                
                sucesso, tamanho = self.baixar_pdf_automatico(
                    oficio['url'], 
                    nome_arquivo
                )
                
                if sucesso:
                    kb = tamanho // 1024
                    print(f"✅ {kb}KB")
                    baixados += 1
                else:
                    print(f"❌")
            
            if baixados > 0:
                self.sucessos.append(numero_processo)
                return True
            else:
                self.falhas.append(numero_processo)
                return False
            
        except Exception as e:
            print(f"   ❌")
            self.falhas.append(numero_processo)
            return False
    
    def processar_planilha_precatorios(self, arquivo):
        print(f"\n📊 Carregando planilha de precatórios...")
        
        if not os.path.exists(arquivo):
            print(f"❌ Arquivo não encontrado: {arquivo}")
            return
        
        wb = openpyxl.load_workbook(arquivo)
        ws = wb.active
        
        # Mapear colunas
        headers = [cell.value for cell in ws[1]]
        
        col_depre = None
        col_op = None
        
        for idx, header in enumerate(headers, 1):
            if header and 'DEPRE' in str(header).upper():
                col_depre = idx
            if header and 'ORDEM' in str(header).upper() and 'PAGAMENTO' in str(header).upper():
                col_op = idx
        
        if not col_depre:
            print(f"❌ Coluna 'Nº Processo DEPRE' não encontrada!")
            return
        
        print(f"✅ Coluna DEPRE: {col_depre}")
        if col_op:
            print(f"✅ Coluna Ordem de Pagamento: {col_op}")
        
        # Extrair processos
        processos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= col_depre:
                depre = row[col_depre - 1]
                op = row[col_op - 1] if col_op and len(row) >= col_op else ""
                
                if depre and str(depre).strip():
                    num_depre = str(depre).strip()
                    num_op = str(op).strip() if op else ""
                    
                    # Verificar se é processo TJSP
                    if '.8.26.' in num_depre:
                        processos.append({
                            'depre': num_depre,
                            'op': num_op
                        })
        
        total = len(processos)
        print(f"✅ {total} processos DEPRE encontrados")
        
        if total == 0:
            print(f"⚠️  Nenhum processo válido encontrado!")
            return
        
        # Mostrar alguns exemplos
        print(f"\n📋 Primeiros processos:")
        for i, p in enumerate(processos[:5], 1):
            print(f"   {i}. {p['depre']}", end="")
            if p['op']:
                print(f" (OP: {p['op']})")
            else:
                print()
        
        if total > 5:
            print(f"   ... e mais {total - 5}")
        
        print("\n" + "="*70)
        print("⚡ PROCESSAMENTO DE PRECATÓRIOS")
        print("="*70)
        
        confirma = input(f"\n>>> Buscar ofícios de {total} processos? (s/n): ").lower()
        
        if confirma != 's':
            return
        
        inicio = datetime.now()
        
        for idx, proc in enumerate(processos, 1):
            self.processar_processo_automatico(
                proc['depre'], 
                idx, 
                total,
                proc['op']
            )
            
            if idx % 10 == 0:
                print(f"\n{'='*70}")
                print(f"📊 {idx}/{total} ({idx/total*100:.1f}%)")
                print(f"   ✅ {len(self.sucessos)} | ⚠️  {len(self.sem_oficio)} | ❌ {len(self.falhas)}")
                
                decorrido = (datetime.now() - inicio).total_seconds()
                media = decorrido / idx
                restante = (total - idx) * media
                
                print(f"   ⏱️  {int(decorrido/60)}min | ⏳ ~{int(restante/60)}min")
                print(f"{'='*70}\n")
            
            time.sleep(0.8)
        
        self.gerar_relatorio_final(inicio)
    
    def gerar_relatorio_final(self, inicio):
        fim = datetime.now()
        duracao = fim - inicio
        
        print("\n" + "="*70)
        print("🎉 CONCLUÍDO!")
        print("="*70)
        
        total = len(self.sucessos) + len(self.sem_oficio) + len(self.falhas)
        
        print(f"\n   Total: {total}")
        print(f"   ✅ Com ofício: {len(self.sucessos)}")
        print(f"   ⚠️  Sem ofício: {len(self.sem_oficio)}")
        print(f"   ❌ Falhas: {len(self.falhas)}")
        
        if total > 0:
            taxa = (len(self.sucessos) / total) * 100
            print(f"   📊 Taxa: {taxa:.1f}%")
        
        print(f"\n   ⏱️  {int(duracao.total_seconds()/60)} min")
        
        pdfs = [f for f in os.listdir(self.pasta_oficios) if f.endswith('.pdf')]
        print(f"   📄 {len(pdfs)} PDFs baixados")
        print(f"\n📁 {os.path.abspath(self.pasta_oficios)}")
        
        if self.sem_oficio:
            print(f"\n⚠️  SEM OFÍCIO ({len(self.sem_oficio)}):")
            for p in self.sem_oficio[:20]:
                print(f"   {p}")
        
        if self.falhas:
            print(f"\n❌ FALHAS ({len(self.falhas)}):")
            for p in self.falhas[:20]:
                print(f"   {p}")
        
        print("="*70)
        
        # Salvar relatório
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        relatorio = f"relatorio_precatorios_{timestamp}.txt"
        
        with open(relatorio, "w", encoding="utf-8") as f:
            f.write("RELATÓRIO - OFÍCIOS DE PRECATÓRIOS\n")
            f.write("="*70 + "\n\n")
            f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"Duração: {int(duracao.total_seconds()/60)} minutos\n\n")
            f.write(f"Total: {total}\n")
            f.write(f"Com ofício: {len(self.sucessos)}\n")
            f.write(f"Sem ofício: {len(self.sem_oficio)}\n")
            f.write(f"Falhas: {len(self.falhas)}\n")
            f.write(f"PDFs: {len(pdfs)}\n\n")
            
            if self.sucessos:
                f.write("COM OFÍCIO:\n")
                for p in self.sucessos:
                    f.write(f"{p}\n")
                f.write("\n")
            
            if self.sem_oficio:
                f.write("SEM OFÍCIO:\n")
                for p in self.sem_oficio:
                    f.write(f"{p}\n")
                f.write("\n")
            
            if self.falhas:
                f.write("FALHAS:\n")
                for p in self.falhas:
                    f.write(f"{p}\n")
        
        print(f"\n📄 Relatório: {relatorio}")
    
    def fechar(self):
        if self.driver:
            self.driver.quit()

if __name__ == "__main__":
    print("="*70)
    print("⚡ BUSCAR OFÍCIOS - PRECATÓRIOS DO PDF")
    print("="*70)
    
    buscador = BuscadorOficiosPrecatorios()
    
    try:
        # Procurar planilha gerada
        import glob
        planilhas = glob.glob("*precatorios*.xlsx")
        
        if not planilhas:
            print("\n❌ Planilha de precatórios não encontrada!")
            print("   Certifique-se que a planilha foi gerada")
            input("\nENTER para sair...")
        else:
            arquivo = planilhas[0]
            print(f"\n✅ Planilha encontrada: {arquivo}")
            
            input("\nENTER para começar...\n")
            
            buscador.iniciar()
            
            if buscador.fazer_login_unico():
                buscador.processar_planilha_precatorios(arquivo)
            
            input("\n\nENTER para fechar...\n")
        
    except KeyboardInterrupt:
        print("\n⚠️  Interrompido")
        if len(buscador.sucessos) > 0:
            buscador.gerar_relatorio_final(datetime.now())
    
    except Exception as e:
        print(f"\n❌ {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        buscador.fechar()
    
    print("\n✅ FIM!")
