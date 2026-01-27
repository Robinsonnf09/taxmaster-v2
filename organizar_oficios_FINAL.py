"""
ORGANIZAR 472 PDFs ORIGINAIS POR ÓRGÃO DEVEDOR
"""

import os
import shutil
import pdfplumber
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

class OrganizadorFinal:
    
    def __init__(self):
        self.pasta_origem = "oficios_requisitorios_tjsp"
        self.pasta_destino = "oficios_por_orgao_FINAL"
        self.estatisticas = defaultdict(lambda: {'quantidade': 0, 'arquivos': []})
        
        if not os.path.exists(self.pasta_destino):
            os.makedirs(self.pasta_destino)
    
    def extrair_orgao_devedor(self, pdf_path):
        """Extrai órgão devedor usando pdfplumber"""
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                texto = ""
                
                # Ler até 3 páginas
                for i in range(min(3, len(pdf.pages))):
                    texto += pdf.pages[i].extract_text() + "\n"
                
                if not texto or len(texto.strip()) < 30:
                    return 'PDF_Vazio'
                
                texto_lower = texto.lower()
                
                # === ESTADO DE SÃO PAULO ===
                if 'fazenda do estado' in texto_lower or 'secretaria da fazenda' in texto_lower:
                    return 'Estado_SP_Fazenda'
                
                if 'secretaria' in texto_lower and 'educação' in texto_lower:
                    return 'Estado_SP_Educacao'
                elif 'secretaria' in texto_lower and 'saúde' in texto_lower:
                    return 'Estado_SP_Saude'
                elif 'secretaria' in texto_lower and 'segurança' in texto_lower:
                    return 'Estado_SP_Seguranca'
                
                if 'estado de são paulo' in texto_lower or 'governo do estado' in texto_lower:
                    return 'Estado_SP_Geral'
                
                # === AUTARQUIAS ===
                if 'ipesp' in texto_lower:
                    return 'Autarquia_IPESP'
                elif 'spprev' in texto_lower:
                    return 'Autarquia_SPPrev'
                elif 'sabesp' in texto_lower:
                    return 'Autarquia_SABESP'
                elif 'cdhu' in texto_lower:
                    return 'Autarquia_CDHU'
                
                # === UNIVERSIDADES ===
                if 'usp' in texto_lower or 'universidade de são paulo' in texto_lower:
                    return 'Universidade_USP'
                elif 'unesp' in texto_lower:
                    return 'Universidade_UNESP'
                elif 'unicamp' in texto_lower:
                    return 'Universidade_UNICAMP'
                
                # === MUNICÍPIOS ===
                match_mun = re.search(
                    r'município\s+(?:de\s+)?([A-ZÀÁÂÃÇ][a-zàáâãçéêíóôõú\s]{2,30})|'
                    r'prefeitura\s+(?:municipal\s+)?(?:de\s+)?([A-ZÀÁÂÃÇ][a-zàáâãçéêíóôõú\s]{2,30})',
                    texto,
                    re.IGNORECASE
                )
                
                if match_mun:
                    municipio = (match_mun.group(1) or match_mun.group(2)).strip().title()
                    municipio = ' '.join(municipio.split())
                    return f'Municipio_{municipio.replace(" ", "_")}'
                
                # === PODER JUDICIÁRIO ===
                if 'tribunal de justiça' in texto_lower or 'poder judiciário' in texto_lower:
                    return 'Poder_Judiciario'
                
                # Se tem bastante texto mas não identificou
                if len(texto.strip()) > 100:
                    return 'Outros_Orgaos'
                
                return 'Nao_Identificado'
                
        except Exception as e:
            return 'Erro_Leitura'
    
    def organizar(self):
        """Organiza os PDFs"""
        
        print(f"\n📂 Organizando 472 PDFs...")
        print(f"   De: {self.pasta_origem}")
        print(f"   Para: {self.pasta_destino}")
        
        pdfs = [f for f in os.listdir(self.pasta_origem) if f.endswith('.pdf')]
        total = len(pdfs)
        
        print(f"\n   📄 Total: {total} PDFs")
        print(f"\n🔍 Processando...")
        
        for idx, pdf in enumerate(pdfs, 1):
            if idx % 10 == 0 or idx == 1:
                print(f"   📋 {idx}/{total} ({idx/total*100:.1f}%) - {len(self.estatisticas)} categorias", end="\r")
            
            try:
                pdf_path = os.path.join(self.pasta_origem, pdf)
                
                # Extrair órgão
                orgao = self.extrair_orgao_devedor(pdf_path)
                
                # Criar pasta
                pasta_orgao = os.path.join(self.pasta_destino, orgao)
                if not os.path.exists(pasta_orgao):
                    os.makedirs(pasta_orgao)
                
                # Copiar
                destino = os.path.join(pasta_orgao, pdf)
                
                if os.path.exists(destino):
                    base, ext = os.path.splitext(pdf)
                    contador = 1
                    while os.path.exists(os.path.join(pasta_orgao, f"{base}_{contador}{ext}")):
                        contador += 1
                    destino = os.path.join(pasta_orgao, f"{base}_{contador}{ext}")
                
                shutil.copy2(pdf_path, destino)
                
                # Estatísticas
                self.estatisticas[orgao]['quantidade'] += 1
                self.estatisticas[orgao]['arquivos'].append(pdf)
                
            except Exception as e:
                continue
        
        print(f"\n   ✅ {total} PDFs processados!")
        
        return True
    
    def gerar_relatorio(self):
        """Gera relatório"""
        
        print(f"\n📊 Gerando relatório...")
        
        ordenados = sorted(
            self.estatisticas.items(),
            key=lambda x: x[1]['quantidade'],
            reverse=True
        )
        
        print(f"\n" + "="*80)
        print(f"📊 DISTRIBUIÇÃO POR ÓRGÃO DEVEDOR")
        print("="*80)
        
        total = sum(d['quantidade'] for _, d in ordenados)
        
        print(f"\n   #    Orgao Devedor                                     Qtd      %")
        print(f"   ---  ------------------------------------------------  ----  ------")
        
        for idx, (orgao, dados) in enumerate(ordenados, 1):
            perc = (dados['quantidade'] / total) * 100
            orgao_nome = orgao.replace('_', ' ')
            print(f"   {idx:>3}  {orgao_nome:<48}  {dados['quantidade']:>4}  {perc:>5.1f}%")
        
        print(f"\n        {'TOTAL':<48}  {total:>4}  100.0%")
        
        # Criar planilha
        self.criar_planilha(ordenados)
    
    def criar_planilha(self, ordenados):
        """Cria planilha Excel"""
        
        print(f"\n📊 Criando planilha...")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Distribuição Final"
        
        headers = ['#', 'Órgão Devedor', 'Quantidade', '%']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = Alignment(horizontal="center")
        
        total = sum(d['quantidade'] for _, d in ordenados)
        
        for idx, (orgao, dados) in enumerate(ordenados, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=orgao.replace('_', ' '))
            ws.cell(row=idx+1, column=3, value=dados['quantidade'])
            
            perc = (dados['quantidade'] / total) * 100
            ws.cell(row=idx+1, column=4, value=f"{perc:.1f}%")
        
        row_total = len(ordenados) + 2
        ws.cell(row=row_total, column=2, value='TOTAL').font = Font(bold=True)
        ws.cell(row=row_total, column=3, value=total).font = Font(bold=True)
        ws.cell(row=row_total, column=4, value='100%').font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        
        ws.row_dimensions[1].height = 25
        ws.auto_filter.ref = f'A1:D{len(ordenados)+1}'
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo = f"distribuicao_oficios_FINAL_{timestamp}.xlsx"
        
        wb.save(arquivo)
        
        print(f"   ✅ {arquivo}")

if __name__ == "__main__":
    print("="*80)
    print("📂 ORGANIZAR 472 PDFs POR ÓRGÃO DEVEDOR")
    print("="*80)
    
    org = OrganizadorFinal()
    
    confirma = input(f"\nOrganizar {len([f for f in os.listdir(org.pasta_origem) if f.endswith('.pdf')])} PDFs? (s/n): ").lower()
    
    if confirma != 's':
        print("\n❌ Cancelado")
        exit()
    
    try:
        inicio = datetime.now()
        
        if org.organizar():
            org.gerar_relatorio()
            
            fim = datetime.now()
            duracao = int((fim - inicio).total_seconds() / 60)
            
            print("\n" + "="*80)
            print("🎉 ORGANIZAÇÃO CONCLUÍDA!")
            print("="*80)
            print(f"\n⏱️  {duracao} minutos")
            print(f"📂 {org.pasta_destino}/")
            print(f"📊 {len(org.estatisticas)} categorias")
        
    except Exception as e:
        print(f"\n❌ {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nENTER...\n")
    
    print("\n✅ FIM!")
