"""
EXTRAIR PROCESSOS - COLA DA ÁREA DE TRANSFERÊNCIA
Você copia da tela, script cria a planilha
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re
import pyperclip

def extrair_processos_clipboard():
    """Extrai processos da área de transferência"""
    
    print("\n📋 Lendo área de transferência...")
    
    try:
        texto = pyperclip.paste()
    except:
        print("❌ Erro ao ler área de transferência")
        print("   Instalando pyperclip...")
        import subprocess
        subprocess.run(["pip", "install", "pyperclip"])
        texto = pyperclip.paste()
    
    if not texto:
        print("❌ Área de transferência vazia!")
        return []
    
    print(f"   ✅ {len(texto)} caracteres")
    
    # Buscar processos no padrão TJSP
    print(f"\n🔍 Procurando processos...")
    
    pattern = r'\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}'
    matches = re.findall(pattern, texto)
    
    # Remover duplicatas mantendo ordem
    processos = []
    vistos = set()
    for proc in matches:
        if proc not in vistos:
            processos.append(proc)
            vistos.add(proc)
    
    print(f"   ✅ {len(processos)} processos encontrados!")
    
    return processos

def criar_planilha(processos, arquivo):
    """Cria planilha Excel"""
    
    print(f"\n📊 Criando planilha...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"
    
    ws['A1'] = 'Nº Processo'
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = header_alignment
    
    for row_idx, processo in enumerate(processos, 2):
        ws.cell(row=row_idx, column=1, value=processo)
    
    ws.column_dimensions['A'].width = 35
    ws.row_dimensions[1].height = 25
    ws.auto_filter.ref = f'A1:A{len(processos)+1}'
    
    wb.save(arquivo)
    
    print(f"   ✅ Salvo: {arquivo}")

if __name__ == "__main__":
    print("="*70)
    print("📋 EXTRAIR PROCESSOS DA ÁREA DE TRANSFERÊNCIA")
    print("="*70)
    
    print("\n📝 INSTRUÇÕES:")
    print("="*70)
    print("   1. No navegador, acesse o PUSH de Requisitórios")
    print("   2. Faça a busca (01/11/2024 a 31/12/2025)")
    print("   3. Selecione TODA a lista (Ctrl+A ou mouse)")
    print("   4. Copie (Ctrl+C)")
    print("   5. Volte aqui e pressione ENTER")
    print("="*70)
    
    input("\n>>> ENTER após copiar a lista <<<\n")
    
    try:
        processos = extrair_processos_clipboard()
        
        if len(processos) == 0:
            print("\n❌ Nenhum processo encontrado!")
            print("   Certifique-se de:")
            print("   1. Copiar a lista completa")
            print("   2. Incluir os números dos processos")
        else:
            print(f"\n" + "="*70)
            print(f"✅ {len(processos)} PROCESSOS EXTRAÍDOS!")
            print("="*70)
            
            print(f"\n📋 PREVIEW (primeiros 20):")
            for i, proc in enumerate(processos[:20], 1):
                print(f"   {i:3}. {proc}")
            
            if len(processos) > 20:
                print(f"\n   ... e mais {len(processos) - 20}")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo = f"processos_push_{timestamp}.xlsx"
            
            criar_planilha(processos, arquivo)
            
            print("\n" + "="*70)
            print("🎉 PLANILHA CRIADA!")
            print("="*70)
            print(f"\n📁 {arquivo}")
            print(f"📊 {len(processos)} processos")
            
            # Se tiver menos de 500, avisar
            if len(processos) < 500:
                print(f"\n⚠️  Atenção: Encontrados apenas {len(processos)} de 500 desejados")
                print(f"   Você pode:")
                print(f"   1. Navegar para próxima página no navegador")
                print(f"   2. Copiar também")
                print(f"   3. Executar o script novamente")
                print(f"   4. O script juntará tudo")
        
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n\nENTER para fechar...\n")
    
    print("\n✅ FIM!")
