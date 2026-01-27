"""
GERAR PLANILHA SIMPLES - NºPROCESSO + DEPRE
Apenas 2 colunas essenciais
"""

import PyPDF2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
from datetime import datetime

def extrair_processos_pdf(arquivo_pdf):
    """Extrai Nº Processo e DEPRE do PDF"""
    
    print(f"\n📄 Lendo PDF: {arquivo_pdf}")
    
    processos = []
    
    with open(arquivo_pdf, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        
        print(f"   Total de páginas: {len(reader.pages)}")
        
        texto_completo = ""
        for pagina in reader.pages:
            texto_completo += pagina.extract_text()
        
        # Procurar padrões
        # Nº de autos: XXXXXXX-XX.XXXX.X.XX.XXXX
        # Nº Processo DEPRE: XXXXXXX-XX.XXXX.X.XX.XXXX
        
        # Dividir por blocos
        blocos = texto_completo.split("Devedora:")
        
        print(f"   Blocos encontrados: {len(blocos)-1}")
        
        for bloco in blocos[1:]:
            try:
                # Nº de autos (processo original)
                match_autos = re.search(r'Nº de autos:\s*([\d\-\.]+)', bloco)
                
                # Nº Processo DEPRE
                match_depre = re.search(r'Nº Processo DEPRE:\s*([\d\-\.]+)', bloco)
                
                if match_depre:
                    processo = {
                        'Nº Processo': match_autos.group(1) if match_autos else '',
                        'DEPRE': match_depre.group(1)
                    }
                    
                    processos.append(processo)
                    
            except Exception as e:
                continue
    
    return processos

def criar_planilha_simples(processos, arquivo_saida):
    """Cria planilha Excel com 2 colunas"""
    
    print(f"\n📊 Criando planilha Excel...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Processos"
    
    # Cabeçalhos
    headers = ['Nº Processo', 'DEPRE']
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escrever cabeçalhos
    ws['A1'] = headers[0]
    ws['B1'] = headers[1]
    
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = header_alignment
    
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = header_alignment
    
    # Escrever dados
    for row_idx, proc in enumerate(processos, 2):
        ws.cell(row=row_idx, column=1, value=proc['Nº Processo'])
        ws.cell(row=row_idx, column=2, value=proc['DEPRE'])
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    
    # Altura da primeira linha
    ws.row_dimensions[1].height = 25
    
    # Filtros
    ws.auto_filter.ref = f'A1:B{len(processos)+1}'
    
    # Salvar
    wb.save(arquivo_saida)
    
    print(f"   ✅ Planilha salva: {arquivo_saida}")
    print(f"   📊 Total de processos: {len(processos)}")

if __name__ == "__main__":
    print("="*70)
    print("📊 GERAR PLANILHA SIMPLES - NºPROCESSO + DEPRE")
    print("="*70)
    
    arquivo_pdf = "099--Lista-de-Prectorios--22000_22300.pdf"
    
    try:
        # Extrair processos
        processos = extrair_processos_pdf(arquivo_pdf)
        
        if len(processos) == 0:
            print("\n❌ Nenhum processo encontrado!")
        else:
            print(f"\n✅ {len(processos)} processos extraídos")
            
            # Nome do arquivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            arquivo_saida = f"processos_depre_{timestamp}.xlsx"
            
            # Criar planilha
            criar_planilha_simples(processos, arquivo_saida)
            
            print("\n" + "="*70)
            print("🎉 PLANILHA CRIADA COM SUCESSO!")
            print("="*70)
            
            # Preview
            print(f"\n📋 PREVIEW (primeiros 10):")
            print(f"\n   {'Nº Processo':<35} {'DEPRE':<35}")
            print(f"   {'-'*35} {'-'*35}")
            
            for i, proc in enumerate(processos[:10], 1):
                num_proc = proc['Nº Processo'] or 'N/A'
                depre = proc['DEPRE']
                print(f"   {num_proc:<35} {depre:<35}")
            
            if len(processos) > 10:
                print(f"\n   ... e mais {len(processos) - 10}")
            
            print(f"\n📁 Arquivo: {arquivo_saida}")
            print(f"📊 Colunas: Nº Processo | DEPRE")
        
    except FileNotFoundError:
        print(f"\n❌ PDF não encontrado: {arquivo_pdf}")
        print(f"   Certifique-se que o arquivo está na pasta atual")
    
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n✅ FIM!")
