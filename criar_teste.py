import openpyxl

# Criar planilha de teste
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Processos"

# Cabeçalho
ws['A1'] = "Número do Processo"

# 3 processos de teste
processos_teste = [
    "0051675-54.2023.8.26.0500",
    "0052802-27.2023.8.26.0500",
    "0053687-41.2023.8.26.0500"
]

for idx, proc in enumerate(processos_teste, 2):
    ws[f'A{idx}'] = proc

wb.save("processos_TESTE_3.xlsx")
print("✅ Planilha de teste criada: processos_TESTE_3.xlsx")
print(f"📋 {len(processos_teste)} processos incluídos")
